#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""i3dna-lint — I3DNA 目录体系的引用完整性检查器（只读旁挂，第一层引擎）。

权威永远在文件（xlsx 索引表 + 目录 + __结果.json），本工具只读、只报错，
不写任何数据。数据流单向：xlsx → 内存图 → 报告。

检查项：
  [绑定] 索引文件每行 = 一条名字绑定边：
         近端（外部引用列=*）→ 名称必须存在于本目录；
         远端（\\...\\ 相对锚）→ 展开后目录必须存在、名称必须存在于其中；
         机器绝对路径（D:\\ 等）→ ⚠ 警告（根变量未绑定，本机不可解析）。
  [取值] 目录文件类型 ∈ {0,1}；模型存储模式 ∈ {0,1}；
         独立存储(1) 时输出文件名不得为 *。
  [参数] 任务参数表每行：参数文件目录可解析；名称非 * 时文件必须存在。
  [对账] __结果.json 的输入/产物清单：文件存在、字节数相符、sha256 相符。

用法：python3 i3dna_lint.py <树根目录>   # 树根 = 含 _简化版 的目录，\\...\\ 锚于此
退出码：0 = 干净（可有警告）；1 = 有错误。
"""
import fnmatch
import glob
import hashlib
import json
import os
import re
import sys

openpyxl = None      # 仅 xlsx 需要（懒加载，md 族树零此依赖）


def _need_openpyxl():
    global openpyxl
    if openpyxl is None:
        try:
            import openpyxl as _o
        except ImportError:
            sys.exit("需要 openpyxl：pip install openpyxl（仅 xlsx）")
        openpyxl = _o
    return openpyxl

ERR, WARN, INFO = "✗", "⚠", "·"


class Report:
    def __init__(self):
        self.errors, self.warnings, self.infos = [], [], []
        self.edges = 0

    def err(self, where, msg):
        self.errors.append((where, msg))

    def warn(self, where, msg):
        self.warnings.append((where, msg))

    def info(self, where, msg):
        self.infos.append((where, msg))

    def dump(self):
        for tag, items in [(ERR, self.errors), (WARN, self.warnings),
                           (INFO, self.infos)]:
            for where, msg in items:
                print(f"  {tag} [{where}] {msg}")
        print(f"  统计：绑定边 {self.edges} 条，错误 {len(self.errors)}，"
              f"警告 {len(self.warnings)}，信息 {len(self.infos)}")
        return len(self.errors)


def sha256(path):
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def norm_cell(v):
    """单元格值 → 去空白字符串（None → ''）。"""
    return "" if v is None else str(v).strip()


def is_machine_path(p):
    return bool(re.match(r"^[A-Za-z]:[\\/]", p))


def resolve_anchor(p, root):
    r"""把 \...\a\b 形式的相对锚路径解析为 root 下的绝对路径。"""
    p = p.replace("\\", "/")
    p = re.sub(r"^/?\.\.\./", "", p)
    return os.path.join(root, *[seg for seg in p.split("/") if seg])


def is_coord_row(cells):
    """列坐标行判定：非空单元格恰为从 0 或 * 起的连续整数序列。

    博士样式：* 0 1 2 3 4 5；亦兼容变体：0 1 2 3 4 5 6。"""
    vals = [c for c in cells if c != ""]
    if len(vals) < 3:
        return False
    if vals[0] == "*":
        vals = vals[1:]
    return all(re.fullmatch(r"\d+", v) for v in vals) and \
        [int(v) for v in vals] == list(range(len(vals)))


def data_rows(ws):
    """跳过表头三行（列名行 / 约束提示行 / 列坐标行），产出数据行。

    判据：第一列【序号】是纯数字、且整行不是列坐标行。"""
    for row in ws.iter_rows(values_only=True):
        cells = [norm_cell(c) for c in row]
        first = cells[0] if cells else ""
        if re.fullmatch(r"\d+", first) and not is_coord_row(cells):
            yield cells


def headers_of(ws):
    for row in ws.iter_rows(values_only=True):
        return [norm_cell(c) for c in row]
    return []


# ── 索引文件（绑定表）──────────────────────────────────────────

def lint_index_sheet(ws, xlsx_path, root, rep):
    where = os.path.relpath(xlsx_path, root)
    cwd = os.path.dirname(xlsx_path)
    for row in data_rows(ws):
        row += [""] * (7 - len(row))
        seq, desc, name, ftype, smode, outname, extref = row[:7]
        rep.edges += 1
        tag = f"{where}#行{seq}({desc or name})"

        if ftype not in ("0", "1"):
            rep.err(tag, f"目录文件类型非法：{ftype!r}（应为 0=目录|1=参数文件）")
        if smode not in ("0", "1"):
            rep.err(tag, f"模型存储模式非法：{smode!r}（应为 0=集成|1=独立）")
        if smode == "1" and outname in ("", "*"):
            rep.err(tag, "独立存储但【独立存储输出文件名称】为 *")
        if not name or name == "*":
            rep.err(tag, "【目录-文件名称】为空或 *，绑定无目标名")
            continue

        if extref in ("", "*"):
            target = os.path.join(cwd, name)
            if not os.path.exists(target):
                rep.err(tag, f"近端绑定悬空：本目录不存在 {name!r}")
        elif is_machine_path(extref):
            rep.warn(tag, f"机器绝对路径（根变量未绑定，本机不可解析）：{extref}")
        else:
            tdir = resolve_anchor(extref, root)
            if not os.path.isdir(tdir):
                rep.err(tag, f"远端绑定悬空：目录不存在 {extref}")
            elif not os.path.exists(os.path.join(tdir, name)):
                rep.err(tag, f"远端绑定悬空：{extref} 内不存在 {name!r}")


# ── 任务参数表 ────────────────────────────────────────────────

def lint_param_sheet(ws, xlsx_path, root, rep):
    where = os.path.relpath(xlsx_path, root)
    heads = headers_of(ws)
    idx = {h: i for i, h in enumerate(heads)}
    c_dir, c_name = idx.get("【参数文件目录】"), idx.get("【参数文件名称】")
    if c_dir is None or c_name is None:
        rep.warn(where, "参数表缺少【参数文件目录】/【参数文件名称】列，跳过")
        return
    for row in data_rows(ws):
        seq = row[0]
        pdir = row[c_dir] if c_dir < len(row) else ""
        pname = row[c_name] if c_name < len(row) else ""
        rep.edges += 1
        tag = f"{where}#行{seq}"
        if not pdir or pdir == "*":
            continue
        if is_machine_path(pdir):
            rep.warn(tag, f"机器绝对路径（根变量未绑定）：{pdir}")
            continue
        tdir = resolve_anchor(pdir, root)
        if not os.path.isdir(tdir):
            rep.err(tag, f"参数目录不存在：{pdir}")
        elif pname not in ("", "*") and not os.path.exists(
                os.path.join(tdir, pname)):
            rep.err(tag, f"参数文件不存在：{pdir} 内无 {pname!r}")


# ── __结果.json 对账 ─────────────────────────────────────────

def _dir_manifest(dpath):
    """目录盘点单（§8.12，与引擎同款；lint 零依赖自含）：递归逐文件 sha，
    `__`/点前缀目录不入。"""
    out = {}
    for dp, dns, fns in os.walk(dpath):
        dns[:] = sorted(d for d in dns if not d.startswith((".", "__")))
        for fn in sorted(fns):
            p = os.path.join(dp, fn)
            out[os.path.relpath(p, dpath)] = sha256(p)
    return out


def _product_cover(root):
    """全树产物覆盖表 {名称: {sha,…}}——任何账的产物清单都是一次覆盖。
    用途：旧账产物 sha 漂移时，若**现字节**被某账覆盖（改弧案卷把改后
    文件再入账＝合法重审），出处链未断——对账只对真悬空（无任何账盖
    住现态）报错（103号 立域拆站引入改写旧产物文件的合法通道）。"""
    cover = {}
    for j in glob.glob(os.path.join(root, "**", "__结果.json"), recursive=True):
        try:
            data = json.load(open(j, encoding="utf-8"))
        except (json.JSONDecodeError, OSError, UnicodeDecodeError):
            continue
        for it in data.get("产物清单", []):
            if it.get("名称") and it.get("sha256"):
                cover.setdefault(it["名称"], set()).add(it["sha256"])
    return cover


def _message_type_names(root):
    """消息类型名全集（**/消息/*.md 文件名去扩展；排除无 frontmatter 的
    正文型说明）——「疑似漏吃」判据的种名表（工单107 B 档）。"""
    names = set()
    for typef in glob.glob(os.path.join(root, "**", "消息", "*.md"),
                           recursive=True):
        try:
            t = open(typef, encoding="utf-8").read()
        except OSError:
            continue
        if re.match(r"^---\s*\n.*?\n---", t, re.S):
            names.add(os.path.splitext(os.path.basename(typef))[0])
    return names


def _ticket_type(fn, msg_types):
    """文件名去 __ 段逐层剥后缀寻种（与引擎 _type_file 同款）——命中
    消息类型名即为单据；非单据（普通知识件）返 None。"""
    stem = os.path.splitext(os.path.basename(fn))[0]
    parts = stem.split("__")
    for i in range(len(parts), 0, -1):
        cand = "__".join(parts[:i])
        if cand in msg_types:
            return cand
    return None


def _recycle_cover(root):
    """回收盖住集（{(名称, sha256)}，8-20 用户实证撤域配套）：产物被回收
    弧案卷依法销账——文件缺席是**退役**不是悬空（旧 sha 在回收清单、git
    历史留尸可对勘），对账不得再报「清单文件不存在」。"""
    out = set()
    for jpath in glob.glob(os.path.join(root, "**", "__结果.json"),
                           recursive=True):
        try:
            data = json.load(open(jpath, encoding="utf-8"))
        except (json.JSONDecodeError, UnicodeDecodeError, OSError):
            continue
        for it in data.get("回收清单") or []:
            if isinstance(it, dict) and it.get("sha256"):
                out.add((it.get("名称"), it.get("sha256")))
    return out


def lint_result_json(jpath, base, root, rep, prod_cover=None, msg_types=None,
                     recycle_cover=None):
    where = os.path.relpath(jpath, root)
    try:
        data = json.load(open(jpath, encoding="utf-8"))
    except (json.JSONDecodeError, UnicodeDecodeError) as e:
        rep.err(where, f"JSON 解析失败：{e}")
        return
    for kind in ("输入清单", "产物清单"):
        for item in data.get(kind, []):
            name = item.get("名称", "")
            tag = f"{where}·{kind}·{name}"
            path = name if os.path.isabs(name) else os.path.join(base, name)
            if item.get("目录"):                    # 盘点单（§8.12）
                if os.path.isdir(path):
                    # 主题车道（形状定律 8-21·工单2）：账带消费清单＝引擎
                    # 侧一火一单——已消费单（按名对上）的缺席是依法核销，
                    # 不算目录漂移；在场其余单＝下轮燃料，不算漏吃。
                    consumed = {os.path.basename(it2.get("名称") or "")
                                for it2 in (data.get("消费清单") or [])}
                    expect = {k: v for k, v in (item.get("清单") or {}).items()
                              if os.path.basename(k) not in consumed}
                    if _dir_manifest(path) != expect:
                        rep.warn(tag, "目录已变（收入/移出/改动）"
                                      "——待重审,推进后自平")
                    # 收件箱消费执法 B 档（缺陷15/工单107）：收盘盘点里
                    # 的单据过了本火仍在场且 sha 未变 ＝ 疑似漏吃——「吃单
                    # 删除」是任务正文里的 prose 嘱咐非机制，漏删→单永远
                    # 在场→站点永远有活反复点火白烧钱。在场但 sha 变＝
                    # 新单顶替（静默）；不在场＝已消费（静默）；非单据
                    # 文件不涉。主题车道（账有消费清单）跳过：删除已机制
                    # 化（A 档引擎执行），在场的其余单是队列余额不是漏吃。
                    if not data.get("消费清单"):
                        for fn, sha in (item.get("清单") or {}).items():
                            if not _ticket_type(fn, msg_types or ()):
                                continue
                            cur = os.path.join(path, fn)
                            if os.path.isfile(cur) and sha256(cur) == sha:
                                rep.warn(tag, f"疑似漏吃：单据「{fn}」过了"
                                              "本火仍在场（收盘盘点 sha 未变）"
                                              "——消费删除未执行或未覆盖")
                else:
                    rep.warn(tag, "输入目录不在场")
                continue
            if not os.path.isfile(path):
                if (name, item.get("sha256")) in (recycle_cover or ()):
                    rep.info(tag, "产物已依法回收（回收弧案卷销账）——"
                                  "旧 sha 在回收清单、git 历史留尸")
                    continue
                path2 = os.path.join(os.path.dirname(jpath), os.path.basename(name))
                if os.path.isfile(path2):
                    path = path2
                elif item.get("可缺"):
                    continue   # 可缺消息:开→消费→收回是设计内生命周期,缺席=已收回
                else:
                    rep.err(tag, "清单文件不存在")
                    continue
            size, digest = os.path.getsize(path), sha256(path)
            # 输入漂移=待重审（正常在途,converge 会消化）→ 警告；
            # 产物漂移=记账后被手改（出处断链）→ 错误。事件账的两种读法别再混
            if item.get("状态"):
                continue      # 字段区(键级属主):多方合法更新,不作漂移判据
            soft = (kind == "输入清单") or item.get("可缺")
            covered = prod_cover is not None and digest in \
                (prod_cover.get(name) or ())
            if "字节" in item and item["字节"] != size and not soft \
                    and not covered:
                rep.err(tag, f"字节数不符：账面 {item['字节']} / 实际 {size}")
            if "sha256" in item and item["sha256"] != digest:
                if item.get("可缺"):
                    rep.warn(tag, "可缺消息已更新（在途新单待消费,推进后自平）")
                elif kind == "输入清单":
                    rep.warn(tag, "输入已变（相对上次点火）——待重审,推进后自平")
                elif covered:
                    rep.info(tag, "产物已被后账重审覆盖（改弧案卷合法改写）"
                                  "——现态出处链完整")
                else:
                    rep.err(tag, f"sha256 不符：账面 {item['sha256'][:12]}… / "
                                 f"实际 {digest[:12]}…（产物在记账后被改动）")


# ── 主流程 ────────────────────────────────────────────────────

def lint_doc_type_keys(root, rep):
    """消息类型键 schema 对照：方法弧上的使能条件所取的键，必须出现在
    同类 消息/<消息名>.md 的「键」表里——拼错的键从静默 fail-closed
    升级为对账红（有类型文件才查，无文件不强求）。"""
    import yaml

    def fm_of(path):
        try:
            t = open(path, encoding="utf-8").read()
            m = re.match(r"^---\s*\n(.*?)\n---", t, re.S)
            return yaml.safe_load(m.group(1)) if m else {}
        except Exception:
            return {}
    for taskmd in glob.glob(os.path.join(root, "**", "方法", "*", "任务.md"),
                            recursive=True):
        croot = os.path.dirname(os.path.dirname(os.path.dirname(taskmd)))
        fm = fm_of(taskmd) or {}
        for item in (fm.get("输入") or []) + (fm.get("产物") or []):
            if not isinstance(item, dict) or not item.get("使能条件"):
                continue
            if not isinstance(item["使能条件"], dict):
                rep.err(os.path.relpath(taskmd, root),
                        f"使能条件不是映射：{item['使能条件']!r}——"
                        "文法是 {取值: 键, 算子: 字面量}（标量会让引擎静默不使能）")
                continue
            key = (item["使能条件"] or {}).get("取值")
            doc = os.path.basename(item.get("路径", ""))
            tname = os.path.splitext(doc)[0] + ".md"
            typef = next((c for c in (
                os.path.join(croot, "消息", tname),
                os.path.join(croot, "状态", tname),
                os.path.join(root, "消息", tname),
                os.path.join(root, "状态", tname)) if os.path.isfile(c)), None)
            if typef is None:
                continue
            keys = (fm_of(typef) or {}).get("键") or []
            if key not in keys:
                rep.err(os.path.relpath(taskmd, root),
                        f"使能条件取值键「{key}」不在类型 "
                        f"{os.path.relpath(typef, root)} 的键表 {keys} 里"
                        "（拼错的键会静默不使能）")
    # 消息路径对照：同一消息散写在多条弧上，谁改了一条消息就静默断流
    # （发的写一处、收的等另一处）——法定路径在类型文件，弧逐条对照
    for typef in (glob.glob(os.path.join(root, "**", "消息", "*.md"),
                            recursive=True)
                  + glob.glob(os.path.join(root, "**", "状态", "*.md"),
                              recursive=True)):
        law = fm_of(typef).get("路径")
        if not law:
            continue
        doc = os.path.basename(typef)
        croot = os.path.dirname(os.path.dirname(typef))
        enterprise = os.path.abspath(croot) == os.path.abspath(root)
        scope = (os.path.join(root, "**", "方法", "*", "任务.md")
                 if enterprise
                 else os.path.join(croot, "方法", "*", "任务.md"))
        for taskmd in glob.glob(scope, recursive=True):
            # 遮蔽链:企业级法不辖已在本类根立同名类型的方法
            # (类根先于企业根,99号父子context语义)
            tcroot = os.path.dirname(os.path.dirname(
                os.path.dirname(taskmd)))
            if enterprise and (
                    os.path.isfile(os.path.join(tcroot, "状态", doc)) or
                    os.path.isfile(os.path.join(tcroot, "消息", doc))):
                continue
            fm = fm_of(taskmd)
            krn = os.path.basename(tcroot) \
                if os.path.abspath(tcroot) != os.path.abspath(root) else ""

            def _norm(p, _krn=krn):
                # 类感知归一:{实例}按 taskmd 自己的类根展开,
                # {案卷号}→占位——开发的 {实例}/测试/审查单.md 与
                # 根法 实例/研发/{案卷号}/… 判同指(共享库所协同)
                p = p.replace("{实例}",
                              f"实例/{_krn}/«CASE»" if _krn else "实例/«CASE»")
                return p.replace("{案卷号}", "«CASE»")

            for item in (fm.get("输入") or []) + (fm.get("产物") or []):
                rel = item.get("路径") if isinstance(item, dict) else item
                if not isinstance(rel, str) or \
                        os.path.splitext(os.path.basename(rel))[0] \
                        != os.path.splitext(doc)[0]:
                    continue
                if _norm(rel) != _norm(law):
                    rep.err(os.path.relpath(taskmd, root),
                            f"消息「{doc}」弧路径 {rel} ≠ 法定路径 {law}"
                            f"（消息类型 消息/{doc}）——发收两侧会静默断流")


def lint_seq_type_arcs(root, rep):
    """动态名单×文件弧错配（104修订2/工单106＋形状定律 8-21·工单2 共用门）：
    声明 顺号: 真（__rN 量子名）或 命名: uuid（引擎代起名）的消息类型，
    单名对消费者不可预期——若消费弧是文件弧（按固定名读单），消费者永远
    读不到新单（审查单/持有单炸点）。凡输入弧的文件名（去 __ 段）命中此类
    型名 → 警告应改目录弧（队列消费）；目录弧不匹配类型名，静默。"""
    import yaml

    def fm_of(path):
        try:
            t = open(path, encoding="utf-8").read()
            m = re.match(r"^---\s*\n(.*?)\n---", t, re.S)
            return yaml.safe_load(m.group(1)) if m else {}
        except Exception:
            return {}

    seq_names = {}
    for typef in glob.glob(os.path.join(root, "**", "消息", "*.md"),
                           recursive=True):
        fm = fm_of(typef) or {}
        attr = "顺号" if str(fm.get("顺号")) in ("真", "True", "true") else (
            "uuid" if str(fm.get("命名") or "").strip().lower() == "uuid"
            else None)
        if attr:
            stem = os.path.splitext(os.path.basename(typef))[0]
            seq_names[stem] = (attr, typef)
    if not seq_names:
        return
    for taskmd in glob.glob(os.path.join(root, "**", "方法", "*", "任务.md"),
                            recursive=True):
        fm = fm_of(taskmd) or {}
        for item in (fm.get("输入") or []):
            rel = item.get("路径") if isinstance(item, dict) else item
            if not isinstance(rel, str) or not rel:
                continue
            stem, ext = os.path.splitext(os.path.basename(rel))
            if not ext:
                continue          # 目录弧（队列消费）——动态名单的正当消费方式
            parts = stem.split("__")
            names = {stem} | {"__".join(parts[:i])
                              for i in range(len(parts) - 1, 0, -1)}
            hit = next((n for n in names if n in seq_names), None)
            if hit:
                attr, typef = seq_names[hit]
                rep.warn(os.path.relpath(taskmd, root),
                         f"{attr}类型「{hit}」的消费弧是文件弧——应改目录弧，"
                         f"否则 __rN/uuid 单错配"
                         f"（类型 {os.path.relpath(typef, root)}）")


def lint_theme_types(root, rep):
    """形状定律主题判型对账（8-21·工单1号）：消息类型文件声明「主题:」＝
    目录即类型——该目录下一切非 __/点 前缀文件皆此类型（判型先目录后
    文件名，引擎 _type_file ①级同判据、独立实现——独立见证人惯例）。
    对账两条：①主题模式悬空（全树无一目录命中）→ 警告——立法指空处，
    单据无处落位；②主题目录里的单据 frontmatter 须键齐（按类型「键:」
    清单）——目录即类型，乱名也是单，缺键＝单不合规。"""
    import yaml

    def fm_of(path):
        try:
            t = open(path, encoding="utf-8").read()
            m = re.match(r"^---\s*\n(.*?)\n---", t, re.S)
            return yaml.safe_load(m.group(1)) if m else {}
        except Exception:
            return {}

    laws = []
    for typef in sorted(glob.glob(os.path.join(root, "**", "消息", "*.md"),
                                  recursive=True)):
        law = (fm_of(typef) or {}).get("主题")
        if isinstance(law, str) and law.strip():
            laws.append((law.strip(), typef))
    if not laws:
        return                    # 零主题声明＝检查休眠（逐字节兼容律）
    all_dirs = [os.path.relpath(d, root).replace(os.sep, "/")
                for d in glob.glob(os.path.join(root, "**", "*"), recursive=True)
                if os.path.isdir(d)]
    for law, typef in laws:
        pat = (law.replace("{案卷号}", "*").replace("{实例}", "*")
                  .replace("{类}", "*"))
        hits = [d for d in sorted(set(all_dirs)) if fnmatch.fnmatch(d, pat)]
        if not hits:
            rep.warn(os.path.relpath(typef, root),
                     f"主题悬空：类型「{os.path.basename(typef)}」模式"
                     f"「{law}」全树无一目录命中——立法指空处，单据无处落位")
            continue
        keys = (fm_of(typef) or {}).get("键") or []
        for d in hits:
            dabs = os.path.join(root, *d.split("/"))
            for fn in sorted(os.listdir(dabs)):
                if fn.startswith(("__", ".")) \
                        or not os.path.isfile(os.path.join(dabs, fn)):
                    continue        # 账不进账/隐藏件豁免（§8.12 同源）
                fm = fm_of(os.path.join(dabs, fn)) or {}
                miss = [k for k in keys if k not in fm]
                if miss:
                    rep.warn(f"{d}/{fn}",
                             f"主题单据「{fn}」缺键 {miss}——目录即类型"
                             f"（{os.path.relpath(typef, root)} 的键表），"
                             "乱名也是单，单要合规")


def lint_table_types(root, rep):
    """表声明对账（形状定律 8-21·工单3：目录即表——费米子聚合升格立法）：
    实体类 类.md frontmatter 声明 表: "实例/人员" → 该目录＝表、子目录＝行
    （Pauli：一路径一份）、行内键＝列（现行惯例一实体一目录由此立法化）。
    声明住法典（类.md＋其指名的 schema.md），不住目录里。对账三条：
    ①表悬空（目录不存在）→ 警告；②行缺列——schema「键说明」要求的键，
    行主文件 frontmatter 或同槽文件（档案袋形：实例/人员/刘亦菲/工作单位.md）
    二形其一在场即可 → 双缺席警告；③一类一表（Pauli）：两类声明同表 →
    错误。引擎零行为——纯读侧对账（为血缘/查询铺路，行为按真实用例再立）。"""
    import yaml

    def fm_of(path):
        try:
            t = open(path, encoding="utf-8").read()
            m = re.match(r"^---\s*\n(.*?)\n---", t, re.S)
            return yaml.safe_load(m.group(1)) if m else {}
        except Exception:
            return {}

    # 类根＝**/类/* 下带 类.md 的目录（与 lint_logical_model 判据同源）
    claims, keys_of = {}, {}
    for cmd_ in sorted(glob.glob(os.path.join(root, "**", "类", "*", "类.md"),
                                 recursive=True)):
        c = os.path.dirname(cmd_)
        fm = fm_of(cmd_) or {}
        table = fm.get("表")
        if not isinstance(table, str) or not table.strip():
            continue
        claims.setdefault(table.strip(), []).append(c)
        # 列清单＝schema.md「键说明」（类.md schema 指针解析；无 schema/
        # 无键说明＝零列要求，只对账悬空与撞表）
        keys = []
        sp = fm.get("schema")
        spath = os.path.join(root, str(sp)) if isinstance(sp, str) and sp else \
            os.path.join(c, "schema.md")
        if os.path.isfile(spath):
            ks = (fm_of(spath) or {}).get("键说明") or []
            keys = list(ks.keys()) if isinstance(ks, dict) else \
                [str(k) for k in ks]
        keys_of[c] = keys
    if not claims:
        return                    # 零表声明＝检查休眠（逐字节兼容律）
    for table, cs in sorted(claims.items()):
        tdir = os.path.join(root, *table.split("/"))
        if len(cs) > 1:
            rep.err(table, f"一表一类（Pauli）被破：{table} 同时被 "
                    + "、".join(os.path.relpath(c, root) for c in cs)
                    + " 声明——一路径一份，行归属只能一个类")
        if not os.path.isdir(tdir):
            rep.warn(os.path.relpath(cs[0], root),
                     f"表悬空：{table} 目录不存在——立法指空处，行无处落位")
            continue
        for row in sorted(os.listdir(tdir)):
            rdir = os.path.join(tdir, row)
            if row.startswith((".", "__")) or not os.path.isdir(rdir):
                continue          # 账不进账/隐藏件豁免（§8.12 同源）
            main_fm = fm_of(os.path.join(rdir, f"{row}.md")) or {}
            slots = {os.path.splitext(f)[0] for f in os.listdir(rdir)
                     if f.endswith(".md")}
            miss = [k for k in keys_of.get(cs[0], [])
                    if k not in main_fm and k not in slots]
            if miss:
                rep.warn(f"{table}/{row}",
                         f"表行「{row}」缺列 {miss}——目录即表（类 "
                         f"{os.path.relpath(cs[0], root)} schema 键说明），"
                         "主文件键值区或同槽文件二形其一须在场")


def lint_echo_types(root, rep):
    """回音双向门（形状定律 8-21·工单4：绿任务 send-and-wait / fire-and-
    forget 立法，要式性互咬）：消息类型声明 回音: 有＝要回音——全树无人
    挂指向其主题夹的空夹门（清空: 真）＝**空承诺**（声明了回音却没人等
    它），错误；回音: 无＝收讫——有人挂空夹门等它＝**死等**（等一个不会
    来的回音），错误。门弧与主题法两侧同规一化（{案卷号}/{实例}/{类} 记
    号）后 fnmatch 对上即认。无 回音 声明或无主题夹者不进门。"""
    import yaml

    def fm_of(path):
        try:
            t = open(path, encoding="utf-8").read()
            m = re.match(r"^---\s*\n(.*?)\n---", t, re.S)
            return yaml.safe_load(m.group(1)) if m else {}
        except Exception:
            return {}

    def _n(p):
        return (p.replace("{实例}", "实例/«C»").replace("{案卷号}", "«C»")
                 .replace("{类}", "«K»"))

    laws = []                      # [(类型文件, 回音值, 主题模式)]
    for typef in sorted(glob.glob(os.path.join(root, "**", "消息", "*.md"),
                                  recursive=True)):
        fm = fm_of(typef) or {}
        echo = str(fm.get("回音") or "").strip()
        law = fm.get("主题")
        if echo not in ("有", "无") or not isinstance(law, str) or not law.strip():
            continue
        laws.append((typef, echo, law.strip()))
    if not laws:
        return                    # 零回音声明＝检查休眠（逐字节兼容律）
    gates = []                     # [(任务md, 门弧路径)]
    for taskmd in sorted(glob.glob(os.path.join(root, "**", "方法", "*",
                                                "任务.md"), recursive=True)):
        for item in (fm_of(taskmd) or {}).get("输入") or []:
            if isinstance(item, dict) \
                    and str(item.get("清空") or "").strip() in ("真", "True",
                                                                "true") \
                    and isinstance(item.get("路径"), str):
                gates.append((taskmd, item["路径"]))
    for typef, echo, law in laws:
        hit = next((t for t, gp in gates
                    if fnmatch.fnmatch(_n(gp), _n(law))
                    or fnmatch.fnmatch(_n(law), _n(gp))), None)
        if echo == "有" and hit is None:
            rep.err(os.path.relpath(typef, root),
                    f"空承诺：类型「{os.path.basename(typef)}」声明 回音: 有，"
                    f"全树无人挂指向其主题夹（{law}）的空夹门（清空: 真）"
                    "——声明了回音却没人等它")
        if echo == "无" and hit is not None:
            rep.err(os.path.relpath(typef, root),
                    f"死等：类型「{os.path.basename(typef)}」声明 回音: 无"
                    f"（收讫），却有人挂空夹门等它（{os.path.relpath(hit, root)}"
                    f" 指向 {law}）——等一个不会来的回音")


def lint_owner_table(root, rep):
    """字段区属主表死指针：属主须是「人」或现存方法（类级查本类，
    企业级查任一类）。蓝任务也可以当属主——它有定点笔（i3dna_kv 命令行），
    行为层另有引擎的越权改键守卫；lint 只查声明指得到人。"""
    import yaml

    for typef in glob.glob(os.path.join(root, "**", "状态", "*.md"),
                           recursive=True):
        try:
            t = open(typef, encoding="utf-8").read()
            m = re.match(r"^---\s*\n(.*?)\n---", t, re.S)
            fm = yaml.safe_load(m.group(1)) if m else {}
        except Exception:
            continue
        owners = (fm or {}).get("属主")
        if not isinstance(owners, dict):
            continue
        croot = os.path.dirname(os.path.dirname(typef))
        scopes = ([croot] if os.path.abspath(croot) != os.path.abspath(root)
                  else (glob.glob(os.path.join(root, "类", "*")) or [root]))
        methods = {os.path.basename(d) for sc in scopes
                   for d in glob.glob(os.path.join(sc, "方法", "*"))
                   if os.path.isdir(d)}
        for k, owner in owners.items():
            if str(owner) == "人" or str(owner) in methods:
                continue
            rep.err(os.path.relpath(typef, root),
                    f"属主表死指针：键「{k}」的属主「{owner}」既不是「人」"
                    f"也不是现存方法（现有：{'、'.join(sorted(methods)) or '无'}）")



def lint_entity_edges(root, rep):
    """实体类关系边(图网的线)检查——91 号图网闭合三件套第三件。
    与 i3dna_model.边 同一语义、独立实现:检查器与被检查方不得共享实现
    (独立见证人原则,同 sha256 对账的读法)。边:类型=关系名(自由词);
    种∈{继承,关联,聚合,组合}(缺省 关联);方向 A → B / A ← B 解析目标类;
    目标类必须全树存在,否则悬空引用。"""
    import yaml

    def fm_of(path):
        try:
            t = open(path, encoding="utf-8").read()
            m = re.match(r"^---\s*\n(.*?)\n---", t, re.S)
            return yaml.safe_load(m.group(1)) if m else {}
        except Exception:
            return {}

    classes = {os.path.basename(d) for d in
               glob.glob(os.path.join(root, "类", "*"))
               + glob.glob(os.path.join(root, "域", "*", "类", "*"))}
    整体对 = set()
    for clsmd in sorted(glob.glob(os.path.join(root, "**", "类.md"),
                                   recursive=True)):
        fm = fm_of(clsmd) or {}
        for item in (fm.get("关系") or []):
            tag = os.path.relpath(clsmd, root)
            if not isinstance(item, dict):
                rep.err(tag, f"关系声明不是映射：{item!r}——文法是 "
                             "{类型: 名, 方向: A → B, 种: 四型之一}")
                continue
            kind = item.get("种") or item.get("类别") or "关联"
            if kind not in ("继承", "关联", "聚合", "组合"):
                rep.err(tag, f"边「{item.get('类型')}」的种非法：{kind!r}"
                             "（应为 继承|关联|聚合|组合 四型之一）")
            d = (item.get("方向") or "").strip()
            parts = [p.strip() for p in re.split(r"[→←]", d) if p.strip()]
            owner = os.path.basename(os.path.dirname(clsmd))
            target = None
            if len(parts) >= 2:
                a, b = parts[0], parts[-1]
                target = b if a == owner else (a if b == owner
                                               else (b if "→" in d else a))
            elif parts:
                target = parts[0]
            if not target:
                rep.err(tag, f"边「{item.get('类型')}」方向无法解析：{d!r}")
                continue
            if target == owner:
                rep.warn(tag, f"边「{item.get('类型')}」方向两端同名：{d!r}")
                continue
            if target not in classes:
                rep.err(tag, f"悬空引用：边「{item.get('类型')}」指向 "
                             f"「{target}」，全树无此类（{d}）")
            if kind in ("聚合", "组合"):
                整体对.add((owner, target))
    # 整体-部分成环守卫（级联展开的 visited 由 lint 提前把关）
    for a, b in sorted(整体对):
        if (b, a) in 整体对:
            rep.err(f"{a}◇{b}", f"整体-部分成环：{a}◇{b} 且 {b}◇{a}"
                    "——UML 聚合/组合是无环的整体-部分层级，环是建模错误")



def lint_domain_context(root, rep):
    """进化边界律(93 号)的执法面:域依赖图推导 + 跨域写凭证检查。
    域管语言,case 管事务;跨域投影=泛衍,必须经凭证——域级白名单(写)或
    字段级属主表(目标键属主=本类名)。无凭证的跨域写 → 黄牌(87 §5)。
    本检查只读推导,不存储:context map 随树演化自动更新。"""
    import yaml

    def fm_of(path):
        """双形声明头:标准 frontmatter(--- 开头) / kv 风格(域.md:键值行直起)。
        与 i3dna_model._fm 同口径(独立实现)。"""
        try:
            t = open(path, encoding="utf-8").read()
            m = re.match(r"^---\s*\n(.*?)\n---", t, re.S)
            if m:
                return yaml.safe_load(m.group(1)) or {}
            out = {}
            for ln in t.splitlines():
                s = ln.strip()
                if not s:
                    continue
                if ":" not in s:
                    break
                k, _, v = s.partition(":")
                out[k.strip()] = v.strip() or None
            return out
        except Exception:
            return {}

    # 类 → 域;域 → 白名单(域名列表,写权限)
    class_dom, dom_allow = {}, {}
    for clsd in glob.glob(os.path.join(root, "域", "*", "类", "*")):
        dom = os.path.basename(os.path.dirname(os.path.dirname(clsd)))
        class_dom[os.path.basename(clsd)] = dom
    for dommd in glob.glob(os.path.join(root, "域", "*", "域.md")):
        dom = os.path.basename(os.path.dirname(dommd))
        wl = (fm_of(dommd) or {}).get("跨域白名单") or []
        if isinstance(wl, str):
            wl = [x.strip() for x in wl.strip("[]").split(",") if x.strip()]
        dom_allow[dom] = {}
        for item in wl or []:
            if isinstance(item, str) and ":" in item:
                d, _, perm = item.partition(":")
                dom_allow[dom][d.strip()] = perm.strip()
    if not class_dom:
        return      # 无域层树(平铺形)不查边界律

    def 实体属主(entity, key):
        """目标键的属主表(字段级凭证)。"""
        for sch in glob.glob(os.path.join(root, "域", "*", "类", entity,
                                          "schema.md")):
            owners = (fm_of(sch) or {}).get("属主") or {}
            return owners.get(key) if isinstance(owners, dict) else None
        return None

    for taskmd in glob.glob(os.path.join(root, "域", "*", "类", "*", "方法",
                                         "*", "任务.md"), recursive=False):
        rel_parts = os.path.relpath(taskmd, root).split(os.sep)
        if len(rel_parts) < 5:
            continue
        dom, klass = rel_parts[1], rel_parts[3]
        fm = fm_of(taskmd)
        writes = (fm.get("产物") or []) + (fm.get("投影") or [])
        for item in writes:
            if not isinstance(item, dict):
                continue
            raw = item.get("路径") or item.get("目标") or ""
            raw = raw.replace("{实例}", "实例").replace("{案卷号}", "X")
            m = re.match(r"^实例/([^/]+)/", raw)
            if not m:
                continue
            target = m.group(1)
            tdom = class_dom.get(target)
            if tdom is None or tdom == dom:
                continue     # 本域写 / 非类目标(如案卷内文件)
            # 跨域写:查凭证
            perm = (dom_allow.get(tdom) or {}).get(dom, "")
            字段凭 = None
            key = item.get("键")
            if key is not None:
                字段凭 = 实体属主(target, key)
            where = os.path.relpath(taskmd, root)
            if "写" in perm or 字段凭 == klass:
                rep.info(where, f"跨域边 {dom}→{tdom} "
                         f"({target},凭证={'白名单' if '写' in perm else f'属主:{字段凭}'})")
            else:
                rep.warn(where, f"跨域写无凭证：{dom}→{tdom}（{target}，"
                         f"键 {key}）——域白名单无本域写权、目标键属主"
                         f"也不是本类（{klass}）；进化边界律黄牌")



def lint_lineage(root, rep):
    """血缘文件检查(94号):__血缘.md 逐行 键 :: 值哈希 :: 来源 :: 时间。
    判据:行格式四段、时间可解析、append-only(同键时间非降)。
    (键∈属主表的对照待字段区类型文件解析完成后补,见 94 号悬置。)"""
    from datetime import datetime
    for lf in sorted(glob.glob(os.path.join(root, "**", "__血缘.md"),
                                recursive=True)):
        rel = os.path.relpath(lf, root)
        prev = {}
        try:
            lines = open(lf, encoding="utf-8").read().splitlines()
        except UnicodeDecodeError as e:
            rep.err(rel, f"血缘文件读不了：{e}")
            continue
        for ln in lines:
            if not ln.strip():
                continue
            parts = [p.strip() for p in ln.split("::")]
            if len(parts) != 4:
                rep.err(rel, f"血缘行格式错(应为 键 :: 值哈希 :: 来源 :: 时间)："
                             f"{ln[:60]}")
                continue
            k, h, src, ts = parts
            try:
                datetime.fromisoformat(ts)
            except ValueError:
                rep.err(rel, f"血缘时间不可解析：{ts!r}（行 {ln[:60]}）")
                continue
            if src.count("/") != 1:
                rep.warn(rel, f"血缘来源形状可疑(应为 案卷号/方法)：{src!r}")
            if prev.get(k, "") > ts:
                rep.err(rel, f"血缘时间回退(append-only 违反)：键 {k} 时间 "
                             f"{ts} < {prev[k]}")
            prev[k] = ts



def lint_case_closure(root, rep, case_dirs=None):
    """案卷结账检查（守恒律,93号 case管事务的执法面）:
    结账=事务 commit,检查点在案卷不在全树——全局静止在开放系统几乎
    从不发生,从不成立的检查点等于没有。对每个案卷:
      (a) 路径落在案卷槽位内的消息(如 实例/研发/X/测试/返工单.md)
      (b) 文件名含该案卷号的消息(如 收件箱/请求审查单__X.md)
    在场即悬账——该案卷开出的承诺未兑现/应收的回执未销账。
    无绿任务豁免:结账的案卷不该还有待人办单(没办完就结账=
    提前结账,本身就是错——待人办与结账互斥,特例自然消解)。
    E=mc² 语义:单⇄档换算每次点火守恒,悬账=换算断链;
    转账例:扣款消息在场而入账回执缺席,守恒破坏可指认。"""
    import glob as _g
    import re as _re
    import yaml as _yaml

    def _fm(path):
        try:
            t = open(path, encoding="utf-8").read()
        except OSError:
            return {}
        m = _re.match(r"^---\s*\n(.*?)\n---", t, _re.S)
        return _yaml.safe_load(m.group(1)) or {} if m else {}

    # 全部消息类型的法定路径(通配展开;同名法去重)
    laws = {}
    for typef in _g.glob(os.path.join(root, "**", "消息", "*.md"),
                         recursive=True):
        law = _fm(typef).get("路径")
        if law:
            name = os.path.splitext(os.path.basename(typef))[0]
            laws.setdefault(name,
                            law.replace("{案卷号}", "*")
                               .replace("{实例}", "*"))
    if not laws:
        return
    # 案卷目录枚举:实例/<类>/<case>——类架名必须在树上真有对应过程类
    # 根(收件箱等共享槽不是案卷)。缺省全树,结账时传单个。
    def _class_names():
        names = set()
        for t in _g.glob(os.path.join(root, "**", "任务.md"), recursive=True):
            d = os.path.dirname(os.path.dirname(
                os.path.dirname(t)))       # …/方法/T → …/类/类名
            names.add(os.path.basename(d))
        return names

    # 案卷目录枚举:实例/<类>/<case>——类架名须真有过程类根,且不能
    # 本身是消息法定路径的信箱容器(收件箱是共享槽不是案卷)
    klasses = _class_names()
    box_dirs = {os.path.dirname(os.path.join(root, p))
                for p in laws.values()}
    if case_dirs is None:
        case_dirs = [
            d for d in _g.glob(os.path.join(root, "实例", "*", "*"))
            if os.path.isdir(d)
            and os.path.basename(os.path.dirname(d)) in klasses
            and os.path.realpath(d) not in map(os.path.realpath, box_dirs)
            and not os.path.basename(d).startswith((".", "__"))]
    for cdir in case_dirs:
        case = os.path.basename(cdir)
        crel = os.path.relpath(cdir, root)
        for name, pat in laws.items():
            for h in _g.glob(os.path.join(root, pat)):
                hrel = os.path.relpath(h, root)
                # 归属判据:落案卷槽位内,或文件名按 __<案卷号> 段匹配
                stem = os.path.splitext(os.path.basename(h))[0]
                in_case = hrel.startswith(crel + os.sep) \
                    or any(seg == case
                           for seg in stem.split("__"))
                if in_case:
                    rep.err(hrel,
                            f"案卷悬账:消息「{name}」在结账案卷 {case} "
                            "在场——开出的承诺未兑现或回执未销账"
                            "（守恒破坏;先消费/销单或撤销结账）")

def lint_logical_model(root, rep):
    """逻辑模型合法性总检——i3dna_model 是树的 schema,本函数是 validator:
    树必须长成十概念的样子(ARCHITECTURE.md §5)。独立实现纪律:不用
    i3dna_model(检查器与被检查方不共享实现,同 sha256 对账的读法)。
    判据(逐概念):
    ①树:唯一根,根下只许 已知顶层目录(实例/类/域/消息/状态/知识/部门…);
    ②域:域下须有 类/ 或域.md(空域=悬空 package);
    ③类:根/类/* 与 域/*/类/* 的每个类根须可判类种(有 方法/=过程类;
      类.md 范畴=实体 或 schema.md=实体类);不可判=报错;
    ④消息类:消息/*.md 须有 路径(法定路径)与发送方/接收方——
      无路径的消息类型使存在性判据失效;
    ⑤实例:实例/<类>/<k> 的类架名须指向真类根(过程类)——
      例外白名单:收件箱(共享槽)、部门(档案袋架);
    ⑥场所:装配清单(场所/<名>.md 装配: [...],N:M 一场所装配多域)里的
      域名须真在 域/ 下;缺装配清单=警告(空场所不入拓扑)。"""
    import yaml as _yaml

    def _fm(path):
        try:
            t = open(path, encoding="utf-8").read()
        except OSError:
            return {}
        parts = t.split("---")
        if len(parts) >= 3 and not parts[0].strip():
            try:
                return _yaml.safe_load(parts[1]) or {}
            except _yaml.YAMLError:
                return {}
        return {}

    # ① 顶层目录白名单
    KNOWN_TOP = {"实例", "类", "域", "场所", "消息", "状态", "知识", "部门",
                 "__日志", "__账", "docs", ".git", ".obsidian"}
    for x in os.listdir(root):
        full = os.path.join(root, x)
        if os.path.isdir(full) and x not in KNOWN_TOP \
                and not x.startswith(("_", ".")):
            rep.warn(os.path.relpath(full, root),
                     f"顶层目录「{x}」不在逻辑模型已知目录集——"
                     "要么是未建模概念,要么放错位置")

    # ② 空域检查
    域根 = os.path.join(root, "域")
    if os.path.isdir(域根):
        for d in os.listdir(域根):
            dp = os.path.join(域根, d)
            if not os.path.isdir(dp):
                continue
            has_classes = os.path.isdir(os.path.join(dp, "类"))
            has_decl = os.path.isfile(os.path.join(dp, "域.md"))
            if not has_classes and not has_decl:
                rep.err(os.path.relpath(dp, root),
                        f"空域「{d}」:无 类/ 无 域.md——悬空 package")

    # ③ 类根可判类种
    def _类根们():
        out = []
        for base in (os.path.join(root, "类"), 域根):
            if not os.path.isdir(base):
                continue
            for sub in os.listdir(base):
                sp = os.path.join(base, sub)
                if not os.path.isdir(sp):
                    continue
                if base is 域根 or base.endswith("域"):
                    c = os.path.join(sp, "类")
                    if os.path.isdir(c):
                        out += [os.path.join(c, x)
                                for x in os.listdir(c)
                                if os.path.isdir(os.path.join(c, x))]
                else:
                    out.append(sp)
        return out

    def _是类根(c):
        if os.path.isdir(os.path.join(c, "方法")):
            return "过程"
        p = os.path.join(c, "类.md")
        if os.path.isfile(p) and _fm(p).get("范畴") == "实体":
            return "实体"
        if os.path.isfile(os.path.join(c, "schema.md")):
            return "实体"
        return None

    过程类名集, 实体类名集 = set(), set()
    for c in _类根们():
        种 = _是类根(c)
        if 种 == "实体":
            实体类名集.add(os.path.basename(c))
        if 种 is None:
            rep.err(os.path.relpath(c, root),
                    f"不可判类根:无 方法/、类.md 范畴≠实体、无 schema.md"
                    "——逻辑模型要求每个类根可判类种(过程类或实体类)")
        elif 种 == "过程":
            过程类名集.add(os.path.basename(c))

    # ④ 消息类型法定路径（主题即法定路径——形状定律 8-21·工单5：声明
    # 主题: 的类型，**主题目录就是存在性判据**（夹内有单＝消息在场；票据
    # 夹的成员清单即 marking），文件槽老法 路径: 不辖主题类型——两法并挂
    # 反而双记：路径悬账门无消费方豁免，要吃单的办结被没吃的单拦死
    # （send-and-wait self-deadlock，报销003 靶场实证）。引擎零改动：
    # 悬账门主题车道自带消费方豁免（工单2）。）
    for typef in glob.glob(os.path.join(root, "**", "消息", "*.md"),
                           recursive=True):
        fm = _fm(typef)
        if not fm:
            continue          # 正文型说明(如 类/复盘/消息/说明.md)跳过
        if fm.get("主题"):
            continue          # 主题类型：法定路径住目录，不查 路径: 键
        if not fm.get("路径"):
            rep.err(os.path.relpath(typef, root),
                    "消息类型缺「路径」(法定路径)——存在性判据失效")
        elif not (fm.get("发送方") and fm.get("接收方")):
            rep.warn(os.path.relpath(typef, root),
                     "消息类型缺 发送方/接收方——跨岗流向不可判")

    # ⑤ 实例架指向真类根(过程类=案卷架,实体类=档案袋架;白名单:收件箱/部门)
    库 = os.path.join(root, "实例")
    if os.path.isdir(库):
        for 架 in os.listdir(库):
            ap = os.path.join(库, 架)
            if not os.path.isdir(ap) or 架.startswith((".", "__")):
                continue
            if 架 in ("收件箱", "部门"):
                continue
            if 架 not in 过程类名集 and 架 not in 实体类名集:
                rep.err(os.path.relpath(ap, root),
                        f"实例架「{架}」不指向任何类根(过程类或实体类)"
                        "——实例/<类>/<k> 的类须真实在树(消息类不实例化;"
                        "白名单:收件箱/部门)")

    # ⑥ 场所装配清单(声明制,N:M):装配的域名须真在 域/ 下;认领互斥
    #   (费米律:一域至多被一份声明认领——与树内校验程序同判,通用
    #   检查器侧的执法面,8-19 审计补)
    场所根 = os.path.join(root, "场所")
    if os.path.isdir(场所根):
        真域 = ({d for d in os.listdir(域根)
                 if os.path.isdir(os.path.join(域根, d))}
                if os.path.isdir(域根) else set())
        认领 = {}
        for f in sorted(os.listdir(场所根)):
            fp = os.path.join(场所根, f)
            if not f.endswith(".md") or not os.path.isfile(fp):
                continue
            装配 = _fm(fp).get("装配")
            if not 装配:
                rep.warn(os.path.relpath(fp, root),
                         "场所声明缺「装配」清单——空场所不入拓扑")
                continue
            doms = [str(x) for x in 装配]
            for d in doms:
                if d != "全部" and d not in 真域:
                    rep.err(os.path.relpath(fp, root),
                            f"场所装配的域「{d}」不在树(域/ 下无此域)"
                            "——装配清单须指向真域")
            展认领 = 真域 if "全部" in doms else set(doms) - {"全部"}
            for d in sorted(展认领):
                if d in 认领:
                    rep.err(os.path.relpath(fp, root),
                            f"费米律破坏：域「{d}」被 {认领[d]} 与 {f} "
                            "两份声明认领——一域至多一份声明")
                else:
                    认领[d] = f


def _parse_baseline(path, root):
    """宪法时刻基线表 → {路径: sha}；无可解析行返回 None。"""
    out = {}
    for line in open(path, encoding="utf-8", errors="replace").read().splitlines():
        s = line.strip()
        if not s.startswith("|") or set(s) <= {"|", "-", " ", ":"}:
            continue
        cells = [c.strip().strip("*").strip() for c in s.strip("|").split("|")]
        if len(cells) >= 2 and cells[0] not in ("路径", "") and len(cells[1]) >= 8:
            out[cells[0]] = cells[1]
    return out or None


def lint_unlicensed_surgery(root, rep):
    """⑦ 无照手术（8-19 裁定 C 档·警告级劝告不闸门）：女娲在场的树，
    手术面（类.md/schema.md/方法/*/任务.md）须有出处——宪法时刻基线
    （女娲领域无关，两表合并：自举骨架表 女娲/知识/宪法时刻.md ＋
    树根 知识/宪法时刻-领域面.md 的领域面大赦）或任一账目产物清单。
    无女娲的树检查整体休眠（无手术通道可言）；基线在而表不可解析＝
    休眠报信息。覆盖了出处的文件归既有对账检查管漂移，此处只管
    「从没进过账」。"""
    base_p = None
    nuwa = False
    for pat in (os.path.join(root, "域", "*", "类", "女娲"),
                os.path.join(root, "类", "女娲")):
        if glob.glob(os.path.join(pat, "类.md")):
            nuwa = True
            hit = sorted(glob.glob(os.path.join(pat, "知识", "宪法时刻.md")))
            if hit:
                base_p = hit[0]
            break
    if not nuwa:
        return                     # 无女娲：手术通道不存在，检查休眠
    if not base_p:
        rep.info("域/*/类/女娲", "女娲在场而宪法时刻基线缺席"
                                 "——无照手术检查休眠（落基线以启用）")
        return
    基线 = _parse_baseline(base_p, root)
    if 基线 is None:
        rep.info(os.path.relpath(base_p, root),
                 "宪法时刻基线无可解析表——无照手术检查休眠")
        return
    # 两表合并（8-19 裁定：女娲领域无关）——女娲自举骨架表只含元层；
    # 领域面大赦住树根 知识/宪法时刻-领域面.md（企业层承载领域）。
    领域面p = os.path.join(root, "知识", "宪法时刻-领域面.md")
    if os.path.isfile(领域面p):
        领域面 = _parse_baseline(领域面p, root)
        if 领域面:
            基线.update(领域面)
    账盖, 账回 = set(), set()
    for j in glob.glob(os.path.join(root, "**", "__结果.json"), recursive=True):
        try:
            data = json.load(open(j, encoding="utf-8"))
        except (json.JSONDecodeError, OSError, UnicodeDecodeError):
            continue
        账盖 |= {it.get("名称") for it in data.get("产物清单", [])
                 if it.get("名称")}
        账回 |= {it.get("名称") for it in data.get("回收清单", [])
                 if it.get("名称")}
    faces, rels = set(), set()
    for pre in (os.path.join(root, "域", "*", "类", "*"),
                os.path.join(root, "类", "*")):
        for f in ("类.md", "schema.md"):
            faces |= set(glob.glob(os.path.join(pre, f)))
        faces |= set(glob.glob(os.path.join(pre, "方法", "*", "任务.md")))
    rels = {os.path.relpath(f, root) for f in faces}
    for rel in sorted(rels):
        if rel in 基线:
            p = os.path.join(root, rel)
            if sha256(p) != 基线[rel]:
                rep.warn(rel, "手术后手改（基线 sha 不符，未经女娲）"
                              "——改类/改法请走女娲案卷")
        elif rel in 账盖:
            pass                  # 有账目出处：漂移归对账检查（lint_result_json）
        else:
            rep.warn(rel, "无照手术：不在宪法时刻基线也无任何账目出处"
                          "——立类/加方法请走女娲案卷")
    for rel in sorted(基线):
        if rel not in 账回 and not os.path.exists(os.path.join(root, rel)):
            rep.warn(rel, "基线文件已不在场（删除/搬移应走回收弧或女娲案卷）")


def lint_tree(root):
    root = os.path.abspath(root)
    rep = Report()
    lint_logical_model(root, rep)
    lint_case_closure(root, rep)
    lint_doc_type_keys(root, rep)
    lint_seq_type_arcs(root, rep)
    lint_theme_types(root, rep)
    lint_table_types(root, rep)
    lint_echo_types(root, rep)
    lint_entity_edges(root, rep)
    lint_domain_context(root, rep)
    lint_lineage(root, rep)
    lint_unlicensed_surgery(root, rep)
    bases = sorted(glob.glob(os.path.join(root, "_*版")))  # _简化版 等
    if not bases:
        bases = [root]

    for xlsx in sorted(glob.glob(os.path.join(root, "**", "*.xlsx"),
                                 recursive=True)):
        if os.path.basename(xlsx).startswith("~$"):
            continue
        try:
            wb = _need_openpyxl().load_workbook(xlsx, data_only=True)
        except Exception as e:
            rep.err(os.path.relpath(xlsx, root), f"xlsx 打不开：{e}")
            continue
        for ws in wb.worksheets:
            heads = headers_of(ws)
            if "【目录-文件名称】" in heads:
                lint_index_sheet(ws, xlsx, root, rep)
            elif "【参数文件类型】" in heads:
                lint_param_sheet(ws, xlsx, root, rep)
            else:
                rep.info(os.path.relpath(xlsx, root),
                         f"未识别的表头，跳过 sheet {ws.title!r}")

    prod_cover = _product_cover(root)
    msg_types = _message_type_names(root)
    recycle_cover = _recycle_cover(root)
    for jpath in sorted(glob.glob(os.path.join(root, "**", "__结果.json"),
                                  recursive=True)):
        base = next((b for b in bases if jpath.startswith(b + os.sep)), root)
        lint_result_json(jpath, base, root, rep, prod_cover, msg_types,
                         recycle_cover)

    return rep


def main():
    if len(sys.argv) != 2:
        print(__doc__)
        sys.exit(2)
    root = sys.argv[1]
    print(f"i3dna-lint 检查树：{os.path.abspath(root)}")
    rep = lint_tree(root)
    n_err = rep.dump()
    print("  结论：" + ("✗ 有错误" if n_err else "✓ 干净"))
    sys.exit(1 if n_err else 0)


if __name__ == "__main__":
    main()
