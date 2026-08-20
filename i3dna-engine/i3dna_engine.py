#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""i3dna_engine — I3DNA 微任务通用执行引擎（写半边；读半边=i3dna-lint）。

对客户三代包格式通用：
  A. 参数表微任务（8.5 模板族）：`__*大模型智能体*版本.xlsx` 声明输入/产物 → 全自动。
  B. 任务定义微任务（7.26 族）：`__*任务定义*.md` 作任务指令；产物需 --output 显式声明。
  C. 纯索引微任务（7.17 族）：位于 _任务文本/_微任务 轴下、索引外联行=输入；产物需 --output。

树根不再写死：**锚探针自校准**——用配方里的 `\\...\\` / 相对锚路径对任务目录的
各级祖先打分，哪个祖先能让最多锚解析成功，哪个就是树根；全不解析（如 7.17 的
D 盘机器路径）时退回"最顶层下划线目录的父目录"规则；--root 永远可强制指定。

引擎默认 **OMP（国产模型，合规车道）**：`omp -p --no-session @{prompt_file}`。
可用 --engine 或环境变量 I3DNA_LLM_CMD 换任意 CLI：命令串含 `{prompt_file}` 时替换为
提示词文件路径；不含时提示词走 stdin（如 `claude -p`）。

IO 两模式（--io，默认 write）：**write=引擎 agent 直写产物绝对路径**，本引擎只做
后置条件验收（存在/非空/被改动）——产物不经 stdout，免疫解释前缀污染与多产物切分；
stdout=产物内容回传、本引擎落盘（裸 LLM API 用，带前缀清洗兜底）。

用法：
  python3 i3dna_engine.py list      <包根目录>
  python3 i3dna_engine.py preflight <微任务目录> [--root D]
  python3 i3dna_engine.py run       <微任务目录> [--sandbox D] [--engine CMD]
                                    [--output 名称@锚目录 ...] [--timeout 秒] [--root D]
  python3 i3dna_engine.py backfill  <微任务目录> [--note 备注] [--root D]
  python3 i3dna_engine.py draft     <微任务目录> --case <案卷号> [--root D]
                                    （stdin JSON：[{路径, 内容|源}]——101号 起草车道，
                                     草稿落案卷零入账；产物槽落位=审批前半步）

多产物协议：声明 >1 个产物时，要求模型用分隔行
  <<<I3DNA-产物:文件名>>>
包裹每个文件，引擎按名切分落盘；单产物时 stdout 整体即文件。

纪律：预检不过不点火（无 oracle 不得 PASS）；产物写临时→rename；索引只许机器
代笔追加；每次执行写 __结果.json（名称=相对树根路径+字节+sha256，与 lint 对账兼容）；
backfill 只记账不伪造出处，缺失如实入「缺失输入或产物」。
"""
import argparse
import fnmatch
import glob
import hashlib
import json
import os
import re
import shlex
import shutil
import sqlite3
import subprocess
import sys
import tempfile
import uuid as _uuid
from datetime import datetime

openpyxl = None      # 仅 xlsx 参数表族需要——md 族树零此依赖（懒加载：
                     # Obsidian 插件用裸 python 也能跑 md 树）


def _need_openpyxl():
    global openpyxl
    if openpyxl is None:
        try:
            import openpyxl as _o
        except ImportError:
            sys.exit("需要 openpyxl：pip install openpyxl（仅 xlsx 参数表族）")
        openpyxl = _o
    return openpyxl

# --thinking high（8-19 裁定默认）：glm-5.3 输出+思考共池 128K（omp 目录
# maxTokens:131072 ＝官方规格；`omp models` 表显的 33K 是显示层陈旧映射，
# 不是发给 API 的值）。max 档沉思可整池烧穿——stopReason=max_tokens
# （8-19 K1 两跑皆然；z.ai bench：max 档 ~75K/任务、high 档 ~50K/任务），
# high 档装得下常规火、质量代价小；仍烧穿时引擎自动降档重试
# （THINKING_LADDER），见 cmd_run 重试块。
DEFAULT_ENGINE = "omp -p --no-session --thinking high @{prompt_file}"

TEXT_EXT = {".md", ".txt", ".py", ".sh", ".yaml", ".yml", ".json", ".csv"}
TASK_AXES = ("_任务文本", "_微任务")
SPLIT_MARK = "<<<I3DNA-产物:{name}>>>"
SPLIT_RE = re.compile(r"^<<<I3DNA-产物:(.+?)>>>\s*$", re.M)
INDEX_HEADERS = ["【序号】", "【描述】", "【目录-文件名称】", "【目录文件类型】",
                 "【模型存储模式】", "【独立存储输出文件名称】", "【外部引用文件目录】"]
INDEX_CONSTRAINT = ["*", "|>******<|", "必须为I3DNAI定义的有效文件名称",
                    "0(目录)|1(参数文件)", "0(集成存储)|1(独立存储)", "没有时为*", "没有时为*"]
INDEX_COORD = ["*", "0", "1", "2", "3", "4", "5"]
INDEX_SHEET = "__基础MN叉树模型_索引文件"


# ── 表格解析（与 i3dna-lint 同款）────────────────────────────


def is_machine_path(p):
    return bool(re.match(r"^[A-Za-z]:[\\/]", p))


def resolve_anchor(p, root):
    """`\\...\\a\\b`、`/.../a/b` 或裸相对路径 → root 下绝对路径。"""
    p = p.replace("\\", "/")
    p = re.sub(r"^/?\.\.\./", "", p)
    return os.path.join(root, *[seg for seg in p.split("/") if seg])


def sha256(path):
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def relroot(path, root):
    return os.path.relpath(path, root)


# ── 树根自校准 ───────────────────────────────────────────────

def ancestors(d, n=12):
    d = os.path.abspath(d)
    out = []
    for _ in range(n):
        out.append(d)
        parent = os.path.dirname(d)
        if parent == d:
            break
        d = parent
    return out


def calibrate_root(task_dir, probe_dirs, forced=None):
    """probe_dirs = 配方/索引里的锚目录串；对祖先打分，探针解析数最多者当树根。"""
    if forced:
        return os.path.abspath(forced)
    probes = [p for p in probe_dirs if p and p not in ("", "*") and not is_machine_path(p)]
    cands = ancestors(task_dir)
    if probes:
        best, best_score = None, -1
        for c in cands:                    # 由近及远，平分时取最近祖先
            score = sum(1 for p in probes if os.path.isdir(resolve_anchor(p, c)))
            if score > best_score:
                best, best_score = c, score
        if best_score > 0:
            return best
    # 兜底：最顶层以 _ 开头的祖先目录，其父即包根（__创建窗口/___勘察系统-01/_简化版 通吃）
    top_underscore = None
    for c in cands:
        if os.path.basename(c).startswith("_"):
            top_underscore = c
    return os.path.dirname(top_underscore) if top_underscore else os.path.dirname(
        os.path.abspath(task_dir))


# ── 三种微任务适配器 ─────────────────────────────────────────








_KV_DIR = os.path.dirname(os.path.abspath(__file__))
if _KV_DIR not in sys.path:
    sys.path.insert(0, _KV_DIR)
# M2 取值运行时。引擎与树里的符号任务**共用同一份**访问器——读写同源，
# 「key 是多方共识」才不会退化成两边各搓一套正则然后悄悄漂移。
from i3dna_kv import get_value, set_value, render, norm_key   # noqa: E402,F401
from i3dna_kv import _fields as _kv_fields                     # noqa: E402
import i3dna_store as store                                   # noqa: E402
from i3dna_store import (norm, is_coord_row, data_rows,       # noqa: E402,F401
                         headers_of, classify_kind,
                         param_rows as _param_rows,
                         index_rows_from_file as _index_rows_from_file,
                         index_input_rows as _index_input_rows,
                         load_account, save_account,
                         account_exists as _account_exists)


def _frontmatter_rows(md_path):
    """兼容垫片：旧签名 (rows, instruction)。"""
    rows, instruction, _fm = store.frontmatter_rows(md_path)
    return rows, instruction


def _numeric(v):
    try:
        float(v)
        return True
    except (TypeError, ValueError):
        return False


_CMP = {"小于": lambda a, b: a < b, "不大于": lambda a, b: a <= b,
        "大于": lambda a, b: a > b, "不小于": lambda a, b: a >= b,
        "等于": lambda a, b: a == b, "不等于": lambda a, b: a != b}


def _eval_cond(path, cond):
    """求值一条使能条件，返回 (通过?, 说明)。

    文法**刻意不可嵌套**：一次取值 + 一个比较 + 一个字面量，没有与或非、
    没有算术、没有跨文档引用。需要复杂判断？退回符号任务算好、写成一个键，
    引擎再来取那个键——协调层只做取值与比较，不做求值。
    （可嵌套的文法 = frontmatter 里长出一门语言 = 计算从侧门溜回协调层。）

    取不到值一律 fail-closed：宁可停机，不可无界。
    """
    key = cond.get("取值")
    if not key:
        return False, "使能条件缺【取值】键"
    raw = get_value(path, key)
    if raw is None:
        return False, f"取值失败：{os.path.basename(path)} 无键「{key}」"
    if isinstance(raw, list):
        return False, f"使能条件不比较列表值：{key} 取到 {len(raw)} 项"
    for op, fn in _CMP.items():
        if op not in cond:
            continue
        want = cond[op]
        # 陈述事实关系（不X ↔ X 取反）：ok=False 时显示实际成立的关系——
        # 「认定金额=300 小于 5000」而不是「不不小于」。
        neg = op[1:] if op.startswith("不") else "不" + op
        try:
            ok = (fn(float(raw), float(want))
                  if _numeric(raw) and _numeric(want) else fn(str(raw), str(want)))
        except (TypeError, ValueError):
            return False, f"使能条件不可比较：{key}={raw!r} {op} {want!r}"
        return ok, f"{key}={raw} {op if ok else neg} {want}"
    return False, f"使能条件无可识别比较算子（支持：{'、'.join(_CMP)}）"


def _cond_block(task):
    """使能条件闸门。条件绑在**输入弧**上：弧不在场则条件不参与判定——
    与可缺弧同构（没有消息，就没有消息上的字段）。任一条不满足即不使能。"""
    for r in task["rows"]:
        cond = r.get("cond")
        if not cond or r["kind"] != "输入" or not r["path"]:
            continue
        if not os.path.isfile(r["path"]):
            continue
        ok, why = _eval_cond(r["path"], cond)
        if not ok:
            return f"使能条件不满足：{why}"
    return None




EXEC_DIR = "执行程序"
TREE_TOPS = ("域", "类", "实例", "场所", "消息", "状态", "知识", "部门")
                                # 树顶命名空间：draft 路径首段命中＝树目标
                                # （须恰为本任务产物槽，防起草面任意树写）


def exec_entry(task_dir):
    """符号主义判据（结构非声明）：任务目录下 执行程序/主程序.py|.sh 存在即符号任务。
    删掉该目录即退回联结——标签可以撒谎，结构不能。跨全部四种任务格式。"""
    for name in ("主程序.py", "主程序.sh"):
        p = os.path.join(task_dir, EXEC_DIR, name)
        if os.path.isfile(p):
            return p
    return None


def task_def_files(task_dir):
    """任务定义文件（弧声明+指令的载体）——编译过期账的必入依据。"""
    md_decl = os.path.join(task_dir, "任务.md")
    if os.path.isfile(md_decl):
        return [md_decl]
    hits = sorted(glob.glob(os.path.join(task_dir, "__*大模型智能体*版本.xlsx"))) \
        or sorted(glob.glob(os.path.join(task_dir, "__*任务定义*.md")))
    return hits


def compile_stale(task_dir, root):
    """编译过期核查（符号同一性，机械可核）：__编译记录.json 里的依据 sha
    与现况比对，返回变了的文件名列表。无记录（手写执行程序）→ 空列表。"""
    rec_path = os.path.join(task_dir, "__编译记录.json")
    if not os.path.isfile(rec_path):
        return []
    try:
        rec = json.load(open(rec_path, encoding="utf-8"))
    except Exception:
        return ["__编译记录.json 不可解析"]
    changed = []
    for rel, old in (rec.get("依据sha") or {}).items():
        p = os.path.join(root, rel)
        if not os.path.isfile(p) or sha256(p) != old:
            changed.append(rel)
    return changed


CASE_MARK = "{实例}"
CASE_NUM_MARK = "{案卷号}"     # 只代入案卷号本身——跨架弧写显式路径用:
                               # 实例/研发/{案卷号}/代码/核心.py。目录即协同
                               # (bean经共享库所协作,无注入语义,99号修订)
KLASS_MARK = "{类}"
CONTENT_MARK = re.compile(r"\{([^{}./]+)\.([^{}./]+)\}")
                                # 内容记号 {<输入名>.<键>}:产物名按输入内容
                                # 取（立场所:场所名住申请里,不挂案卷号）。
                                # 仅产物弧可用;取不到值=载荷缺失,响亮拒绝


def case_rel(kr, case):
    """实例库寻址：实例/<类名>/<实例号>（类根=树根的单类树退化为 实例/<实例号>）。
    实例=实例，住企业实例库不住类目录——档案室独立于制度汇编：
    保管期限与制度废止解耦（删类不删档），类目录保持纯 M1，
    法的 git 史不被事务 journal 淹没。按类分架保住结构可推导性。"""
    if case and ("/" in case or case.startswith(".")):
        sys.exit(f"案卷号不干净：{case!r}（无 / 不以 . 开头）——防路径注入写出树外")
    shelf = os.path.basename(kr) + "/" if kr else ""
    return f"实例/{shelf}{case}"


def sealed_class(task):
    """封存门（8-20 立法口归一配套）：类根在场 封存.md＝类已封存。
    封不删＝证据保留，不是通道保留——新点火/新办结拒绝（历史账与
    git 史不动）；接棒法见封存碑。无类根（三代包任务）恒假。"""
    kr = klass_rel(task["task_dir"], task["root"])
    return bool(kr) and os.path.isfile(
        os.path.join(task["root"], kr, "封存.md"))


def klass_rel(task_dir, root):
    """类根 = 方法目录的父目录（结构判据，与 exec_entry 同款式）：
    task_dir=企业/类X/方法/T → 类X；task_dir=root/方法/T（单类树）→ root。
    对旧 任务/ 轴同理。返回树根相对路径，根本身返回 ""。"""
    r = os.path.relpath(os.path.dirname(os.path.dirname(task_dir)), root)
    return "" if r == "." else r.replace(os.sep, "/")


def load_task(task_dir, forced_root=None, extra_outputs=(), case=None,
              executor=None, tolerate_content_marks=False):
    """微任务目录 → 统一 Task 结构；kind ∈ {frontmatter, 参数表, 任务定义, 纯索引}。

    M1/M0 分层（8-09）：弧路径可含 `{实例}` 模板记号——任务定义是 M1（类型），
    实例是 M0（实例）。点火签名从 规则×执行者 升为 规则×实例×执行者。
    引擎对实例号**只代入不解释**（与 get_value 同一条 M2 纪律）：
    `{实例}` → `实例/<case>`。声明了记号却没给 --case ＝ 载荷缺失，响亮拒绝。"""
    task_dir = os.path.abspath(task_dir)
    md_decl = os.path.join(task_dir, "任务.md")
    recipes = store.recipe_paths(task_dir)
    mds = store.taskdef_md_paths(task_dir)
    # 持久化已隔离（i3dna_store）：任务定义底物按目录自动识别，引擎只见弧列表
    def_store = store.open_def_store(task_dir)
    if def_store is not None:
        kind = def_store.kind
        instruction, rows = def_store.load_arcs()
    else:
        idx_rows = _index_input_rows(task_dir)
        under_axis = any(a in task_dir.split(os.sep) for a in TASK_AXES)
        if idx_rows and under_axis:
            kind, rows = "纯索引", idx_rows
        else:
            sys.exit(f"{task_dir} 不是可识别的微任务目录"
                     "（无参数表/任务定义，且不在任务轴下或索引无外联行）")
    for spec in extra_outputs:                 # --output 名称@锚目录
        name, _, pdir = spec.partition("@")
        rows.append({"desc": "显式产物", "ptype": "", "pdir": pdir or "*",
                     "pver": "", "pname": name, "kind": "产物"})
    marked = any((CASE_MARK in (r["pdir"] + r["pname"])
                  or CASE_NUM_MARK in (r["pdir"] + r["pname"]))
                 for r in rows) or CASE_MARK in instruction
    if marked and not case:
        sys.exit(f"{task_dir} 的弧声明含实例记号（M1 定义），"
                 "点火须给 --case <实例号>（M0 实例）")
    # 定根：无记号走老校准；有记号（{类}/{实例}）改用"代入后计分"——
    # 对每个候选祖先 R，按 R 试代入记号后数命中（文件命中 2 分 > 目录 1 分）。
    # 文件级权重解决两层同名目录（公司知识/ vs 类知识/）打平手的歧义。
    has_marker = any(KLASS_MARK in (r["pdir"] + r["pname"])
                     or CASE_MARK in (r["pdir"] + r["pname"])
                     or CASE_NUM_MARK in (r["pdir"] + r["pname"])
                     for r in rows)
    if not has_marker:
        root = calibrate_root(task_dir, [r["pdir"] for r in rows], forced_root)
    else:
        cands, cur = [], os.path.dirname(task_dir)
        if forced_root:
            cands.append(os.path.abspath(forced_root))
        for _ in range(8):
            if cur not in cands:
                cands.append(cur)
            if os.path.isdir(os.path.join(cur, ".git")):
                break
            nxt = os.path.dirname(cur)
            if nxt == cur:
                break
            cur = nxt

        def _root_score(R):
            kr0 = klass_rel(task_dir, R)
            cp = case_rel(kr0, case) if case else None
            sc = 0
            for r in rows:
                pd, pn = r["pdir"], r["pname"]
                pd = pd.replace(KLASS_MARK + "/", kr0 + "/" if kr0 else "")
                pd = pd.replace(KLASS_MARK, kr0 if kr0 else ".")
                if cp:
                    pd, pn = pd.replace(CASE_MARK, cp), pn.replace(CASE_MARK, cp)
                    pd, pn = pd.replace(CASE_NUM_MARK, case), pn.replace(CASE_NUM_MARK, case)
                if not pd or pd in ("*",) or is_machine_path(pd):
                    continue
                full = os.path.join(R, pd)
                if os.path.isfile(os.path.join(full, pn)):
                    sc += 2
                elif os.path.isdir(full):
                    sc += 1
            return sc
        # 平手取最浅祖先：根是结构事实。全实例作用域的任务（如质检）从类根
        # 和企业根解析都通——点火无歧义但**账的键有歧义**，取最浅根让账名canonical。
        # 但显式 forced_root 平手优先：未出生案卷（实体尚未实例化）让全候选零分时，
        # 用户给的根就是唯一权威（trade-v3 实测：产物未出生→零分平手→最浅根抢走）。
        best = max(cands, key=lambda R: (_root_score(R),
                                         -len(os.path.abspath(R))))
        if forced_root and os.path.abspath(forced_root) in cands \
                and _root_score(forced_root) >= _root_score(best):
            root = os.path.abspath(forced_root)
        else:
            root = best
    kr = klass_rel(task_dir, root)
    casep = case_rel(kr, case) if case else None

    def _sub(text):
        out = text.replace(KLASS_MARK + "/", (kr + "/" if kr else ""))
        out = out.replace(KLASS_MARK, kr if kr else ".")
        if case:
            out = out.replace(CASE_NUM_MARK, case)
            out = out.replace(CASE_NUM_MARK, case)
        if casep:
            out = out.replace(CASE_MARK, casep)
        return out
    for r in rows:
        r["pdir"] = _sub(r["pdir"])
        r["pname"] = _sub(r["pname"])
    instruction = _sub(instruction)

    # 内容记号（产物按输入内容取名）：{<输入名>.<键>} ← get_value(输入文件,键)
    # ——按内容取名（用户裁决 2026-08-19）。纪律同 {实例}：只许产物弧；
    # 无此输入/无此键/值不干净（非标量、含 / 或 ..）＝载荷缺失，响亮拒绝
    # ——产物名是文件系统安全边界，不猜。
    if any(CONTENT_MARK.search(r["pdir"] + r["pname"]) for r in rows):
        ins = {}
        for r in rows:
            if CONTENT_MARK.search(r["pdir"] + r["pname"]) \
                    and r["kind"] != "产物":
                sys.exit(f"内容记号只许在产物弧：{r['pdir']}/{r['pname']}")
            if r["kind"] == "输入" and not is_machine_path(r["pdir"]):
                pth = (os.path.join(task_dir, r["pname"])
                       if r["pdir"] in ("", "*")
                       else os.path.join(resolve_anchor(r["pdir"], root),
                                         r["pname"]))
                ins.setdefault(os.path.splitext(r["pname"])[0], pth)

        def _csub(text):
            def rep(m):
                stem, key = m.group(1), m.group(2)
                if stem not in ins or not os.path.isfile(ins[stem]):
                    if tolerate_content_marks:
                        # 起草车道自举（8-20 真用例「加撤域」）：往空案卷
                        # 起草 申请.md 正是 draft 的使命——此刻申请未落盘，
                        # 记号留原样不炸（产物槽带记号判不了树目标，落位
                        # 本就无从谈起；批准时申请在场，严格解析照旧）。
                        return m.group(0)
                    sys.exit(f"内容记号 {m.group(0)} 无对应在场输入"
                             f"「{stem}」——载荷缺失")
                v = get_value(ins[stem], key)
                if not v or isinstance(v, (list, dict)) \
                        or "/" in str(v) or ".." in str(v):
                    sys.exit(f"内容记号 {m.group(0)} 取值失败：{v!r}"
                             "——产物名须是干净标量（无 / 无 ..）")
                return str(v)
            return CONTENT_MARK.sub(rep, text)

        for r in rows:
            r["pdir"], r["pname"] = _csub(r["pdir"]), _csub(r["pname"])

    def resolve(r):
        if is_machine_path(r["pdir"]):
            r["path"] = None
        elif r["pdir"] in ("", "*"):
            r["path"] = os.path.join(task_dir, r["pname"])
        else:
            r["path"] = os.path.join(resolve_anchor(r["pdir"], root),
                                     r["pname"])
        # 树根包含性复检（8-20 对抗验收）：resolve_anchor 只剥一个前导
        # ../，弧声明里残留的 .. 段会把产物/输入解析到树根之外（写面=
        # 落盘出树、读面=树外文件进 prompt）。越出=不可解析（同机器
        # 绝对路径待遇：预检✗拒绝执行），不猜不修。两侧都 realpath——
        # macOS /var→/private/var 这类前缀符号链接不消解会假阴性。
        if r["path"] is not None:
            rp = os.path.realpath(r["path"])
            rr = os.path.realpath(root)
            if not (rp == rr or rp.startswith(rr + os.sep)):
                print(f"  ⚠ 弧路径越出树根（含 ../ 段）："
                      f"{r['pdir']}/{r['pname']} → {rp}"
                      "——按不可解析处理（拒绝执行）")
                r["path"] = None
        return r

    # 两级寻址展开（目录寻册、册内归查询参数）：输入若是索引文件，其外联行也是输入
    expanded = []
    for r in map(resolve, rows):
        expanded.append(r)
        if (r["kind"] == "输入" and r["path"] and os.path.isfile(r["path"])
                and re.search(r"索引文件.*\.xlsx$", r["pname"])):
            for sub in _index_rows_from_file(r["path"]):
                sub["desc"] = f"↳{sub['desc']}"
                expanded.append(resolve(sub))
    entry = exec_entry(task_dir)
    def_path = (md_decl if os.path.isfile(md_decl)
                else recipes[0] if recipes else mds[0] if mds else None)
    # M2 第九词「执行者」：人工工位牌。引擎只认"人"（不自动点火），
    declared = get_value(md_decl, "执行者") if kind == "frontmatter" else None
    # executor 值域（开放集）三分:
    #   机制值: 人|agent|程序 —— 声明执行机制(绿/蓝/红),点火权归属由此定
    #   主体值: 实例/<类>/<k>(如 部门/D01) —— 谁的作业(审计问责),机制按结构判
    #   空:     未声明(=旧默认)
    # 规则: CLI --executor 可给主体值(换"哪个部办"),但不得覆盖声明的机制值——
    #       "人"的工位不许替成部门绕过人工守卫;"程序/agent"同守
    MECH = {"人", "agent", "程序"}
    if declared in MECH:
        # 声明机制值:CLI 给主体值(实例/…) → 主体入账,机制值留在声明(结构+声明共裁);
        # CLI 给另一机制值 → 声明受保护(点火权不可被 CLI 篡改)
        if executor is None or executor in MECH:
            executor = declared
        # else: executor 是主体值,直接用(部门雇佣 agent/程序/人)
    elif executor is None:
        executor = declared
    # 主体值声明(D01)而结构无执行程序 → 蓝任务由 agent 代跑(部门雇佣 agent);
    # 结构有执行程序 → 红(程序是部门的工具)。机制始终由结构+机制值共同裁定
    # 结账站（缺陷19·悬账门互锁 8-21）：任务卡声明 结账: 真 → 本站办结/
    # 点火承担案卷打烊全夹盘点（_closure_pending）；普通站不问此门。
    closing = kind == "frontmatter" and \
        str(get_value(md_decl, "结账") or "").strip() in ("真", "True", "true")
    return {"kind": kind, "task_dir": task_dir, "root": root, "case": case,
            "executor": executor, "executor_declared": declared,
            "def_path": def_path, "closing": closing,
            "instruction": instruction, "rows": expanded,
            "exec_entry": entry, "paradigm": "符号" if entry else "联结"}


def _theme_laws(type_files):
    """主题法收集（形状定律 8-21·工单1号）：消息类型文件声明「主题:」＝
    **目录即类型**——该目录下一切非 __ 文件皆此类型。值为路径模式，文法
    同「路径:」（可含 {案卷号}/{实例}/{类}/*）。零声明＝空表＝判型走老路
    （逐字节兼容，向后兼容律）。"""
    out = []
    for tf in type_files:
        v = get_value(tf, "主题")
        if isinstance(v, str) and v.strip():
            out.append((v.strip(), tf))
    return out


def _theme_pat(law, case=None, casep=None, kr=None):
    """主题模式代入（文法同 路径: 的记号集）：{类}/{案卷号}/{实例} 按本帧
    代入，未知记号退 *——判型是读侧提问不是落位断言，缺参宽松匹配；
    落位仍由产物弧显式声明（uuid 命名是 2 号单）。"""
    p = law.replace(KLASS_MARK, kr or "*").replace(CASE_NUM_MARK, case or "*")
    return p.replace(CASE_MARK, casep or "*")


def _theme_hit(drel, laws, case=None, casep=None, kr=None, name=None):
    """①主题级判型：文件的树内相对**目录**命中某主题模式 → 该类型文件
    （不看文件名——形状定律：类型判据先目录后文件名）。drel 为空（文件
    直躺树根）不参与；`__`/点前缀文件豁免（账不进账老钉，§8.12 同源——
    主题目录里的 __ 件/隐藏件不是单据）。"""
    if not drel:
        return None
    if name and os.path.basename(name).startswith(("__", ".")):
        return None
    for law, tf in laws:
        if fnmatch.fnmatch(drel, _theme_pat(law, case, casep, kr)):
            return tf
    return None


def _queue_tickets(dpath):
    """主题队列在场单（形状定律 8-21·工单2）：字典序——选单规则确定性
    可复算（要时序就给文件名加时间戳前缀）。`__` 前缀＝账不进账
    （§8.12 同源）、点前缀＝隐藏件，皆豁免。"""
    try:
        return sorted(f for f in os.listdir(dpath)
                      if not f.startswith((".", "__"))
                      and os.path.isfile(os.path.join(dpath, f)))
    except OSError:
        return []


def _theme_laws_global(root):
    """全树主题法一表（**/消息/*.md 递归）——点火/办结/draft 三处共用。"""
    return _theme_laws(sorted(glob.glob(os.path.join(root, "**", "消息",
                                                     "*.md"), recursive=True)))


def _theme_queue(task, r, laws):
    """目录弧行指向主题目录 → 该主题类型文件（判「这个队列是什么单的
    队列」——目录弧 drel＝目录自身）；否则 None。"""
    if not r.get("path") or not os.path.isdir(r["path"]):
        return None
    kr = klass_rel(task["task_dir"], task["root"])
    c = task.get("case")
    drel = relroot(r["path"], task["root"]).replace(os.sep, "/")
    return _theme_hit(drel, laws, c, case_rel(kr, c) if c else None, kr,
                      r["path"])


def _closure_pending(task):
    """结账站盘点（缺陷19·悬账门互锁 8-21）：返回 [(目录, 张数)] ——空表
    ＝可打烊。案卷**全部**案卷寻址主题夹清点：非本任务弧的夹须零张；本
    任务消费弧的夹容恰一张（本火/本结即吃，N 张 N 轮火到空，最后一张
    那轮才轮到关门）；本任务门弧（清空: 真）的夹零张（门开才谈得上打烊）。
    普通站不问此门（对非消费目录休眠）——结账不再是每次办结的事，是
    终点站的事；lint_case_closure 同语义第二实现，互为见证。"""
    _case = task.get("case")
    if not _case:
        return []
    root = task["root"]
    kr = klass_rel(task["task_dir"], root)
    casep = case_rel(kr, _case)
    mine = {}      # realpath(夹) → 是否门弧（多弧同夹从严：任一 drain 即门）
    for r in task["rows"]:
        if r["kind"] == "输入" and r["path"] and os.path.isdir(r["path"]):
            rp = os.path.realpath(r["path"])
            mine[rp] = mine.get(rp, False) or bool(r.get("drain"))
    bad = []
    for tf in glob.glob(os.path.join(root, "**", "消息", "*.md"), recursive=True):
        law = get_value(tf, "主题")
        if not isinstance(law, str) or not law.strip():
            continue
        if "{案卷号}" not in law and CASE_MARK not in law:
            continue          # 非案卷寻址的公共队列——不是本案卷的在途
        pat = law.replace("{案卷号}", _case).replace(CASE_MARK, casep) \
                 .replace(KLASS_MARK, kr or "*")
        for d in glob.glob(os.path.join(root, pat)):
            if not os.path.isdir(d):
                continue
            n = len(_queue_tickets(d))
            cap = 1 if mine.get(os.path.realpath(d)) is False else 0
            if n > cap:
                bad.append((relroot(d, root), n))
    return bad


def _uuid_named(task, r, laws):
    """uuid 落位代起名判定（形状定律 8-21·工单2）：产物弧是目录弧
    （无扩展名、指向主题目录）且该主题类型声明 命名: uuid → 代起名
    <种名>__<uuid4>.md——并发零碰撞（缺陷1 抢名非原子/缺陷3 同名覆盖
    釜底抽薪），agent 只交内容不交名。无声明＝现状固定名，零迁移；
    固定名消息（审查单/返工单文件槽 0/1 语义）不受扰。"""
    if os.path.splitext(r["pname"])[1] or not r.get("path"):
        return None                    # 文件弧（有扩展名）＝老行为
    kr = klass_rel(task["task_dir"], task["root"])
    _case = task.get("case")
    drel = relroot(r["path"], task["root"]).replace(os.sep, "/")
    tf = _theme_hit(drel, laws, _case,
                    case_rel(kr, _case) if _case else None, kr)
    if tf is None:
        return None
    if str(get_value(tf, "命名") or "").strip().lower() != "uuid":
        return None
    return tf


def _type_file(task, r, kind):
    """槽的类型声明查找链（判型两级·形状定律 8-21）：**①先目录**——消息
    类型文件声明「主题:」且文件所在目录命中主题模式 → 类型＝该主题类型，
    不看文件名（目录即类型，uuid 身份让文件名退休）；**②后文件名**（老
    路）：类根/<kind>/ → 企业根/<kind>/（Object 兜底）。类型文件名=槽名去
    扩展+.md（状态.json→状态.md，返工单.md→返工单.md）。同键消息后缀
    剥离（104号 缺陷3）：类型件命名**单种**（消息/请求审查单.md），实例形
    请求审查单__<案卷>__r2.md 按 __ 段逐层剥后缀寻种——带案卷号/顺号
    后缀的单此前寻不到种，is_message 恒 False（悬账门键规则同款
    split("__")，两处判定由此对齐）。"""
    if not r.get("path"):
        return None
    kr = klass_rel(task["task_dir"], task["root"])
    if kind == "消息":                       # ①主题级（消息类型立法载体）
        klass_root = os.path.join(task["root"], *(kr.split("/") if kr else []))
        laws = _theme_laws(
            sorted(glob.glob(os.path.join(klass_root, "消息", "*.md")))
            + sorted(glob.glob(os.path.join(task["root"], "消息", "*.md"))))
        _case = task.get("case")
        _rel = os.path.relpath(r["path"], task["root"]).replace(os.sep, "/")
        # 目录弧（队列本体）：drel＝目录自身——判「这个队列是什么单的队列」；
        # 文件弧（单据）：drel＝所在目录——判「这张单是什么单」。
        drel = _rel if os.path.isdir(r["path"]) else os.path.dirname(_rel)
        tf = _theme_hit(drel, laws, _case,
                        case_rel(kr, _case) if _case else None, kr,
                        r["path"])
        if tf is not None:
            return tf
    stem = os.path.splitext(os.path.basename(r["path"]))[0]
    parts = stem.split("__")
    names = [stem + ".md"] + ["__".join(parts[:i]) + ".md"
                              for i in range(len(parts) - 1, 0, -1)]
    for name in names:
        cands = [os.path.join(task["root"], *(kr.split("/") if kr else []),
                              kind, name),
                 os.path.join(task["root"], kind, name)]
        for c in cands:
            if os.path.isfile(c):
                return c
    return None


def is_message(task, r):
    """消息（生命周期件：开出→消费→销毁）——账面条目缺席=已消费。"""
    return _type_file(task, r, "消息") is not None


def is_state(task, r):
    """状态字段（常驻）——可缺产物缺席=本轮未更新，**不许当收回删除**。"""
    return _type_file(task, r, "状态") is not None


def is_field_area(task, r):
    """字段区＝状态类型且带「属主」表（键级多方更新件，如 状态.json）。
    只有它享受「守门不催火」——需求/报告等文档状态是内容依赖，变更必须催火。
    判据是 dict 形状不是词义（引擎不解释「属主」写了谁）。"""
    tf = _type_file(task, r, "状态")
    if not tf:
        return False
    return isinstance(get_value(tf, "属主"), dict)


def _next_seq(path):
    """同键消息顺号（104号 缺陷3）：目标在场→落为 __rN 新量子。r1＝无
    后缀形，存量零迁移。每次发射＝一张新单（账房律：同额同日两张汇款单
    是两张）。"""
    base, ext = os.path.splitext(path)
    n = 2
    while os.path.exists(f"{base}__r{n}{ext}"):
        n += 1
    return f"{base}__r{n}{ext}"


def _seq_enabled(task, r):
    """顺号（104修订2/工单106）：消息类型的投递属性——类型文件 frontmatter
    声明「顺号: 真」才启用，默认关闭（＝现状行为，安全兜底）。要式性：
    类型即效力，投递方式住在单据的法定格式（消息/<单名>.md）里。修订1
    判据「跨主角＝非本任务输入的消息」有洞：文件弧按固定名消费的单
    （审查单/持有单）会被误编号 __rN——消费者永远读不到新单，故废止。"""
    tf = _type_file(task, r, "消息")
    return tf is not None and str(get_value(tf, "顺号")) in ("真", "True", "true")


def _area_fields(path):
    """字段区读器：与 M2 取值运行时同一张读法（json 顶层 / md 机器面）——
    此前硬编码 json.load，md 字段区两侧恒 {}，守卫形同虚设（8-11 评审 #4）。"""
    try:
        return dict(_kv_fields(path))
    except Exception:
        return {}


def _field_snap(task, outs):
    """点火前给字段区产物拍快照：键值 dict（执法比对）＋原始字节（还原）。"""
    snaps = {}
    for r in outs:
        if not is_field_area(task, r) or not r["path"]:
            continue
        raw = None
        if os.path.isfile(r["path"]):
            with open(r["path"], "rb") as f:
                raw = f.read()
        snaps[r["pname"]] = (r, _area_fields(r["path"]), raw)
    return snaps


def _field_guard(task, snaps):
    """键级属主执法：点火前后对字段区做键 diff——改了非属主的键=响亮失败。
    属主表来自类型文件（声明）；引擎只比较字符串，不解释谁是谁。
    这是运行时执法，比声明期 lint 强：lint 查不了行为，diff 查得了。
    出生豁免：点火前文件不存在=实体出生（造档案），不 diff——
    出生是造实体不是改字段；已有档案的键变更才在属主辖权内
    （85号弧角色契约：生命周期的"开"端不受守门，中端受）。"""
    me = os.path.basename(task["task_dir"])
    for name, (r, old, raw) in snaps.items():
        if raw is None:
            continue                    # 出生：点火前无文件，无"被改的旧值"
        tf = _type_file(task, r, "状态")
        owners = get_value(tf, "属主") if tf else None
        if not isinstance(owners, dict):
            continue                        # 无属主表=不设防（老树宽容）
        own_n = {norm_key(k): v for k, v in owners.items()}
        new = _area_fields(r["path"])
        changed = {k for k in set(old) | set(new) if old.get(k) != new.get(k)}
        rogue = sorted(k for k in changed if str(own_n.get(k, "")) != me)
        if rogue:
            # 先还原后拦账：把字段区整文件回滚到点火前字节——检测而不撤销
            # 等于让被拒的点火留下真实污染（8-11 评审 #2）。还原的是快照，
            # 不是谁的新值，引擎作为执法者持有这支笔。
            if raw is None:
                if os.path.isfile(r["path"]):
                    os.remove(r["path"])
            else:
                tmpf = r["path"] + ".tmp"
                with open(tmpf, "wb") as f:
                    f.write(raw)
                os.replace(tmpf, r["path"])
            sys.exit(f"越权改键：{name} 的 {rogue} 属主是 "
                     f"{[own_n.get(k, '?') for k in rogue]}，不是本任务「{me}」"
                     "——字段区只许属主定点写；已回滚该字段区到点火前快照")


def _lineage_file(field_path):
    """血缘文件:与字段区文件同目录的 __血缘.md(append-only,机器代笔)。"""
    return os.path.join(os.path.dirname(field_path), "__血缘.md")


def _append_lineage_for(task, snaps, t):
    """血缘附注(94号):键值区每个被改键记一行历史出身——
    键 :: 值哈希 :: 来源(案卷号/方法) :: 时间。append-only,引擎代笔,
    不信任 agent 手写。出生(点火前无文件)记一条「出生」。
    只在执法(_field_guard)通过后调用——血缘只记合法键。"""
    case = task.get("case") or ""
    me = os.path.basename(task["task_dir"])
    src = f"{case}/{me}" if case else me
    for name, (r, old, raw) in snaps.items():
        if not r.get("path") or not os.path.isfile(r["path"]):
            continue
        lf = _lineage_file(r["path"])
        lines = []
        if raw is None:
            lines.append(f"出生 :: {sha256(r['path'])[:16]} :: {src} :: {t}")
        else:
            new = _area_fields(r["path"])
            for k in sorted(set(old) | set(new)):
                if old.get(k) == new.get(k):
                    continue
                v = new.get(k)
                h = hashlib.sha256(repr(v).encode("utf-8")).hexdigest()[:16] \
                    if v is not None else "None"
                lines.append(f"{k} :: {h} :: {src} :: {t}")
        if not lines:
            continue
        os.makedirs(os.path.dirname(lf), exist_ok=True)
        with open(lf, "a", encoding="utf-8") as f:
            for ln in lines:
                f.write(ln + "\n")
        print(f"  🩸 血缘 → {os.path.relpath(lf, task['root'])}（{len(lines)} 条）")


def seed_state_defaults(root, kroot, cdir):
    """实例字段区零值初始化（运行时职责，JVM 式：字段先零值，构造参数归人）。
    声明的状态类型带「默认」表且实例内文件尚不存在 → 按默认播种；
    已存在的文件**绝不触碰**（值是活的）。确定性、来自声明——
    这是构造的机械补全，不是引擎替人做主。"""
    seen = set()
    for base in (kroot, root):
        for tf in sorted(glob.glob(os.path.join(base, "状态", "*.md"))):
            name = os.path.basename(tf)
            if name in seen:
                continue
            seen.add(name)
            law, defaults = get_value(tf, "路径"), get_value(tf, "默认")
            if not law or not isinstance(defaults, dict):
                continue
            rel = str(law).split(CASE_MARK + "/", 1)[-1]
            fp = os.path.join(cdir, *rel.split("/"))
            if os.path.isfile(fp):
                continue
            os.makedirs(os.path.dirname(fp) or cdir, exist_ok=True)
            for k, v in defaults.items():
                set_value(fp, k, v)
            print(f"  🌱 字段区零值初始化 → {os.path.relpath(fp, root)}")


def blank_slot(path):
    """空槽判据：文件不存在，或内容为空/纯空白——空槽是落点不是内容，
    必需输入的 token 未到（企业镜像：空白表单≠已提交）。
    目录槽=信箱（Actor 收件箱）：无文件=空箱（token 未到），
    有文件=有待处理消息（token 在场）——目录的存在不算内容。"""
    if not path:
        return True
    if os.path.isdir(path):
        return not any(not f.startswith(".") for f in os.listdir(path))
    if not os.path.isfile(path):
        return True
    try:
        size = os.path.getsize(path)
        if size == 0:
            return True
        if size <= 64:
            with open(path, encoding="utf-8", errors="replace") as f:
                return not f.read().strip()
    except OSError:
        return True
    return False


def rec_dir(task):
    """点火记录（账）的落点。账随 bean 的宿主：有状态 bean（有案卷）
    账落案卷 实例/<类>/<case>/__账/；无状态 bean（无案卷，类方法）账落
    场所（类根下任务目录）——ARCHITECTURE.md §5「无状态 bean 账落场所」。"""
    if task.get("case"):
        kr = klass_rel(task["task_dir"], task["root"])
        rel = case_rel(kr, task["case"])
        return os.path.join(task["root"], *rel.split("/"), "__账",
                            os.path.basename(task["task_dir"]))
    return task["task_dir"]


def find_tasks(pkg_root):
    """包根下发现四种微任务目录。"""
    hits = {}
    for f in glob.glob(os.path.join(pkg_root, "**", "任务.md"), recursive=True):
        hits[os.path.dirname(f)] = "frontmatter"
    for pat, kind in (("__*大模型智能体*版本.xlsx", "参数表"),
                      ("__*任务定义*.md", "任务定义")):
        for f in glob.glob(os.path.join(pkg_root, "**", pat), recursive=True):
            hits.setdefault(os.path.dirname(f), kind)
    for f in glob.glob(os.path.join(pkg_root, "**", "__*索引文件*.xlsx"),
                       recursive=True):
        d = os.path.dirname(f)
        if d in hits or not any(a in d.split(os.sep) for a in TASK_AXES):
            continue
        if _index_input_rows(d):
            hits[d] = "纯索引"
    return dict(sorted(hits.items()))


# ── 输入读取与 prompt ────────────────────────────────────────

def read_input(path):
    if os.path.isdir(path):     # 目录=树的子图:文件树+小文本文件全文
        # 泰勒式输入:类目录是 skill/context 包(类.md/schema.md/知识件),给 LLM 全读
        out = []
        for r, ds, fs in os.walk(path):
            ds[:] = [d for d in ds if not d.startswith((".", "__"))]
            rel = os.path.relpath(r, path)
            for f in sorted(fs):
                fp = os.path.join(r, f)
                rp = os.path.join(rel, f) if rel != "." else f
                out.append(f"─── {rp} ───")
                if os.path.getsize(fp) <= 16384 and \
                        os.path.splitext(f)[1].lower() in TEXT_EXT:
                    out.append(open(fp, encoding="utf-8",
                                    errors="replace").read())
                else:
                    out.append(f"（{os.path.getsize(fp)}B,非文本或过大,只列不读）")
        return "\n".join(out[:800])
    ext = os.path.splitext(path)[1].lower()
    if ext in TEXT_EXT:
        return open(path, encoding="utf-8", errors="replace").read()
    if ext == ".xlsx":
        wb = _need_openpyxl().load_workbook(path, read_only=True, data_only=True)
        out = []
        for ws in wb.worksheets:
            out.append(f"[sheet] {ws.title}")
            for row in ws.iter_rows(values_only=True):
                cells = [norm(c) for c in row]
                if any(cells):
                    out.append("\t".join(cells))
        wb.close()
        return "\n".join(out)
    if ext == ".db":
        con = sqlite3.connect(path)
        out = [r[0] for r in con.execute(
            "SELECT sql FROM sqlite_master WHERE sql IS NOT NULL")]
        con.close()
        return "\n".join(out)
    if ext == ".docx" and shutil.which("textutil"):
        r = subprocess.run(["textutil", "-convert", "txt", "-stdout", path],
                           capture_output=True, text=True)
        if r.returncode == 0:
            return r.stdout
    return f"（二进制文件，{os.path.getsize(path)} 字节，sha256 {sha256(path)[:16]}…）"


def _part_index(task):
    """类知识索引（读桥第二腿：模板+边 → 按需加载，skill 同构）。
    ①**本类（模板）**——实例由类生，类的知识=实例的操作手册（用户裁决
    2026-08-19）：本类根 知识/*.md＋类.md＋schema.md；②**部件类**——沿
    本类 聚合/组合 边一跳，部件根 知识/＋类.md。确定性、零第二登记；
    已作显式输入弧的文件**去重**（全文已在 prompt）；关联/继承不级联。
    账记清单级「类知识索引」，不参与对账漂移——必读材料仍走显式弧
    （缺席不使能/变更待重审）。"""
    import yaml
    root = task["root"]
    declared = {r["path"] for r in (task.get("rows") or [])
                if r["kind"] == "输入" and r["path"]
                and os.path.isfile(r["path"])}
    kr = klass_rel(task["task_dir"], root)

    def _files(proot, with_schema=False):
        fs = []
        kdir = os.path.join(proot, "知识")
        if os.path.isdir(kdir):
            fs += [f for f in sorted(os.listdir(kdir))
                   if f.endswith(".md")
                   and os.path.join(kdir, f) not in declared]
        extras = ("类.md", "schema.md") if with_schema else ("类.md",)
        for ext in extras:
            p = os.path.join(proot, ext)
            if os.path.isfile(p) and p not in declared:
                fs.append(ext)
        return fs

    out = []
    cls_md = os.path.join(root, *kr.split("/"), "类.md") if kr else ""
    if kr:
        proot = os.path.join(root, *kr.split("/"))
        fs = _files(proot, with_schema=True)
        if fs:
            out.append({"类": kr.split("/")[-1], "种": "本类（模板）",
                        "路径": kr, "文件": fs})
    if kr and os.path.isfile(cls_md):
        m = re.match(r"^---\s*\n(.*?)\n---",
                     open(cls_md, encoding="utf-8").read(), re.S)
        rels = ((yaml.safe_load(m.group(1)) or {}).get("关系") or []) if m \
            else []
        for r in rels:
            if not isinstance(r, dict) \
                    or (r.get("种") or r.get("类别") or "关联") \
                    not in ("聚合", "组合"):
                continue
            d = (r.get("方向") or "").strip()
            parts = [p.strip() for p in re.split(r"[→←]", d) if p.strip()]
            owner = kr.split("/")[-1]
            target = None
            if len(parts) >= 2:
                a, b = parts[0], parts[-1]
                target = b if a == owner else (a if b == owner
                                               else (b if "→" in d else a))
            elif parts:
                target = parts[0]
            if not target or target == owner:
                continue
            cands = glob.glob(os.path.join(root, "类", target)) \
                + glob.glob(os.path.join(root, "域", "*", "类", target))
            if not cands:
                continue                               # 悬空由 lint 报，此处不越权
            fs = _files(cands[0])
            if fs:
                out.append({"类": target, "种": r.get("种") or r.get("类别"),
                            "路径": os.path.relpath(cands[0], root),
                            "文件": fs})
    return out


def build_prompt(task, outs, io_mode, dst_of=None):
    """io_mode='write'：给出绝对路径让引擎 agent 直写文件（默认，产物不经 stdout，
    免疫前缀污染/多产物切分）；'stdout'：产物内容回传 stdout，本引擎落盘（裸 LLM 用）。"""
    ins = [r for r in task["rows"] if r["kind"] == "输入"
           and (not r.get("optional")
                or (r["path"] and os.path.isfile(r["path"])))]   # 可缺弧未在场不进 prompt
    p = ["你是 I3DNA 目录体系的微任务执行程序。微任务=执行一个操作，完成一项任务：",
         "依据下列输入参数文件，生成产物文件。产物必须自包含、可直接落盘使用。",
         "若你的环境装有 i3dna-* skills，先读 i3dna-task-contract（执行契约）"
         "与 i3dna-literacy（体系识字课）并严格遵守。"]
    if io_mode == "write":
        p.append(f"本任务有 {len(outs)} 个产物文件。用你的文件写入工具把每个产物"
                 "**直接写到下面给出的绝对路径**（这是引擎的暂存区，验收通过后由引擎"
                 "原子落位到正树；暂存区为空，必须新建完整写出——即使你认为树里已有"
                 "旧版或任务已完成）。产物较长时**分多次写入**（先写前半，再用追加/"
                 "编辑工具续写），不要把超长内容塞进单次工具调用——会撞输出上限被截断。"
                 "写完后只回复一行：完成。不要把文件内容打到回复里，不要解释。")
        p.append("")
        me = os.path.basename(task["task_dir"])
        for r in outs:
            if is_field_area(task, r):
                tf = _type_file(task, r, "状态")
                owners = get_value(tf, "属主") if tf else None
                mine = ([k for k, v in owners.items() if str(v) == me]
                        if isinstance(owners, dict) else [])
                p.append(f"【字段区→定点写】{dst_of[r['pname']]}　（{r['desc']}）："
                         "多岗共享键值区，**禁止用文件写入工具整文件重写**"
                         "（会碾掉他岗的键）。需要更新时用 shell 定点写：\n"
                         f"  python3 {os.path.join(_KV_DIR, 'i3dna_kv.py')} set "
                         f"{dst_of[r['pname']]} <键> <值>\n"
                         f"你的属主键：{('、'.join(mine)) or '（无——不该写它）'}；"
                         "含空格/特殊字符的值要加引号；无需更新则不动它。")
            else:
                p.append(f"【产物→写到】{dst_of[r['pname']]}"
                         f"　（{r['desc']}，类型 {r['ptype'] or '未声明'}，版本 {r['pver'] or '-'}）")
    elif len(outs) > 1:
        marks = "、".join(SPLIT_MARK.format(name=r["pname"]) for r in outs)
        p.append(f"本任务有 {len(outs)} 个产物文件。每个文件之前必须单独一行输出分隔标记"
                 f"（严格照抄）：{marks}。标记行之外只输出文件内容本身，不要解释。")
    else:
        p.append("要求：只输出产物文件的完整内容本身——不要解释、不要 markdown 代码围栏、"
                 "不要输出文件名。第一行就是文件第一行。")
    if io_mode != "write":
        p.append("")
        for r in outs:
            p.append(f"【产物】{r['desc']}：{r['pname']}"
                     f"（类型 {r['ptype'] or '未声明'}，版本 {r['pver'] or '-'}）")
    p.append("")
    if task["instruction"]:
        p += ["===== 任务定义 =====", task["instruction"], ""]
    if io_mode == "write":
        idx = _part_index(task)
        if idx:
            p.append("===== 类知识索引（按需加载）=====")
            p.append("本类（模板）与整体◇部件类的知识材料。需要时用你的读"
                     "文件工具自行查看，用不到的不必读：")
            for it in idx:
                p.append(f"· {it['类']}（{it['种']}）→ {it['路径']}/"
                         f"{{{','.join(f[:-3] for f in it['文件'])}}}.md")
            p.append("")
    for r in ins:
        if r.get("_pick"):     # 主题车道（形状定律 8-21·工单2）：一火一单
            p.append(f"===== 输入：{r['desc']}（主题目录 {r['pname']}——"
                     f"本轮恰消费一张：{os.path.basename(r['_pick'])}；"
                     "队列里其余单下轮再来，不要自己删单——消费由引擎记账）=====")
            p.append(read_input(r["_pick"]))
        else:
            p.append(f"===== 输入：{r['desc']}（{r['pname']}）=====")
            p.append(read_input(r["path"]))
        p.append("")
    return "\n".join(p)


# ── 引擎调用 ─────────────────────────────────────────────────

# 验证动作的机械判据（符号匹配，零幻觉）：命中即记入 结果.json 的「验证动作」
VERIFY_PAT = re.compile(
    r"py_compile|pytest|unittest|test_|offscreen|QT_QPA|sqlite3|\.schema|"
    r"verify|COMPILE_OK|smoke", re.I)


def _pretty_event(obj, reply, trace=None):
    """OMP --mode json 事件 → 人读单行（💭思考 🔧工具 ↩结果 🗣回复）；
    trace 非空时同步记录过程摘要（工具调用清单＋思考段数）。"""
    if obj.get("type") != "message_end":
        return
    m = obj.get("message", {})
    role = m.get("role")
    for c in m.get("content", []):
        ct = c.get("type")
        if role == "assistant" and ct == "thinking":
            print("💭 " + " ".join(c.get("thinking", "").split())[:400], flush=True)
            if trace is not None:
                trace["思考段"] = trace.get("思考段", 0) + 1
        elif role == "assistant" and ct == "toolCall":
            args = json.dumps(c.get("arguments", {}), ensure_ascii=False)
            print(f"🔧 {c.get('name')} {args[:200]}", flush=True)
            if trace is not None:
                trace.setdefault("工具调用", []).append(
                    {"工具": c.get("name"), "参数摘要": args[:160]})
        elif role == "assistant" and ct == "text":
            reply.append(c.get("text", ""))
            print("🗣 " + " ".join(c.get("text", "").split())[:300], flush=True)
        elif role in ("toolResult", "tool") and ct == "text":
            print("↩ " + " ".join(c.get("text", "").split())[:200], flush=True)


class _AcpPrinter:
    """ACP 是 token 级 chunk 流——逐条打印会刷屏，按种类切换时聚合成
    与 --mode json 同一词汇表的人读单行（💭思考 🔧工具 ↩结果 🗣回复）。"""

    def __init__(self, stream, trace):
        self.kind, self.buf = None, []
        self.stream, self.trace = stream, trace

    def flush(self):
        text, self.buf = "".join(self.buf), []
        if not text.strip():
            return
        if self.kind == "thought":
            if self.stream:
                print("💭 " + " ".join(text.split())[:400], flush=True)
            if self.trace is not None:
                self.trace["思考段"] = self.trace.get("思考段", 0) + 1
        elif self.kind == "message" and self.stream:
            print("🗣 " + " ".join(text.split())[:300], flush=True)

    def feed(self, u, reply):
        k = u.get("sessionUpdate")
        if k == "agent_thought_chunk":
            if self.kind != "thought":
                self.flush()
                self.kind = "thought"
            self.buf.append(u.get("content", {}).get("text", ""))
        elif k == "agent_message_chunk":
            t = u.get("content", {}).get("text", "")
            reply.append(t)
            if self.kind != "message":
                self.flush()
                self.kind = "message"
            self.buf.append(t)
        elif k in ("tool_call", "tool_call_update"):
            self.flush()
            self.kind = None
            title = str(u.get("title") or u.get("kind") or "")
            if k == "tool_call":
                if self.stream:
                    print(f"🔧 {title[:200]}", flush=True)
                if self.trace is not None:
                    self.trace.setdefault("工具调用", []).append(
                        {"工具": u.get("kind") or "tool", "参数摘要": title[:160]})
            elif u.get("status") in ("completed", "failed") and self.stream:
                print(f"↩ {title[:160]} [{u.get('status')}]", flush=True)


def call_engine_acp(prompt, server_cmd, timeout, stream=False, trace=None):
    """ACP 车道：spawn 服务端（如 `omp --thinking high acp`），JSON-RPC over stdio。
    价值不在换模型在失败可读：stopReason 进过程摘要——max_tokens（思考烧光预算）、
    服务端断线、协议错各是各的死因，不再一律静默空返。"""
    import time
    toks = [os.path.expandvars(os.path.expanduser(a))
            for a in shlex.split(server_cmd)]
    p = subprocess.Popen(toks, stdin=subprocess.PIPE, stdout=subprocess.PIPE,
                         stderr=subprocess.DEVNULL, text=True,
                         encoding="utf-8", errors="replace", bufsize=1)
    reply, rid_box = [], [0]
    printer = _AcpPrinter(stream, trace)
    deadline = time.time() + timeout

    def send(obj):
        p.stdin.write(json.dumps(obj) + "\n")
        p.stdin.flush()

    def dispatch(msg):
        m = msg.get("method")
        if m == "session/update":
            printer.feed(msg.get("params", {}).get("update", {}), reply)
        elif m == "session/request_permission" and "id" in msg:
            opts = msg.get("params", {}).get("options", [])
            allow = next((o for o in opts if "allow" in str(o.get("kind", ""))),
                         opts[0] if opts else None)
            send({"jsonrpc": "2.0", "id": msg["id"], "result": {"outcome": {
                "outcome": "selected",
                "optionId": allow["optionId"] if allow else ""}}})
        elif "id" in msg and "method" in msg:       # 其它服务端请求：空成功
            send({"jsonrpc": "2.0", "id": msg["id"], "result": {}})

    def rpc(method, params):
        rid_box[0] += 1
        rid = rid_box[0]
        send({"jsonrpc": "2.0", "id": rid, "method": method, "params": params})
        while True:
            if time.time() > deadline:
                p.terminate()
                sys.exit(f"ACP 车道超时（{timeout}s）于 {method}")
            line = p.stdout.readline()
            if not line:
                p.wait()
                sys.exit(f"ACP 服务端中途退出（码 {p.returncode}）——"
                         "车道断线，与模型空返是两种病")
            line = line.strip()
            if not line:
                continue
            try:
                msg = json.loads(line)
            except json.JSONDecodeError:
                continue
            if msg.get("id") == rid and ("result" in msg or "error" in msg):
                if "error" in msg:
                    p.terminate()
                    sys.exit(f"ACP {method} 出错：{str(msg['error'])[:300]}")
                return msg["result"]
            dispatch(msg)

    try:
        rpc("initialize", {"protocolVersion": 1, "clientCapabilities": {
            "fs": {"readTextFile": False, "writeTextFile": False}}})
        sid = rpc("session/new", {"cwd": os.getcwd(), "mcpServers": []})["sessionId"]
        res = rpc("session/prompt", {"sessionId": sid,
                                     "prompt": [{"type": "text", "text": prompt}]})
    finally:
        try:
            p.terminate()
        except Exception:
            pass
    printer.flush()
    stop = res.get("stopReason", "?")
    if trace is not None:
        trace["停机原因"] = stop
    if stop != "end_turn":
        print(f"  ⚠ ACP stopReason={stop}"
              + ("（思考/输出烧光 token 预算——降思考档或换车道）"
                 if stop == "max_tokens" else ""), flush=True)
    return "".join(reply).strip()


# max_tokens 死是确定性死（思考烧光输出预算），原样重跑只会再死一遍——
# 降一档思考重试；降档车道与首跑死因都进账（过程摘要·降档重试）。
# glm-5.3 档位 low/high/max 全在列，梯子只配会烧穿的 high 以上档。
THINKING_LADDER = {"max": "high", "high": "low"}


def _degrade_thinking(engine):
    """降档映射：命中梯内档位 → (新车道, 旧档, 新档)；无可降返回 None。"""
    for hi, lo in THINKING_LADDER.items():
        old, new = f"--thinking {hi}", f"--thinking {lo}"
        if old in engine:
            return engine.replace(old, new), hi, lo
    return None


def call_engine(prompt, engine, timeout, stream=False, trace=None):
    if engine.startswith("acp:"):
        return call_engine_acp(prompt, engine[4:].strip(), timeout, stream, trace)
    fd, pfile = tempfile.mkstemp(suffix=".txt", prefix="i3dna_prompt_")
    try:
        with os.fdopen(fd, "w", encoding="utf-8") as f:
            f.write(prompt)
        # 无 shell 执行，逐 token 展开 ~ 与 $VAR，让
        # `env CLAUDE_CONFIG_DIR=~/.claude-glm claude -p` 这类合规组合直接可用
        toks = [os.path.expandvars(os.path.expanduser(a)) for a in shlex.split(engine)]
        if stream and toks and os.path.basename(toks[0]) == "omp" \
                and "--mode" not in engine:
            toks = toks[:1] + ["--mode", "json"] + toks[1:]   # 事件流
        if "{prompt_file}" in engine:
            cmd = [a.replace("{prompt_file}", pfile) for a in toks]
            stdin_text = None
        else:
            cmd, stdin_text = toks, prompt

        if stream:
            # 流式：逐行泵出——JSON 事件美化，普通行透传；超时交由调用方掐进程
            p = subprocess.Popen(cmd, stdin=subprocess.PIPE if stdin_text else None,
                                 stdout=subprocess.PIPE,
                                 stderr=subprocess.DEVNULL, text=True,
                                 encoding="utf-8", errors="replace")
            if stdin_text:
                p.stdin.write(stdin_text)
                p.stdin.close()
            reply, raw = [], []
            for line in p.stdout:
                line = line.rstrip("\n")
                if not line:
                    continue
                try:
                    _pretty_event(json.loads(line), reply, trace)
                except (json.JSONDecodeError, AttributeError):
                    raw.append(line)
                    print(line, flush=True)
            p.wait()
            if p.returncode != 0:
                sys.exit(f"引擎失败（{engine}）：退出码 {p.returncode}")
            out = ("\n".join(reply) or "\n".join(raw)).strip()
            os.unlink(pfile)
            return out

        r = subprocess.run(cmd, input=stdin_text, capture_output=True,
                           text=True, encoding="utf-8", errors="replace",
                           timeout=timeout)
    finally:
        if os.path.exists(pfile):
            os.unlink(pfile)
    if r.returncode != 0:
        sys.exit(f"引擎失败（{engine}）：{(r.stderr or r.stdout)[:500]}")
    out = r.stdout.strip()
    if out.startswith("```"):
        out = re.sub(r"^```[^\n]*\n", "", out)
        out = re.sub(r"\n```\s*$", "", out)
    # 前缀清洗（stdout 模式的兜底）：开头若是闲话，切到首个标题/围栏/产物标记
    if not re.match(r"^(#|```|<<<I3DNA|#!|\s*<)", out):
        m = re.search(r"^(#|```|<<<I3DNA|#!)", out, re.M)
        if m and m.start() > 0:
            out = out[m.start():]
    return out


def split_products(content, outs):
    """按分隔标记切分多产物；单产物直接整体返回。

    可缺产物（如返工单/审查单）块缺席＝**收回**，不是错误——
    存在性即信息，本次点火的输出就是产物在场性的全权声明
    （write 车道里 agent 直接删文件表达同一语义，stdout 车道靠块缺席表达）。
    必需产物缺块仍然拒绝落盘。"""
    if len(outs) == 1:
        return {outs[0]["pname"]: content}
    parts, marks = {}, list(SPLIT_RE.finditer(content))
    if not marks:
        sys.exit("多产物任务但输出中无 <<<I3DNA-产物:...>>> 分隔标记，拒绝落盘")
    for i, m in enumerate(marks):
        end = marks[i + 1].start() if i + 1 < len(marks) else len(content)
        parts[m.group(1).strip()] = content[m.end():end].strip() + "\n"
    missing = [r["pname"] for r in outs
               if r["pname"] not in parts and not r.get("optional")]
    if missing:
        sys.exit(f"多产物输出缺文件：{missing}，拒绝落盘")
    return parts


# ── 预检 / 索引登记 / 结果.json ──────────────────────────────

def preflight_rows(task):
    """结构化预检内核：逐行返回 (kind, desc, pname, state, ok)；
    末尾追加汇总元组 ('_总判', '', '', 提示或'', 全局ok)。UI 与 CLI 共用。"""
    out, ok, has_out = [], True, False
    for r in task["rows"]:
        if r["kind"] == "产物":
            has_out = True
            state = ("⚠ 产物落点在远端机器" if r["path"] is None else
                     "⚠ 产物名未定（名称=*）" if r["pname"] == "*" else
                     "（产物槽）已有旧产物" if os.path.isfile(r["path"]) else "（产物槽）空")
            row_ok = True
        elif r["path"] is None:
            state, row_ok = "✗ 机器绝对路径不可解析", False
            ok = False
        elif r.get("optional"):
            # 可缺弧：存在性即 token——缺席不阻使能，在场即消费
            row_ok = True
            state = ("（可缺弧）✓ 在场" if os.path.isfile(r["path"])
                     else "（可缺弧）未在场·静默")
        else:
            if os.path.isdir(r["path"]):
                row_ok, state = True, "✓ 存在"
            elif blank_slot(r["path"]):
                # 空槽=落点非内容:与 converge 使能同口径,直接点火同样拒空
                row_ok = False
                state = ("✗ 空槽（有落点无内容）"
                         if os.path.isfile(r["path"]) else "✗ 缺失")
            else:
                row_ok, state = True, "✓ 存在"
            ok = ok and row_ok
        out.append((r["kind"], r["desc"], r["pname"], state, row_ok))
    hint = "" if has_out else "⚠ 产物未声明（任务定义/纯索引族需 run --output 名称@锚目录）"
    out.append(("_总判", "", "", hint, ok and has_out))
    return out


def preflight(task, quiet=False):
    rows = preflight_rows(task)
    summary = rows[-1]
    if not quiet:
        for kind, desc, pname, state, _ in rows[:-1]:
            print(f"  [{kind}] {desc:<10} {pname}  {state}")
        if summary[3]:
            print(f"  {summary[3]}")
    return summary[4]


def register_in_index(target_dir, name, desc):
    hits = glob.glob(os.path.join(target_dir, "__*索引文件*.xlsx"))
    if hits:
        path = hits[0]
        wb = _need_openpyxl().load_workbook(path)
        ws = wb.active
        seqs = []
        for cells in data_rows(ws):
            if len(cells) > 2 and cells[2] == name:
                wb.close()
                return f"已登记（{os.path.basename(path)}），跳过"
            seqs.append(int(cells[0]))
        ws.append([str(max(seqs) + 1 if seqs else 0), desc, name, "1", "0", "*", "*"])
    else:
        base = "__" + os.path.basename(target_dir).lstrip("_") + "_索引文件.xlsx"
        path = os.path.join(target_dir, base)
        wb = _need_openpyxl().Workbook()
        ws = wb.active
        ws.title = INDEX_SHEET
        for row in (INDEX_HEADERS, INDEX_CONSTRAINT, INDEX_COORD):
            ws.append(row)
        ws.append(["0", desc, name, "1", "0", "*", "*"])
    wb.save(path)
    wb.close()
    return f"已追加登记 → {os.path.basename(path)}"


def manifest_item(path, root):
    return {"名称": relroot(path, root), "字节": os.path.getsize(path),
            "sha256": sha256(path)}


def write_result(rec_dir_, payload, root=None):
    """账写入已隔离底物（i3dna_store）：默认 json，包配置 xlsx。原子落位。"""
    return save_account(rec_dir_, payload, root=root or rec_dir_)


# ── 命令 ─────────────────────────────────────────────────────

def _run_symbolic(task, sandbox, timeout):
    """符号点火：跑 执行程序/主程序，不调 LLM。确定性、可重跑，直写产物槽。
    沙盒语义与联结点火一致（正树零副作用）：按输入清单拷副本树，cwd=沙盒。"""
    entry, root = task["exec_entry"], task["root"]
    # 实例模式：cwd=实例根——执行程序里的相对路径（测试/报告.md 等）
    # 自动落进本实例，M1 程序一行不改就实例无关
    if task.get("case"):
        kr = klass_rel(task["task_dir"], root)
        cwd = os.path.join(root, *case_rel(kr, task["case"]).split("/"))
    else:
        cwd = root
    if sandbox:
        for r in task["rows"]:
            if r["kind"] == "输入" and r["path"] and os.path.isfile(r["path"]):
                q = os.path.join(sandbox, relroot(r["path"], root))
                os.makedirs(os.path.dirname(q), exist_ok=True)
                shutil.copy2(r["path"], q)
        # 实例模式的 cwd 要落沙盒里的实例根，不是沙盒根——否则程序的
        # 相对读全部落空（8-11 评审 #10）
        cwd = os.path.join(sandbox, relroot(cwd, root)) if cwd != root \
            else sandbox
        os.makedirs(cwd, exist_ok=True)
    runner = [sys.executable] if entry.endswith(".py") else ["/bin/bash"]
    print(f"  符号点火：{os.path.basename(runner[0])} "
          f"{relroot(entry, root)}（cwd={'沙盒' if sandbox else '包根'}）")
    env = dict(os.environ)          # M2 取值运行时随点火注入：树里的主程序
    env["PYTHONPATH"] = os.pathsep.join(   # `from i3dna_kv import get_value` 即可，
        [_KV_DIR] + ([env["PYTHONPATH"]] if env.get("PYTHONPATH") else []))
    r = subprocess.run(runner + [entry], cwd=cwd, capture_output=True,
                       text=True, encoding="utf-8", errors="replace",
                       timeout=timeout, env=env)   # 不必自己搓正则读键
    out = (r.stdout or "") + (("\n" + r.stderr) if r.stderr.strip() else "")
    if out.strip():
        print("  │ " + out.strip().replace("\n", "\n  │ "))
    if r.returncode != 0:
        sys.exit(f"符号点火失败：主程序退出码 {r.returncode}")
    return out


def cmd_run(task, sandbox, engine, timeout, io_mode, run_id="",
            stream=False, intent=""):
    root = task["root"]
    print(f"微任务：{relroot(task['task_dir'], root)}（{task['kind']}族，树根 {root}）")
    if sealed_class(task):
        sys.exit("类已封存（类根 封存.md 在场）：点火被拒——立法通道已封，"
                 "接棒法见封存碑")
    if task.get("executor") == "人" or task.get("executor_declared") == "人":
        sys.exit("本方法执行者=人（人工工位）：请人工完成产物后用 backfill 办结入账，"
                 "引擎不代人点火（声明机制值不因 --executor 主体覆盖而失守）")
    if not preflight(task):
        sys.exit("预检失败：输入缺失/不可解析或产物未声明，拒绝执行")
    # 三车道同门（缺陷20·8-21）：fire 前置接使能条件（弧上逐一核，
    # fail-closed——取不到值一律停机）＋空夹门（门弧非空＝悬账在场，拒火）
    # ＋结账站盘点（结账: 真 的终点站，点火即打烊）。直火拒火（响亮，
    # 零副作用）；推进侧 _task_needs_fire 同一副门挂起（converge 不炸）。
    blocked = _cond_block(task)
    if blocked:
        sys.exit(blocked + "（零副作用，未点火）")
    _dq = [f"{r['pname']}（{len(_queue_tickets(r['path']))} 张）"
           for r in task["rows"]
           if r["kind"] == "输入" and r.get("drain") and r["path"]
           and _queue_tickets(r["path"])]
    if _dq:
        sys.exit("点火被空夹门拒绝：" + "、".join(_dq[:3])
                 + ("…" if len(_dq) > 3 else "")
                 + "——门弧（清空: 真）非空＝悬账在场，等回音核销"
                 "（零副作用，未点火）")
    if task.get("closing"):
        _bad = _closure_pending(task)
        if _bad:
            sys.exit("点火被结账门拒绝：案卷 " + str(task.get("case"))
                     + " 主题目录 "
                     + "、".join(f"「{_d}」还有 {_n} 张在途"
                                 for _d, _n in _bad[:3])
                     + "——终点站打烊全夹盘点，先消费/销单（零副作用，未点火）")

    def place(p):
        if not sandbox:
            return p
        q = os.path.join(sandbox, relroot(p, root))
        os.makedirs(os.path.dirname(q), exist_ok=True)
        return q

    outs = [r for r in task["rows"] if r["kind"] == "产物" and r["path"] is not None
            and r["pname"] != "*"]
    if not outs:
        sys.exit("产物不可写：全部产物行落点在远端机器或名称=*（未定名）。"
                 "用 --output 名称@锚目录 显式声明可写产物后重试。")

    # ── 主题车道（形状定律 8-21·工单2）──全树主题法一表，行为三件套用：
    # ①uuid 落位代起名 ②一火一单消费 ③选单告知 agent。零主题声明＝空表
    # ＝三件全休眠，行为与现状逐字节同（向后兼容律）。
    theme_laws = _theme_laws_global(root)
    # ①uuid 代起名：目录弧产物指向 命名: uuid 的主题目录 → 本火产物名＝
    # <种名>__<uuid4>.md（引擎代起，agent 只交内容不交名）；账/落位/清单
    # 全记实名。write 主道与 stdout 车道共用本处（dst_of 一份，两道同源）。
    for r in outs:
        tf = _uuid_named(task, r, theme_laws)
        if tf is not None:
            stem = os.path.splitext(os.path.basename(tf))[0]
            r["path"] = os.path.join(r["path"], f"{stem}__{_uuid.uuid4()}.md")
    # ③选单（消费前瞻）：目录弧输入指向主题目录 → 本火恰消费字典序最小
    # 一张；先把选定单喂给 agent（prompt 只见这一张，队列其余下轮再来）。
    # 提交点删除的也是它——若中途被别火偷吃，退而选当前最小（不变式：
    # 每火至多消费一张仍成立）。门弧（清空: 真）不入选（缺陷20）：门弧
    # 只读不消费——核销权归声明消费它的方法，别火吃了＝核销权旁落。
    for r in task["rows"]:
        if r["kind"] == "输入" and not r.get("drain") \
                and _theme_queue(task, r, theme_laws):
            tickets = _queue_tickets(r["path"])
            if tickets:
                r["_pick"] = os.path.join(r["path"], tickets[0])

    dst_of = {r["pname"]: place(r["path"]) for r in outs}
    for d in dst_of.values():
        os.makedirs(os.path.dirname(d), exist_ok=True)
    # write 模式＝暂存-验收-落位：agent 只写包根下的私有暂存区（同卷，rename 原子），
    # 验收通过后逐产物原子落位。失败零副作用；并发双跑后者原子覆盖，无锁不坏数据。
    stage = None
    agent_dst = dst_of
    if io_mode == "write" and not task.get("exec_entry"):
        stage = os.path.join(root, f".i3dna_stage_{os.getpid()}")
        os.makedirs(stage, exist_ok=True)
        # 暂存区镜像正式树结构：产物按弧的相对路径落子目录（代码/应用.py、测试/测试_应用.py）
        # 这样 LLM 在暂存区自测环境与正式树一致——import 路径错误在暂存区就会暴露
        fa_names = {r["pname"] for r in outs if is_field_area(task, r)}
        # 字段区不进暂存流水线：暂存从空文件起步、落位=整文件替换——
        # 恰是字段区要禁的动作。字段区由 agent 用 kv 命令定点直写正树。
        agent_dst = {n: (dst_of[n] if n in fa_names
                         else os.path.join(stage, relroot(p, root)))
                     for n, p in dst_of.items()}
        for d in agent_dst.values():
            os.makedirs(os.path.dirname(d), exist_ok=True)

    t0 = datetime.now().isoformat(timespec="seconds")
    # OCC 快照：t0 记输入清单（真实消费的版本），落账前复验漂移（不锁而验）
    in_manifest = []
    for r in task["rows"]:
        if r["kind"] == "输入" and r["path"] and os.path.isfile(r["path"]):
            it = manifest_item(r["path"], root)
            if r.get("optional") or is_message(task, r):
                it["可缺"] = True   # 生命周期件（可缺弧/消息）：缺席=已收回/已消费
            if is_state(task, r):
                it["状态"] = True   # 状态件：lint 漂移宽容
            if is_field_area(task, r):
                it["字段区"] = True  # 仅字段区享受「守门不催火」
            if store.is_fact_arc(r):
                it["事实"] = True   # 事实件弧：漂移语义见 _needs_fire（违规非过期）
            in_manifest.append(it)
    # 任务定义也是变换的输入（法）：记账不声明——改指令/改弧=账面变旧=重审。
    # 与执行程序同一逻辑；此前只有红任务的"法"（主程序）被追踪，蓝任务的漏了
    if task.get("def_path") and os.path.isfile(task["def_path"]):
        in_manifest.append(manifest_item(task["def_path"], root))

    def _drift():
        return [os.path.basename(m["名称"]) for m in in_manifest
                if os.path.isfile(os.path.join(root, m["名称"]))
                and sha256(os.path.join(root, m["名称"])) != m["sha256"]]

    field_snaps = _field_snap(task, outs)   # 属主执法快照（点火前）

    if task.get("exec_entry"):
        # ── 符号主义分支：执行体=确定性程序，账（OCC/结果/journal）与联结同套 ──
        # 程序也是变换的输入：记进输入清单，改程序=账面变旧=推进自动重点火
        for f in sorted(glob.glob(os.path.join(task["task_dir"], EXEC_DIR, "*"))):
            if os.path.isfile(f):
                in_manifest.append(manifest_item(f, root))
        stale = compile_stale(task["task_dir"], root)
        if stale:
            print(f"  ⚠ 编译过期：{','.join(stale)} 在编译后变过——"
                  "现行程序照跑；本引擎不自动重编译")
        content = _run_symbolic(task, sandbox, timeout)
        t1 = datetime.now().isoformat(timespec="seconds")
        drift = _drift()
        opt_names = {r["pname"] for r in outs if r.get("optional")}
        bad = [f"{n}（不存在或空）" for n, p in dst_of.items()
               if n not in opt_names and blank_slot(p)]
        if bad:
            sys.exit("符号点火验收失败（程序跑完但产物槽没长东西）："
                     + "；".join(bad))
        engine = f"符号执行：{relroot(task['exec_entry'], root)}"
        io_mode, trace = "符号直写", None
        if drift:
            print(f"  ⚠ OCC：点火期间输入被改（{','.join(drift)}）——记录标「出处存疑」，"
                  "下次推进将自动重算")
    else:
        prompt = build_prompt(task, outs, io_mode, agent_dst)
        print(f"  prompt {len(prompt)} 字符 → 引擎 `{engine}`（{io_mode} 模式）…")
        trace = {} if stream else None
        content = call_engine(prompt, engine, timeout, stream, trace)
        t1 = datetime.now().isoformat(timespec="seconds")
        drift = _drift()
        if drift:
            print(f"  ⚠ OCC：点火期间输入被改（{','.join(drift)}）——记录标「出处存疑」，"
                  "下次推进将自动重算")

    placed = set()                             # 本火实际落盘的产物名（104
    # 修订2：顺号消息槽在场≠本火开的——缺席火不得认领前火之单入产物清单）
    if task.get("exec_entry"):
        pass                                   # 符号分支已直写并验收
        placed = {r["pname"] for r in outs
                  if r["path"] and os.path.isfile(r["path"])}
    elif io_mode == "write":
        # 验收在暂存区：agent 自称写完不算数，逐个验真落地；过了才原子落位（提交点）。
        # 引擎空返/未写（模型瞬时抽风，8-06 两例同因）自动重试一次再判死。
        opt_names = {r["pname"] for r in outs if r.get("optional")}

        def _bad():
            return [f"{n}（不存在或空）" for n, p in agent_dst.items()
                    if n not in opt_names and blank_slot(p)]
        bad = _bad()
        if bad:
            # 先看病历再重试：首跑 stopReason=max_tokens＝确定性死，原样重跑
            # 白烧一遍——降思考档重试（车道带梯内档位才有得降，见
            # THINKING_LADDER）；降档车道与死因入过程摘要，账不替引擎隐过。
            first_stop = trace.get("停机原因") if trace else None
            deg = _degrade_thinking(engine) if first_stop == "max_tokens" else None
            if deg:
                low, hi, lo = deg
                print(f"  ⚠ 验收未过——首跑 stopReason=max_tokens（思考烧光输出"
                      f"预算），降档重试 --thinking {hi} → {lo}…", flush=True)
                trace["降档重试"] = {"车道": engine, "停机原因": first_stop}
                engine = low
            else:
                print(f"  ⚠ 验收未过（{'；'.join(bad)}）——疑似引擎空返，"
                      "自动重试一次…", flush=True)
            content = call_engine(prompt, engine, timeout, stream, trace)
            t1 = datetime.now().isoformat(timespec="seconds")
            bad = _bad()
        if bad:
            shutil.rmtree(stage, ignore_errors=True)
            sys.exit("直写验收失败（真槽未动，零副作用，已重试一次）："
                     + "；".join(bad)
                     + f"\n引擎回复（末 300 字）：{content[-300:]}")
        row_of = {r["pname"]: r for r in outs}
        for n, p in agent_dst.items():
            if p == dst_of[n]:
                continue      # 字段区定点直写正树，不经暂存落位
            if n in opt_names and not os.path.isfile(p):
                # 可缺产物未产出=收回（如审查单全通过、返工单全绿）。
                # 删旧目标，否则残留文件+不变的 sha256 会让下游误判新鲜。
                # 状态件缺席=本轮未更新，保持不动——stdout 车道早有此保护，
                # write 车道此前漏了（字段区差点被当收回删掉）。
                # 顺号消息除外（104修订2）：本次没开 ≠ 撤销前次——
                # 前火开给别家的单，消费者没吃到之前本火不撤。
                if _seq_enabled(task, row_of[n]):
                    continue
                if os.path.isfile(dst_of[n]) and not is_state(task, row_of[n]):
                    os.remove(dst_of[n])
                continue
            if _seq_enabled(task, row_of[n]) and os.path.exists(dst_of[n]):
                dst_of[n] = _next_seq(dst_of[n])   # 顺号：不覆盖前单，新量子
                row_of[n]["path"] = dst_of[n]      # 账/清单记实名（捕获前改）
            os.replace(p, dst_of[n])
            placed.add(n)
        shutil.rmtree(stage, ignore_errors=True)
    else:
        parts = split_products(content, outs)
        for r in outs:
            dst = dst_of[r["pname"]]
            if is_field_area(task, r):
                # 字段区无定点笔的车道不落盘——stdout 块是整文件重写，
                # 恰是字段区的禁手（8-11 评审 #3）。键值更新走 write 车道。
                print(f"  字段区 {r['pname']} 不经 stdout 车道落盘，保持不动")
                continue
            if r["pname"] not in parts:
                # 可缺产物块缺席：消息＝收回删除；状态＝本轮未更新，保持不动
                # （顺号消息除外，同 write 车道——104修订2）
                if _seq_enabled(task, r):
                    continue
                if os.path.isfile(dst) and not is_state(task, r):
                    os.remove(dst)
                    print(f"  收回 → {dst}（可缺消息本轮未开）")
                continue
            if _seq_enabled(task, r) and os.path.exists(dst):
                dst = _next_seq(dst)               # 顺号：不覆盖前单，新量子
                dst_of[r["pname"]] = dst           # 落位表同步（账记实名）
                r["path"] = dst
            tmp = f"{dst}.tmp.{os.getpid()}"   # pid 后缀：并发不撞暂存名（8-19）
            with open(tmp, "w", encoding="utf-8") as f:
                f.write(parts[r["pname"]])
            os.replace(tmp, dst)
            placed.add(r["pname"])

    products, opt_of, st_of, fa_of = [], {}, {}, {}
    for r in outs:
        dst = dst_of[r["pname"]]
        opt_of[dst] = bool(r.get("optional"))
        st_of[dst] = is_state(task, r)
        fa_of[dst] = is_field_area(task, r)
        if r.get("optional") and (not os.path.isfile(dst)
                                  or (r["pname"] not in placed
                                      and _seq_enabled(task, r))):
            continue                           # 可缺产物本轮未开（如全绿无
            # 返工单）；顺号槽在场≠本火开的——缺席火不认领前火之单
        tgt_dir = os.path.dirname(dst)
        if sandbox:
            src_hits = glob.glob(os.path.join(os.path.dirname(r["path"]),
                                              "__*索引文件*.xlsx"))
            if src_hits and not glob.glob(os.path.join(tgt_dir, "__*索引文件*.xlsx")):
                shutil.copy2(src_hits[0], tgt_dir)
        if task["kind"] == "frontmatter":
            note = "md 底物免登记（弧在 frontmatter，图归 wikilink/graph）"
        else:
            note = register_in_index(tgt_dir, r["pname"], r["desc"])
        print(f"  产物 → {dst}\n  索引：{note}")
        products.append(dst)
    # 键级属主执法(点火后) + 血缘附注(94号)——执法在前,血缘只记合法键
    if field_snaps:
        _field_guard(task, field_snaps)
        _append_lineage_for(task, field_snaps, t1)
    # 收盘盘点（§8.12）：目录输入弧记盘点单——记点火后的清单（消费即删
    # 不落永久漂移）；账从此说得出「点火时收件箱里有哪几张单」（死信有物证）。
    for r in task["rows"]:
        if r["kind"] == "输入" and r["path"] and os.path.isdir(r["path"]):
            it = {"名称": relroot(r["path"], root), "目录": True,
                  "清单": dir_manifest(r["path"])}
            if r.get("optional") or is_message(task, r):
                it["可缺"] = True
            in_manifest.append(it)
    # ②一火一单消费（形状定律 8-21·工单2）：目录弧输入指向主题目录 →
    # 本火恰消费一张（字典序最小，确定性可复算）。**盘点记「点火时队列里
    # 有哪几张」，消费清单记「哪一火吃了哪张」**——一账一单（缺陷4 消费
    # 零痕迹由此闭合）。删除权收归引擎（缺陷15 A 档）：选单、删除、入账
    # 同提交点，清单外的单不删、新到不误伤；agent 不再自己删单。盘点之后
    # 删＝消费单也在盘上（死信有物证）。
    消费清单 = []
    for r in ([] if sandbox else task["rows"]):
        if r["kind"] != "输入" or r.get("drain") \
                or not _theme_queue(task, r, theme_laws):
            continue          # 非主题目录＝老收件箱；门弧（清空）只读不消费
        pick = r.get("_pick")
        if pick is None or not os.path.isfile(pick):   # 中途被偷吃→退当前最小
            tickets = _queue_tickets(r["path"])
            if not tickets:
                continue
            pick = os.path.join(r["path"], tickets[0])
        消费清单.append(dict(manifest_item(pick, root), 消费=True))
        os.remove(pick)
        print(f"  消费一张 → {relroot(pick, root)}（一火一单；队列余 "
              f"{len(_queue_tickets(r['path']))} 张）")

    res_root = sandbox if sandbox else root
    payload = {"任务": relroot(task["task_dir"], root), "状态": "执行",
               "引擎": engine, "IO模式": io_mode, "开始": t0, "结束": t1,
               **({"执行者": task.get("executor")} if task.get("executor") else {}),
               **({"批次标识": run_id} if run_id else {}),
               **({"意图": intent} if intent else {}),
               **({"类知识索引": _part_index(task)} if _part_index(task) else {}),
               **({"过程摘要": trace,
                   "验证动作": [t for t in trace.get("工具调用", [])
                                if VERIFY_PAT.search(t["参数摘要"])
                                or VERIFY_PAT.search(t["工具"] or "")]}
                  if trace else {}),
               "输入清单": in_manifest,
               **({"消费清单": 消费清单} if 消费清单 else {}),
               **({"出处存疑": drift} if drift else {}),
               "产物清单": [dict(manifest_item(p, res_root),
                                 **({"可缺": True} if opt_of.get(p) else {}),
                                 **({"状态": True} if st_of.get(p) else {}),
                                 **({"字段区": True} if fa_of.get(p) else {}))
                             for p in products],
               "备注": "沙盒执行，产物未入正树" if sandbox else ""}
    out = write_result(place(task["task_dir"]) if sandbox else rec_dir(task),
                     payload, root=task["root"])
    print(f"  点火记录 → {out}")
    # git 事件层：每次真树点火 = 一条 journal 提交（fail-soft，无仓则跳过）
    if not sandbox and os.path.isdir(os.path.join(root, ".git")):
        msg = (f"点火 {relroot(task['task_dir'], root)}"
               + (f" 批次 {run_id}" if run_id else ""))
        try:
            subprocess.run(["git", "-C", root, "add", "-A"],
                           capture_output=True, timeout=30)
            r = subprocess.run(["git", "-C", root, "commit", "-q", "-m", msg],
                               capture_output=True, text=True, timeout=30)
            print("  journal → git 提交" if r.returncode == 0
                  else "  journal：无变更或提交失败（不影响点火）")
        except Exception as e:
            print(f"  journal：git 异常已忽略（{e}）")


def cmd_backfill(task, note, intent=""):
    root = task["root"]
    # 封存门（8-20 立法口归一配套）：封存类旧法不受理新案卷（零副作用）。
    if sealed_class(task):
        sys.exit("类已封存（类根 封存.md 在场）：办结被拒（零副作用，未入账）"
                 "——封存类旧法不再受理新案卷，接棒法见封存碑")
    # 办结校验门（skill 带脚本同款）：任务.md frontmatter 声明「校验: <路径>」
    # → 入账前代跑（argv=树根,案卷号；cwd=树根），非零退出＝拒绝办结
    # （零副作用，未写账）。脚本通常住元知识类——树的自检是树的知识。
    gate = get_value(os.path.join(task["task_dir"], "任务.md"), "校验")
    if gate:
        gp = os.path.join(root, str(gate))
        if not os.path.isfile(gp):
            sys.exit(f"校验程序不在场：{gate}——声明了校验就得有，办结拒绝")
        r = subprocess.run(
            [sys.executable, gp, root, task.get("case") or ""],
            cwd=root, capture_output=True, text=True, timeout=60)
        if r.returncode != 0:
            sys.exit("办结被校验程序拒绝（零副作用，未入账）：\n"
                     + (r.stderr.strip() or r.stdout.strip()))
        if r.stdout.strip():
            print(r.stdout.strip())
    # 三车道同门（缺陷20·8-21）：settle 同接使能条件——弧上条件逐一核
    # （fail-closed，fire 同一副门）。审批矩阵（金额路由门）由此有牙：
    # 12800 的单办结经理审批（门 <5000）＝越权，拒（零副作用）。
    blocked = _cond_block(task)
    if blocked:
        sys.exit(blocked + "（零副作用，未入账）")
    # 办结查使能（形状定律 8-21·工单4，闭缺陷18c）：必选输入缺席 → 拒办结
    # ——追认通道不再无牙（fail-closed）。文件弧：空槽/不在场；机器绝对
    # 路径：不可解析；目录弧：夹本身不在场（余单是下轮燃料不拦，空夹是
    # 回音语义不拦）。可缺输入照旧（缺席＝已收回/未到）。
    _absent = []
    for r in task["rows"]:
        if r["kind"] != "输入" or r.get("optional"):
            continue
        if r["path"] is None:
            _absent.append(f"{r['pname']}（不可解析）")
        elif not os.path.exists(r["path"]):
            _absent.append(relroot(r["path"], root) if os.path.isabs(r["path"])
                           else f"{r['pdir']}/{r['pname']}")
        elif os.path.isfile(r["path"]) and blank_slot(r["path"]):
            _absent.append(f"{r['pname']}（空槽）")
    if _absent:
        sys.exit("办结被使能门拒绝：必选输入缺席 " + "、".join(_absent[:5])
                 + "——先补齐输入或改声明可缺（零副作用，未入账）")
    # 办结悬账门（8-19 符合性审计落地）：结账＝对账点（saga commit）——
    # 本案卷在途消息未清零＝守恒破坏，拒绝办结（零副作用）。与
    # lint_case_closure 同语义的第二实现（互不相识纪律：引擎＝机制闸，
    # lint＝对账）；树上无 消息/*.md 法定路径时检查自然休眠。
    _case = task.get("case")
    if _case:
        _laws = {}
        for _tf in glob.glob(os.path.join(root, "**", "消息", "*.md"),
                             recursive=True):
            _law = get_value(_tf, "路径")
            if _law:
                _laws.setdefault(
                    os.path.splitext(os.path.basename(_tf))[0],
                    _law.replace("{案卷号}", "*").replace("{实例}", "*"))
        if _laws:
            _crel = case_rel(klass_rel(task["task_dir"], root), _case)
            for _名, _pat in _laws.items():
                for _h in glob.glob(os.path.join(root, _pat)):
                    _hrel = os.path.relpath(_h, root)
                    _stem = os.path.splitext(os.path.basename(_h))[0]
                    _in_case = _hrel.startswith(_crel + os.sep) \
                        or any(_s == _case for _s in _stem.split("__"))
                    if _in_case:
                        sys.exit(
                            f"办结被悬账门拒绝：案卷 {_case} 在途消息"
                            f"「{_名}」在场（{_hrel}）——开出的承诺未兑现"
                            "或回执未销账；先消费/销单或撤销结账"
                            "（零副作用，未入账）")
        # 悬账门·主题目录（形状定律 8-21·工单2；缺陷19 收窄 8-21）：普通
        # 办结对主题目录休眠——结账不再是每次办结的事，是结账站的事
        # （A2 后审批夹/知会夹各一单、两站互等＝环形等待，报销003 实证；
        # 消费动作本身即清偿，非消费目录不归本站管）。结账站（任务卡
        # 结账: 真）：打烊全夹盘点（_closure_pending）——lint_case_closure
        # 同语义第二实现，互为见证。
        if task.get("closing"):
            _bad = _closure_pending(task)
            if _bad:
                sys.exit("办结被结账门拒绝：案卷 " + _case + " 主题目录 "
                         + "、".join(f"「{_d}」还有 {_n} 张在途"
                                     for _d, _n in _bad[:3])
                         + "——终点站打烊全夹盘点；先消费/销单"
                         "（零副作用，未入账）")
    # 回收弧（产物带 回收: 真）：办结=撤销——目标在场才可回收（先记
    # 旧 sha256 入账再删，git 历史留尸）；缺席=回收空气，响亮拒绝。
    # 空壳随扫（8-20 用户实证「撤了域树上还在」）：只删文件会留空目录
    # 残壳——树面即目录，残壳渲染成空域节点还触发空域悬空 lint。
    回收清单 = []
    for r in task["rows"]:
        if r["kind"] != "产物" or not r.get("retract") or not r["path"]:
            continue
        if not os.path.isfile(r["path"]):
            sys.exit(f"回收弧目标不在场：{relroot(r['path'], root)}"
                     "——不能回收空气，办结拒绝（零副作用）")
        回收清单.append(dict(manifest_item(r["path"], root), 回收=True))
        os.remove(r["path"])
        _prune_empty(os.path.dirname(r["path"]), root)
    # 草稿转正（形状定律 8-21·工单4，闭缺陷18b）：案卷 __草稿/<产物名>
    # 在途稿 → 办结落位成正式产物（sha 入账）。稿住 `__` 前缀区＝不是
    # 产物、不使能下游、账不进账——「先写草稿、办结转正」收编账外直写。
    # 回音: 无 类型＝**收讫两件套**（核销＋入账），无转正——fire-and-
    # forget：单到了账清了就完，没有回执要等。
    _laws = _theme_laws_global(root)
    收讫 = False
    for r in task["rows"]:
        tf = _theme_queue(task, r, _laws) if r["kind"] == "输入" else None
        if tf is not None and str(get_value(tf, "回音") or "").strip() == "无":
            收讫 = True
    if not 收讫 and task.get("case"):
        cdir0 = os.path.join(
            root, *case_rel(klass_rel(task["task_dir"], root),
                            task["case"]).split("/"))
        for r in task["rows"]:
            if r["kind"] != "产物" or not r["path"] \
                    or os.path.splitext(r["pname"])[1] == "":
                continue
            dp = os.path.join(cdir0, "__草稿", os.path.basename(r["path"]))
            if os.path.isfile(dp):
                os.makedirs(os.path.dirname(r["path"]), exist_ok=True)
                os.replace(dp, r["path"])
                print(f"  草稿转正 → {relroot(r['path'], root)}"
                      "（在途稿落位，sha 入账）")
    ins, missing, prods = [], [], []
    for r in task["rows"]:
        if r["path"] is None or not os.path.exists(r["path"]):
            missing.append({"名称": f"{r['pdir']} :: {r['pname']}",
                            "原因": "机器绝对路径不可解析" if r["path"] is None
                                    else "文件不存在"})
        elif r["kind"] == "输入":
            if os.path.isfile(r["path"]):
                ins.append(dict(manifest_item(r["path"], root),
                                **({"可缺": True}
                                   if r.get("optional") or is_message(task, r)
                                   else {})))
            elif os.path.isdir(r["path"]):          # 盘点单与点火同形（§8.12）
                ins.append({"名称": relroot(r["path"], root), "目录": True,
                            "清单": dir_manifest(r["path"])})
        elif os.path.isfile(r["path"]):
            prods.append(dict(manifest_item(r["path"], root),
                              **({"可缺": True}
                                 if r.get("optional") or is_message(task, r)
                                 else {})))
    if task.get("def_path") and os.path.isfile(task["def_path"]):
        ins.append(manifest_item(task["def_path"], root))   # 法入账，与点火对齐
    # 一火一单消费·办结侧（形状定律 8-21·工单2）：人工工位（话语即签字）
    # 与引擎车道同律——恰消费字典序最小一张，选单、删除、入账同提交点。
    # 盘点（上方 ins）记办结时队列，消费清单记本结吃掉哪张；agent/人不再
    # 自己删单（缺陷15 A 档：删除权收归引擎）。
    消费清单 = []
    for r in task["rows"]:
        if r["kind"] != "输入" or r.get("drain") \
                or not _theme_queue(task, r, _laws):
            continue          # 门弧（清空: 真）只读不消费——核销权归消费方
        tickets = _queue_tickets(r["path"])
        if not tickets:
            continue
        pick = os.path.join(r["path"], tickets[0])
        消费清单.append(dict(manifest_item(pick, root), 消费=True))
        os.remove(pick)
        print(f"  消费一张 → {relroot(pick, root)}（一火一单；队列余 "
              f"{len(_queue_tickets(r['path']))} 张）")
    # 一案卷一手术·机械面（8-20 对抗验收）：同一方法对**同一产物文件**
    # 双案卷办结＝「后账覆盖」洗账面（他案卷 settle 把本案卷产物的漂移
    # 洗成已覆盖）。重审同文件走改弧案卷/换目标，不办结第二遍（零副作用）。
    if task.get("case"):
        mine = {relroot(r["path"], root) for r in task["rows"]
                if r["kind"] == "产物" and r["path"]}
        rd = rec_dir(task)          # 实例/<类>/<案卷>/__账/<方法>
        shelf = os.path.dirname(os.path.dirname(os.path.dirname(rd)))
        for other in glob.glob(os.path.join(
                shelf, "*", "__账", os.path.basename(rd), "__结果.json")):
            if os.path.dirname(other) == rd:
                continue                      # 本案卷旧账（末火覆盖是既有语义）
            try:
                orec = json.load(open(other, encoding="utf-8"))
            except (json.JSONDecodeError, OSError, ValueError):
                continue
            dup = mine & {it.get("名称") for it in orec.get("产物清单", [])}
            if dup:
                ocase = os.path.basename(
                    os.path.dirname(os.path.dirname(os.path.dirname(other))))
                sys.exit(f"办结被一案卷一手术门拒绝：产物 {sorted(dup)} 已由"
                         f"同方法他案卷（{ocase}）办结——重审走改弧案卷，"
                         "不洗第二遍（零副作用，未入账）")
    payload = {"任务": relroot(task["task_dir"], root), "状态": "事后追认",
               "引擎": "图外执行，本记录由执行引擎回填",
               **({"执行者": task.get("executor")}
                  if task.get("executor") else {}),
               "回填时间": datetime.now().isoformat(timespec="seconds"),
               "输入清单": ins, "产物清单": prods}
    if missing:
        payload["缺失输入或产物"] = missing
    if 消费清单:
        payload["消费清单"] = 消费清单
    if 回收清单:
        payload["回收清单"] = 回收清单
    if note:
        payload["备注"] = note
    if intent:
        payload["意图"] = intent
    out = write_result(rec_dir(task), payload, root=task["root"])
    print(f"回填 → {out}（输入 {len(ins)}，产物 {len(prods)}，缺失 {len(missing)}）")
    # git 事件层：办结也是事件（fail-soft）——绿任务手术（女娲七法等）走
    # settle，不留 journal＝改系统的事件层断在半路（8-19 补，对齐点火）
    if os.path.isdir(os.path.join(root, ".git")):
        msg = (f"办结 {relroot(task['task_dir'], root)}"
               + (f"（{note}）" if note else ""))
        try:
            subprocess.run(["git", "-C", root, "add", "-A"],
                           capture_output=True, timeout=30)
            r = subprocess.run(["git", "-C", root, "commit", "-q", "-m", msg],
                               capture_output=True, text=True, timeout=30)
            print("  journal → git 提交" if r.returncode == 0
                  else "  journal：无变更或提交失败（不影响办结）")
        except Exception as e:
            print(f"  journal：git 异常已忽略（{e}）")


def _accounted_products(root):
    """全树已入账产物名集合（任何 __结果.json 的产物清单）＋宪法时刻
    基线路径集合。draft 落位只许写**空槽**：已入账/已立案的产物变更走
    正式通道（改弧/回收弧）——起草面不得覆写活产物（8-20 对抗验收：
    覆写已办结产物＋再办结＝洗账；覆写基线法文件＝无审批改法）。"""
    acc = set()
    for j in glob.glob(os.path.join(root, "**", "__结果.json"), recursive=True):
        try:
            data = json.load(open(j, encoding="utf-8"))
        except (json.JSONDecodeError, OSError, UnicodeDecodeError):
            continue
        acc |= {it.get("名称") for it in data.get("产物清单", [])
                if it.get("名称")}
    for pat in (os.path.join(root, "域", "*", "类", "女娲", "知识",
                             "宪法时刻.md"),
                os.path.join(root, "类", "女娲", "知识", "宪法时刻.md"),
                os.path.join(root, "知识", "宪法时刻-领域面.md")):
        for bp in glob.glob(pat):
            lines = open(bp, encoding="utf-8",
                         errors="replace").read().splitlines()
            for ln in lines:
                s = ln.strip()
                if s.startswith("|") and not set(s) <= {"|", "-", " ", ":"}:
                    cells = [c.strip() for c in s.strip("|").split("|")]
                    if cells:
                        acc.add(cells[0])
    return acc


def _inside(tp, base):
    """物理包含性（realpath——防案卷内预置符号链接把草稿写出界：
    词法 commonpath 不解析链接，8-20 对抗验收实证可逃逸）。"""
    rp = os.path.realpath(tp)
    rb = os.path.realpath(base)
    return rp == rb or rp.startswith(rb + os.sep)


def _msg_type_names(root):
    """消息类型名收集（draft 案卷材料门·工单108 F3）：**/消息/*.md 文件名
    去扩展、无 frontmatter 的正文型说明不算——与 lint 的 _message_type_names
    判据同源、各自实现（独立见证人惯例：两边独立读盘，谁漂移谁露馅）。"""
    names = set()
    for p in glob.glob(os.path.join(root, "**", "消息", "*.md"),
                       recursive=True):
        try:
            with open(p, encoding="utf-8", errors="replace") as f:
                head = f.read(4096)
        except OSError:
            continue
        if re.match(r"^---\s*\n.*?\n---", head, re.S):
            names.add(os.path.splitext(os.path.basename(p))[0])
    return names


def _hits_msg_type(name, msg_types):
    """文件名去 __ 段逐层剥后缀寻种（与 _type_file 同款）——命中消息
    类型名＝单据名（工单108 F3：一张草稿不得伪造 返工单.md 这类槽文件）。"""
    stem = os.path.splitext(name)[0]
    parts = stem.split("__")
    return any("__".join(parts[:i]) in msg_types
               for i in range(len(parts), 0, -1))


def cmd_draft(task, 草稿):
    """写桥 draft（101号 起草车道·103号 审批入图：柜员的手）——草稿
    落**案卷**、零入账、journal 留尸（8-20 裁定②：尸骸即展示-落位绑定
    证据，模型批准步换稿则两笔 journal 两具尸、git diff 即铁证；settle
    才入账）。两类落点：
    · 案卷材料：路径=案卷内相对路径（域意.md/申请要点…）；
    · 产物槽落位：路径=树内全路径，须恰为本任务（本案卷代入后）的
      某条产物弧且是**空槽**（不在任何账面产物清单、不在宪法时刻基线）
      ——审批信封「[draft 落位, settle 办结]」的前半步；已入账产物/
      基线法文件的变更走改弧/回收弧（正式通道），起草面不碰活产物。
    源=案卷内已有草稿（内容抄录，免模型重打全文）。封存门同
    fire/settle；起草零入账（账＝审计红线，draft 不碰）。"""
    root = task["root"]
    if sealed_class(task):
        sys.exit("类已封存（类根 封存.md 在场）：起草被拒（零副作用）"
                 "——封存类旧法不再受理新案卷，接棒法见封存碑")
    case = task.get("case")
    if not case:
        sys.exit("draft 须给 --case 案卷号（草稿落案卷——类级任务无案卷）")
    kr = klass_rel(task["task_dir"], root)
    cdir = os.path.join(root, *case_rel(kr, case).split("/"))
    slots = {relroot(r["path"], root) for r in task["rows"]
             if r["kind"] == "产物" and r["path"] and r["pname"] != "*"}
    frozen = _accounted_products(root)          # 活产物∪基线法——落位禁区
    msg_types = _msg_type_names(root)           # 案卷材料槽家族门（F3）
    theme_laws = _theme_laws_global(root)       # 主题门（形状定律·工单1号）：
                                                # 主题目录里的乱名也是单
    written = []                                # journal 留尸范围（工单109 F2②）
    if not isinstance(草稿, list) or not 草稿:
        sys.exit("draft 载荷须为非空草稿数组（stdin JSON）")
    if len(草稿) > 20:
        sys.exit(f"草稿数 {len(草稿)} 超上限 20——一案一次起草别贪多")
    for d in 草稿:
        if not isinstance(d, dict):
            sys.exit(f"草稿条目不是对象：{d!r}")
        raw = d.get("路径")
        if not isinstance(raw, str) or not raw.strip():
            sys.exit(f"草稿「路径」不是非空字符串：{raw!r}")
        rel = raw.replace("\\", "/")
        segs = [s for s in rel.split("/") if s not in ("", ".")]
        if "\\" in raw or rel.startswith("/") \
                or any(s in ("", "..") or s.startswith((".", "__"))
                       for s in rel.split("/")):
            sys.exit(f"草稿路径不干净：{raw!r}"
                     "（无 .. 无点下划线头段，不越案卷）")
        src = d.get("源")
        if src is not None:
            if not isinstance(src, str) or "/" in src or src.startswith(".") \
                    or not os.path.isfile(os.path.join(cdir, src)):
                sys.exit(f"草稿源不在案卷：{src!r}——落位抄录须案卷内在场文件")
            sp = os.path.join(cdir, src)
            if not _inside(sp, cdir):
                sys.exit(f"草稿源越出案卷（符号链接？）：{src!r}")
            content = open(sp, encoding="utf-8", errors="replace").read()
        elif isinstance(d.get("内容"), str):
            content = d["内容"]
        else:
            sys.exit(f"草稿条目缺「内容」或「源」：{rel}")
        norm = "/".join(segs)
        treeish = segs[0] in TREE_TOPS         # 首段=树顶命名空间→按树目标判
        if treeish and norm not in slots:
            sys.exit(f"草稿路径是树内路径但不是本任务产物槽：{norm}"
                     "——落位只许产物弧（防起草面写出任意树文件）")
        if norm in slots:                     # 产物槽落位（审批前半步）
            if norm in frozen:
                sys.exit(f"落位目标已入账/已立案（活产物或基线法）：{norm}"
                         "——变更走改弧/回收弧（正式通道），起草面不碰")
            tp = os.path.join(root, *norm.split("/"))
            if os.path.exists(tp) and not _inside(tp, tp):
                sys.exit(f"落位目标是符号链接：{norm}——拒绝跟随")
            note = "（覆盖未入账旧稿）" if os.path.exists(tp) \
                else "（产物槽落位）"
        else:                                  # 案卷材料
            tp0 = os.path.join(cdir, *segs)
            _drel = os.path.dirname(relroot(tp0, root)).replace(os.sep, "/")
            if any(s in ("消息", "状态") for s in segs) \
                    or _hits_msg_type(segs[-1], msg_types) \
                    or _theme_hit(_drel, theme_laws, case,
                                  case_rel(kr, case), kr) is not None:
                sys.exit(f"案卷材料不得落槽家族（消息/状态）——那是单据"
                         f"不是材料：{norm}（工单108 F3：伪单在场欺骗使能，"
                         "结账被悬账门拦死；或落进主题目录——目录即类型，"
                         "乱名也是单，形状定律 8-21）")
            tp = os.path.join(cdir, *segs)
            if not _inside(tp, cdir):
                sys.exit(f"草稿路径越出案卷（含符号链接逃逸）：{raw!r}")
            note = "（案卷材料）"
        os.makedirs(os.path.dirname(tp), exist_ok=True)
        tmp = f"{tp}.tmp.{os.getpid()}"        # §8.1 同型修法（工单108 F1）：
        with open(tmp, "w", encoding="utf-8") as f:   # pid 暂存＋原子换入，
            f.write(content)                   # 写一半崩不留半截槽文件——
        os.replace(tmp, tp)                    # 「人过目」看的不许是撕裂盘
        print(f"  草稿 → {relroot(tp, root)} {note}")
        written.append(relroot(tp, root))
    # journal 留尸（8-20 裁定②，工单109）：draft 零账但留 git 尸——尸骸即
    # 展示-落位绑定证据；scoped add 只加本次草稿，不代别人提交在途文件。
    _journal(root, f"起草 {relroot(task['task_dir'], root)} 案卷 {case}",
             paths=written)


def cmd_login(root, principal, status):
    """登录事件入账（树原生登录日志——login.db 的 登录日志 搬进 __日志/）。
    主体值=实例/人员/<k>；一笔一文件：__日志/登录_<ts>_<账号>.log。"""
    root = os.path.abspath(root)
    ts = datetime.now().strftime("%Y%m%d-%H%M%S")
    d = os.path.join(root, "__日志")
    os.makedirs(d, exist_ok=True)
    p = os.path.join(d, f"登录_{ts}_{os.path.basename(principal)}.log")
    with open(p, "w", encoding="utf-8") as f:
        f.write(f"{datetime.now().isoformat(timespec='seconds')} "
                f"账号={principal} 终端=explorer 状态={status}\n")
    print(f"登录日志 → {os.path.relpath(p, root)}")


def _prune_empty(d, root):
    """空目录链清扫（回收弧配套，8-20 用户实证「撤了域树上还在」）：文件
    已删，空父目录逐级收走——树面即目录面，残壳会渲染成空域节点并触发
    空域悬空 lint。只收真空目录（非空即止＝「只吃空域」的机械面），不越
    树根。"""
    rr = os.path.realpath(root)
    while d:
        try:
            if os.path.realpath(d) == rr:
                return
            os.rmdir(d)                 # 非空目录抛 OSError → 止
        except OSError:
            return
        d = os.path.dirname(d)


def _journal(root, msg, paths=None):
    """git 事件层一笔提交（fail-soft，无仓则跳过）——账本即档案柜。
    paths 给定＝scoped add（工单109 起草面：draft 是局部动作，不代别人
    提交在途文件）；缺省全树 -A（fire 语义不变）。"""
    if not os.path.isdir(os.path.join(root, ".git")):
        return
    try:
        subprocess.run(["git", "-C", root, "add"]
                       + (list(paths) if paths else ["-A"]),
                       capture_output=True, timeout=30)
        r = subprocess.run(["git", "-C", root, "commit", "-q", "-m", msg],
                           capture_output=True, text=True, timeout=30)
        print("  journal → git 提交" if r.returncode == 0
              else "  journal：无变更或提交失败")
    except Exception as e:
        print(f"  journal：git 异常已忽略（{e}）")


TEXT_EXT = {".md", ".txt", ".py", ".sh", ".yaml", ".yml", ".json", ".csv", ".j2"}


def _arc_material(task, head_chars=1500):
    """弧现物摘录（文本类头部）——给判官/编译 agent 当理解材料。"""
    root, chunks = task["root"], []
    for r in task["rows"]:
        p = r["path"]
        if not p or not os.path.isfile(p):
            continue
        rel = relroot(p, root)
        if os.path.splitext(p)[1].lower() in TEXT_EXT:
            body = open(p, encoding="utf-8", errors="replace").read()
            cut = body[:head_chars]
            more = f"\n……（共 {len(body)} 字符，以上为头部）" if len(body) > head_chars else ""
            chunks.append(f"◆ {r['kind']}「{r['desc']}」 {rel}\n{cut}{more}")
        else:
            chunks.append(f"◆ {r['kind']}「{r['desc']}」 {rel}"
                          f"（二进制，{os.path.getsize(p)} 字节，不摘录）")
    return "\n\n".join(chunks)


def _arc_lines(task, kind):
    root = task["root"]
    return "\n".join(f"- {r['desc']} → {relroot(r['path'], root)}"
                     for r in task["rows"]
                     if r["kind"] == kind and r["path"] is not None
                     and not r["desc"].startswith("↳"))


def cmd_detect(task, engine, timeout, stream):
    """检测可符号化：LLM 判官报告只投影不落树。重心=拆分重构建议——
    调用方拿反馈去重构微任务让它可符号化（渐进式信息化的循环本体）。"""
    tdir, root = task["task_dir"], task["root"]
    hist = "（无点火记录）"
    j = load_account(rec_dir(task), root)
    if j:
        try:
            hist = (f"最近点火 {j.get('开始','?')}→{j.get('结束','?')}，"
                    f"引擎 {j.get('引擎','?')}，产物 "
                    + "、".join(m["名称"] for m in j.get("产物清单", [])))
        except Exception:
            hist = "（点火记录不可解析）"
    prompt = f"""你是渐进式信息化顾问。评估下面这个微任务的可符号化程度：它的变换能否由一段确定性程序（无 LLM、无网络、无随机）完成。

判据（刀口）：能机械化的是符号同一性——格式转换、编译、校验、复制加适配、渲染、固定 schema 的抽取；不能机械化的是语义指称与言语行为——判断、设计、散文写作、需要理解意图的抽取。

【任务指令】
{task['instruction'] or '（本族格式无自由文本指令，变换语义由弧与参数表承载）'}

【输入弧】
{_arc_lines(task, '输入') or '（无）'}

【产物弧】
{_arc_lines(task, '产物') or '（无）'}

【历届点火】{hist}

【弧现物摘录】
{_arc_material(task)}

要求（纯文本报告，不要写任何文件，不要改树内任何东西）：
1. 判决：可符号化 / 部分可符号化 / 不可符号化
2. 理由——对着上面的刀口说，指认变换里哪些环节是符号同一性、哪些是语义判断
3. 拆分重构建议（重心）：哪一半机械可转符号程序、哪一半语义该留给 LLM、微任务与弧应该怎么切分让机械的那半独立成任务。若整体可符号化，给出程序轮廓（输入怎么读、变换怎么做、产物怎么写）。"""
    print(f"微任务：{relroot(tdir, root)}（检测可符号化，引擎 `{engine}`）")
    print(f"  prompt {len(prompt)} 字符 …")
    trace = {} if stream else None
    report = call_engine(prompt, engine, timeout, stream, trace)
    print("\n═ 可符号化检测报告（只投影，不落树）═\n" + report.strip())


def cmd_inform(task, engine, timeout, stream):
    """检测可信息化（绿→蓝）：人工工位的变换能否交给 LLM agent。
    与 cmd_detect（蓝→红）同构的判官——报告只投影不落树。
    三色谱：绿(人)→蓝(LLM)=信息化，蓝→红(程序)=符号化；
    蓝→绿的逃逸通道是消息单据（澄清单/缺货单），红→蓝是回退联结主义。"""
    tdir, root = task["task_dir"], task["root"]
    if task.get("executor") != "人":
        sys.exit("本方法执行者非人（不是绿任务）：蓝任务用 detect 检测可符号化")
    hist = "（无办结记录）"
    j = load_account(rec_dir(task), root)
    if j:
        try:
            hist = (f"最近办结 {j.get('开始', j.get('时间','?'))}，"
                    f"状态 {j.get('状态','?')}，备注 {j.get('备注','（无）')}")
        except Exception:
            hist = "（办结记录不可解析）"
    prompt = f"""你是渐进式信息化顾问。评估下面这个人工工位（绿任务）的可信息化程度：它的变换能否交给 LLM agent（蓝任务）完成。

判据（刀口）：看信息住在哪、行动在不在物理世界——
- 可信息化：纯信息作业——所需信息全部在输入弧文件里（或可由推理补齐），产物是树内文件，无需接触物理世界、无需联系树外的人。
- 不可信息化·物理钉：需要身体与世界接触（做饭、送货、取样、安装）。信息化无解，只能等机器人。
- 部分可信息化·通道钉：判断本身是信息作业，但关键信息住在树外人的脑子里（顾客要不要换菜、委托人要哪种口径）——缺的不是智能是通道。通道一旦信息化（IM/邮件/表单接口、MCP 工具），任务即可转蓝。
- 责任钉：裁决/签字/担责制度上必须留人的环节——技术可转蓝、制度不许，如实指出。

【任务指令】
{task['instruction'] or '（无自由文本指令）'}

【输入弧】
{_arc_lines(task, '输入') or '（无）'}

【产物弧】
{_arc_lines(task, '产物') or '（无）'}

【历届办结】{hist}

【弧现物摘录】
{_arc_material(task)}

要求（纯文本报告，不要写任何文件，不要改树内任何东西）：
1. 判决：可信息化 / 部分可信息化（写明缺哪条通道）/ 不可信息化（写明物理钉或责任钉）
2. 理由——对着刀口逐条说，指认变换里哪些环节是纯信息作业、哪些卡在物理/通道/责任
3. 转蓝路径（重心）：若可或部分可——任务指令怎么改写成 LLM 可执行的蓝任务指令、缺的通道用什么补（消息单据/表单/MCP）、残余必须留人的环节怎么切成更小的绿任务。若不可——写明钉子，并指出周边有无可剥离的信息作业（如登记、汇总、单据起草）可先转蓝。"""
    print(f"微任务：{relroot(tdir, root)}（检测可信息化 绿→蓝，引擎 `{engine}`）")
    print(f"  prompt {len(prompt)} 字符 …")
    trace = {} if stream else None
    report = call_engine(prompt, engine, timeout, stream, trace)
    print("\n═ 可信息化检测报告（只投影，不落树）═\n" + report.strip())


def cmd_compile(task, engine, timeout, stream, accept=False, discard=False):
    """编译=渐进式符号化：LLM 把任务变换写成确定性程序，进暂存隔离试跑，
    三件证据（程序全文/试跑输出/产物差异）落暂存。验收不自动：
    等待 --accept（落位转红并记编译账）或 --discard（弃暂存零副作用）。"""
    tdir, root = task["task_dir"], task["root"]
    stage = os.path.join(tdir, ".i3dna_compile")
    if discard:
        shutil.rmtree(stage, ignore_errors=True)
        print("已放弃：编译暂存清空，真树零副作用")
        return
    if accept:
        _compile_accept(task, stage, engine)
        return

    outs = [r for r in task["rows"] if r["kind"] == "产物" and r["path"]]
    if not outs:
        sys.exit("无可写产物弧，无从编译")
    shutil.rmtree(stage, ignore_errors=True)
    prog_dir = os.path.join(stage, EXEC_DIR)
    os.makedirs(prog_dir, exist_ok=True)
    golden = [r for r in outs if os.path.isfile(r["path"])]
    golden_note = ("当前产物槽里有上次联结点火的现物，可读它理解变换的目标形态"
                   "（程序必须从输入推导产物，禁止把现物内容硬编码进程序——现物只是理解材料）：\n"
                   + "\n".join(f"- {relroot(r['path'], root)}" for r in golden)
                   ) if golden else "（产物槽为空，无现物参照——试跑差异将缺对照证据）"
    prompt = f"""你被 I3DNA 执行引擎调用来【编译】一个微任务：把它的变换写成一段确定性程序（渐进式符号化）。注意：不是执行任务本身，是生成以后代替 LLM 执行它的程序。

【任务指令】
{task['instruction'] or '（变换语义由弧与参数表承载，见下）'}

【输入弧】程序运行时 cwd=包根，用这些相对路径读：
{_arc_lines(task, '输入') or '（无）'}

【产物弧】程序必须写出这些相对路径：
{_arc_lines(task, '产物')}

【现物参照】{golden_note}

【弧现物摘录】
{_arc_material(task, 2500)}

硬约束：
1. 程序必须确定性：不调 LLM、不访问网络、内容不依赖随机数或时钟。
2. 只写一个文件到这个绝对路径：{os.path.join(prog_dir, '主程序.py')}
   （Python 3，只用标准库；确需第三方库时在假设清单里声明。）
   程序顶部用注释写「假设清单」：所有未由输入决定的技术选择。
3. 程序运行时以包根为 cwd，按上述相对路径读输入、写产物；除声明的产物槽外不写任何文件。
4. 长程序分段写（单次工具调用有输出上限，先骨架后续写，最后通读校验）。
5. 写完后回复两行，不要多余内容：
   完成
   运行时输入：<逗号分隔的名称列表——程序运行时真正会读的输入；没列出的视为已烤进程序，它们一变就该重编译>"""
    print(f"微任务：{relroot(tdir, root)}（编译→符号程序，引擎 `{engine}`）")
    print(f"  prompt {len(prompt)} 字符，程序暂存 → {relroot(prog_dir, root)}")
    trace = {} if stream else None
    reply = call_engine(prompt, engine, timeout, stream, trace)
    prog = os.path.join(prog_dir, "主程序.py")
    if not os.path.isfile(prog) or os.path.getsize(prog) == 0:
        print("  ⚠ 编译空返（agent 未写出主程序）——疑似引擎瞬时抽风，"
              "自动重试一次…", flush=True)
        reply = call_engine(prompt, engine, timeout, stream, trace)
    if not os.path.isfile(prog) or os.path.getsize(prog) == 0:
        shutil.rmtree(stage, ignore_errors=True)
        sys.exit("编译失败（已重试一次，暂存已清，零副作用）——"
                 "建议换引擎车道再编译\n"
                 f"回复末 300 字：{reply[-300:]}")

    m = re.search(r"运行时输入[:：]\s*(.+)", reply)
    runtime_in = [s.strip() for s in re.split(r"[,，、]", m.group(1))
                  if s.strip()] if m else None

    # 隔离试跑：拷输入进暂存副本树，cwd=副本，真树零接触
    trial = os.path.join(stage, "试跑")
    for r in task["rows"]:
        if r["kind"] == "输入" and r["path"] and os.path.isfile(r["path"]):
            q = os.path.join(trial, relroot(r["path"], root))
            os.makedirs(os.path.dirname(q), exist_ok=True)
            shutil.copy2(r["path"], q)
    os.makedirs(trial, exist_ok=True)
    print("  试跑（隔离副本树）…")
    tr = subprocess.run([sys.executable, prog], cwd=trial, capture_output=True,
                        text=True, encoding="utf-8", errors="replace",
                        timeout=min(timeout, 300))
    trial_log = (f"$ python3 {relroot(prog, root)}  （cwd=隔离副本树）\n"
                 f"退出码 {tr.returncode}\n{tr.stdout}{tr.stderr}")
    open(os.path.join(stage, "试跑输出.txt"), "w", encoding="utf-8").write(trial_log)

    import difflib
    diffs, verdicts = [], {}
    for r in outs:
        rel = relroot(r["path"], root)
        tp = os.path.join(trial, rel)
        if not os.path.isfile(tp):
            verdicts[rel] = "试跑未产生该产物 ✗"
            diffs.append(f"═ {rel}\n试跑未产生该产物")
            continue
        if not os.path.isfile(r["path"]):
            verdicts[rel] = f"新产物（无现物可比，{os.path.getsize(tp)} 字节）"
            diffs.append(f"═ {rel}\n（产物槽为空，无 diff；试跑产物 {os.path.getsize(tp)} 字节）")
            continue
        if os.path.splitext(rel)[1].lower() in TEXT_EXT:
            a = open(r["path"], encoding="utf-8", errors="replace").read().splitlines()
            b = open(tp, encoding="utf-8", errors="replace").read().splitlines()
            d = list(difflib.unified_diff(a, b, f"现物/{rel}", f"试跑/{rel}", lineterm=""))
            verdicts[rel] = "与现物逐字节一致 ✓" if not d else f"与现物有差异（{len(d)} 行 diff）"
            diffs.append(f"═ {rel}\n" + ("（完全一致）" if not d else "\n".join(d[:400])))
        else:
            same = sha256(r["path"]) == sha256(tp)
            verdicts[rel] = "与现物 sha 一致 ✓" if same else "与现物 sha 不同（二进制）"
            diffs.append(f"═ {rel}\n二进制比对：{'一致' if same else '不同'}")
    open(os.path.join(stage, "试跑差异.txt"), "w", encoding="utf-8").write(
        "\n\n".join(diffs))
    meta = {"任务": relroot(tdir, root), "编译时间":
            datetime.now().isoformat(timespec="seconds"), "引擎": engine,
            "运行时输入": runtime_in, "试跑退出码": tr.returncode,
            "产物比对": verdicts,
            **({} if runtime_in is not None else
               {"备注": "agent 未自报运行时输入，默认全部按运行时算（编译过期仅看任务定义）"})}
    with open(os.path.join(stage, "编译元.json"), "w", encoding="utf-8") as f:
        json.dump(meta, f, ensure_ascii=False, indent=2)
    print(f"  证据三件套 → {relroot(stage, root)}/（主程序 / 试跑输出 / 试跑差异）")
    for rel, v in verdicts.items():
        print(f"    {rel}：{v}")
    print("  验收未定：暂存保留。落位=compile --accept，放弃=compile --discard")


def _compile_accept(task, stage, engine):
    tdir, root = task["task_dir"], task["root"]
    src = os.path.join(stage, EXEC_DIR)
    if not (os.path.isfile(os.path.join(src, "主程序.py"))
            or os.path.isfile(os.path.join(src, "主程序.sh"))):
        sys.exit("暂存里没有可落位的执行程序（先跑 compile）")
    meta = {}
    mp = os.path.join(stage, "编译元.json")
    if os.path.isfile(mp):
        meta = json.load(open(mp, encoding="utf-8"))
    runtime_in = meta.get("运行时输入")
    basis = {}
    for f in task_def_files(tdir):
        basis[relroot(f, root)] = sha256(f)
    for r in task["rows"]:
        if r["kind"] != "输入" or not r["path"] or not os.path.isfile(r["path"]):
            continue
        if runtime_in is not None and not any(
                k and (k in r["pname"] or k in r["desc"]) for k in runtime_in):
            basis[relroot(r["path"], root)] = sha256(r["path"])
    dst = os.path.join(tdir, EXEC_DIR)
    shutil.rmtree(dst, ignore_errors=True)
    shutil.move(src, dst)
    rec = {"任务": relroot(tdir, root), "编译时间": meta.get("编译时间", ""),
           "落位时间": datetime.now().isoformat(timespec="seconds"),
           "引擎": meta.get("引擎", engine), "运行时输入": runtime_in,
           "产物比对": meta.get("产物比对", {}), "依据sha": basis}
    with open(os.path.join(tdir, "__编译记录.json"), "w", encoding="utf-8") as f:
        json.dump(rec, f, ensure_ascii=False, indent=2)
    shutil.rmtree(stage, ignore_errors=True)
    print(f"已落位转红：{relroot(os.path.join(dst, '主程序.py'), root)}")
    print(f"  编译账（依据 sha {len(basis)} 条）→ __编译记录.json")
    _journal(root, f"编译 {relroot(tdir, root)}（联结→符号，--accept 落位）")


def cmd_revert(task):
    """回退联结主义：删执行程序即退蓝（判据是结构的）。删是安全的——
    journal 就是档案柜，编译落位那笔提交里程序全档都在，复辟走 git。"""
    tdir, root = task["task_dir"], task["root"]
    if not task["exec_entry"]:
        sys.exit("本任务不是符号任务（无执行程序），无从回退")
    shutil.rmtree(os.path.join(tdir, EXEC_DIR))
    rp = os.path.join(tdir, "__编译记录.json")
    if os.path.isfile(rp):
        os.remove(rp)
    print(f"已回退联结主义：{relroot(tdir, root)} 的执行程序已删（journal 有全档可复辟）")
    _journal(root, f"回退联结 {relroot(tdir, root)}（符号→联结，红字冲销）")


def dir_manifest(dpath):
    """目录盘点单（§8.12）：递归逐文件 sha；`__`/点前缀目录不入（账不进账，
    账不能是自己的输入）。记「收盘盘点」（点火后）——消费即删的收件箱
    不落永久漂移；对账侧同函数重算比对。"""
    out = {}
    for dp, dns, fns in os.walk(dpath):
        dns[:] = sorted(d for d in dns if not d.startswith((".", "__")))
        for fn in sorted(fns):
            p = os.path.join(dp, fn)
            out[os.path.relpath(p, dpath)] = sha256(p)
    return out


def _task_needs_fire(tdir, root, case=None):
    """推进判据 = 使能∧(产物缺失∨过期∨无记录)。Petri 出使能，Make 出新鲜度守卫。
    返回 (need, reason)。账外产物（有产物无记录）→ 需重算：没有记录就无法
    证明新鲜度，保守的正确选择是重算而非跳过。想保留旧产物请先 backfill（opt-in）。"""
    task = load_task(tdir, root, case=case)
    root = task["root"]               # 记号树可能把根上提（类目标→企业根），以 task 为准

    def _fire(reason):
        # 人工工位：LHS 满足也不自动点火——待人办是合法等待态，
        # 树在等账外直写+backfill 入账（引擎不替人干活，也不催人）
        if task.get("executor") == "人":
            return False, f"待人办（{reason}）"
        return True, reason
    if task.get("case"):
        kr = klass_rel(task["task_dir"], root)
        cdir = os.path.join(root, *case_rel(kr, task["case"]).split("/"))
        if not os.path.isdir(cdir):
            # 跨类任务({实例:别类})的工作目录在别家实例架——
            # 任一实例记号代入后真实存在的目录即为已实例化
            probes = []
            for r in task["rows"]:
                p = r.get("path")
                if p:
                    rel = os.path.relpath(p, root)
                    seg = rel.split(os.sep)
                    if len(seg) > 2 and seg[0] == "实例":
                        probes.append(os.path.join(root, seg[0], seg[1],
                                                   task["case"]))
            if not any(os.path.isdir(d) for d in probes):
                return False, f"未实例化：{task['case']} 在实例库无目录"
    # 持有单隔离(ACID 的 I):本任务读的实体实例若被其它案卷持有
    # (实例/<类>/<实体>/持有单__<别的案卷号>.md 在场),实体处于撕裂态
    # (案卷进行中:工作单位已改户口未改)——fail-closed,读了就是脏数据。
    # 持有方自己的任务不挡(案卷号匹配本 case 则是自家锁)。
    held = []
    for r in task["rows"]:
        if r["kind"] != "输入" or not r["path"]:
            continue
        seg = os.path.relpath(r["path"], root).split(os.sep)
        if len(seg) < 3 or seg[0] != "实例":
            continue
        ent_dir = os.path.join(root, seg[0], seg[1], seg[2])
        try:
            for f in os.listdir(ent_dir):
                if f.startswith("持有单__"):
                    holder = f[len("持有单__"):-3] if f.endswith(".md") else ""
                    if holder and holder != task.get("case"):
                        tag = f"{seg[2]}←{holder}"
                        if tag not in held:
                            held.append(tag)
                        break
        except OSError:
            continue
    if held:
        return False, ("脏数据隔离：" + "、".join(held[:3])
                       + ("…" if len(held) > 3 else "")
                       + "——实体正被其它案卷持有,等它结账")
    # 环让拓扑序失去意义（审查⇄生成互为上下游），新实例首轮必然有任务
    # 的输入还没出生；它们的火在后续轮次自然到（有界固定点迭代本来就为此）
    # 空夹门（形状定律 8-21·工单4）：目录弧声明 清空: 真 → 夹内（非 __/点
    # 前缀）零文件才使能，非空＝挂起——send-and-wait 的「等」是物理停不是
    # 会计停（闭缺陷18a：下游使能不再只看自身输入弧的在场）。与 blank_slot
    # 目录语义对偶：非空使能（吃单）↔ 清空使能（等回音核销完）。
    drained = []
    for r in task["rows"]:
        if r["kind"] == "输入" and r.get("drain") and r["path"]:
            n = len(_queue_tickets(r["path"]))   # 目录缺席＝空夹（0 张）
            if n:
                drained.append(f"{r['pname']}（{n} 张）")
    if drained:
        return False, ("空夹门挂起：" + "、".join(drained[:3])
                       + ("…" if len(drained) > 3 else "")
                       + "——等回音核销")
    # 三车道同门（缺陷20·8-21）推进侧：条件门挂起（直火拒火/办结拒，
    # 推进挂起——一副门两副牙，converge 不炸）；结账站盘点挂起（终点站
    # 等全夹清零才轮到关门那轮）。
    blocked = _cond_block(task)
    if blocked:
        return False, "条件门挂起：" + blocked[len("使能条件不满足："):]
    if task.get("closing"):
        bad = _closure_pending(task)
        if bad:
            return False, ("结账门挂起：" + "、".join(
                f"「{_d}」{_n} 张" for _d, _n in bad[:3])
                + "——终点站等全夹清零")
    absent = [r["pname"] for r in task["rows"]
              if r["kind"] == "输入" and r["path"] and not r.get("optional")
              and not r.get("drain")            # 空夹门行：空＝使能（对偶）
              and blank_slot(r["path"])]
    if absent:
        return False, "未使能：缺输入（缺席或空槽） " + ",".join(absent[:3])             + ("…" if len(absent) > 3 else "")
    outs = [r for r in task["rows"] if r["kind"] == "产物"
            and r["path"] and r["pname"] != "*"]
    # 目录弧产物（主题队列·形状定律 8-21）：在场＝队列非空（blank_slot
    # 目录语义：无文件＝空箱）；单被消费光＝产物缺失→再开一张（N 张单
    # N 轮火的产线侧对偶）。
    missing = [r["pname"] for r in outs
               if not (os.path.isfile(r["path"])
                       or (os.path.isdir(r["path"])
                           and not blank_slot(r["path"])))
               and not r.get("optional")]      # 可缺产物（返工单类）缺席≠缺失
    if missing:
        return _fire(f"产物缺失：{','.join(missing)}")
    if not _account_exists(rec_dir(task), root):
        return _fire("账外产物（有产物无点火记录）——无记录即无法验证新鲜度，重算")
    try:
        rec = load_account(rec_dir(task), root) or {}
    except (json.JSONDecodeError, OSError, ValueError):
        return False, "账不可解析，不自动重算"
    if rec.get("出处存疑"):
        return _fire("上轮出处存疑（点火期间输入被改），需重算")
    declared_all = {os.path.relpath(r["path"], root) for r in task["rows"]
                    if r["kind"] == "输入" and r["path"]}
    exec_prefix = os.path.relpath(os.path.join(tdir, EXEC_DIR), root) + os.sep
    stale = []
    violations = []
    dirs = []          # 盘点单漂移（§8.12）：目录收入/移出/改动
    for it in rec.get("输入清单", []):
        name = it.get("名称", "")
        # 弧改装后，记录里不再声明的旧输入失去约束力；执行程序条目除外
        # （符号任务刻意记录不声明——程序也是变换的输入）
        def_rel = (os.path.relpath(task["def_path"], root)
                   if task.get("def_path") else None)
        if (name not in declared_all and not name.startswith(exec_prefix)
                and name != def_rel):
            continue
        p = os.path.join(root, name)
        if it.get("目录"):
            # 盘点单对账（§8.12）：清单≠盘上＝目录已变（收入/移出/改动）
            if os.path.isdir(p) and dir_manifest(p) != it.get("清单"):
                dirs.append(os.path.basename(name))
            continue
        if it.get("字段区"):
            continue          # 仅字段区守门不催火；文档状态(需求等)是内容依赖,照常催
        if it.get("事实"):
            # 弧角色契约：事实件漂移 = 违规（篡改历史），不是过期（重算）。
            # 审计语义：报警呈人，不自动重算——重算会把篡改洗白成新事实
            if os.path.isfile(p) and "sha256" in it and sha256(p) != it["sha256"]:
                violations.append(os.path.basename(name))
            continue
        if os.path.isfile(p) and "sha256" in it and sha256(p) != it["sha256"]:
            stale.append(os.path.basename(name))
    if violations:
        return False, "⛔ 事实件被篡改：" + ",".join(violations) \
            + "（审计红线：不重算，呈人裁决；更正走新单据冲销）"
    if stale:
        return _fire("输入已变：" + ",".join(stale))
    if dirs:
        return _fire("目录已变：" + ",".join(dirs))
    # 可缺弧未在场不计入声明集：返工单被质检收回后不再催火（环在此闭合）
    declared = {os.path.relpath(r["path"], root) for r in task["rows"]
                if r["kind"] == "输入" and r["path"]
                and not is_field_area(task, r)     # 仅字段区不入催火比对
                and (not r.get("optional") or os.path.isfile(r["path"]))}
    extra = declared - {it.get("名称") for it in rec.get("输入清单", [])}
    if extra:
        opt = {os.path.relpath(r["path"], root) for r in task["rows"]
               if r["kind"] == "输入" and r["path"] and r.get("optional")}
        names = ",".join(sorted(os.path.basename(x) for x in extra))
        # 同一条检查的两种人话：可缺消息在场未消费=有活找你；必需弧新声明=法改了
        if all(x in opt for x in extra):
            return _fire(f"在途消息待消费：{names}")
        return _fire(f"声明输入未入账（任务改装后需重算）：{names}")
    return False, "新鲜"


def cmd_converge(pkg_root, a):
    """按拓扑序推进全包：该点的点、新鲜的跳、账外的留言。一次决策收敛整网。"""
    root = os.path.abspath(pkg_root)
    tdirs = sorted(find_tasks(root))
    case = getattr(a, 'case', None)

    def _needs(t):
        """帧容错判据：本案卷帧装不下的任务（外类内容记号缺载荷、无案卷
        marked 任务）＝与之无关，按未实例化聚合跳过——不炸整个推进
        （8-19 猎证：女娲内容记号任务曾让全根扇出推进全线失败）。"""
        try:
            return _task_needs_fire(t, root, getattr(a, 'case', None))
        except SystemExit:
            return False, "未实例化（本帧装不下：内容记号缺载荷或缺案卷）"

    # 任务级 DAG → 拓扑层（装不下的任务无弧=孤立点，由 _needs 聚合跳过）
    prod, cons = {}, {}
    for t in tdirs:
        try:
            task = load_task(t, root, case=case)
        except SystemExit:
            continue
        for r in task["rows"]:
            if not r["path"]:
                continue
            (prod if r["kind"] == "产物" else cons).setdefault(t, set()).add(r["path"])
    succ = {t: {u for u in tdirs if u != t
                and prod.get(t, set()) & cons.get(u, set())} for t in tdirs}
    level = {t: 0 for t in tdirs}
    for _ in range(len(tdirs) + 1):
        for x, ys in succ.items():
            for y in ys:
                level[y] = max(level[y], level[x] + 1)
    order = sorted(tdirs, key=lambda t: (level[t], t))

    # 有界固定点迭代：环上的账（守卫弧返工单）一轮消化不完就再对一轮；
    # 到 --max-rounds 仍有火 = 未收敛，停止并以退出码 3 报告状态
    # （停机从"必然"降为"有界"；归因与处置在治理层，不在本引擎）
    max_rounds = 1 if a.plan else max(1, getattr(a, "max_rounds", 3))
    total_fired, skipped, round_no = [], [], 0
    while True:
        round_no += 1
        if round_no > 1:
            print(f"\n═ 第 {round_no} 轮（环上有新账）═")
        fired, skipped = [], []
        alien = {}                    # 未实例化的类：与本实例无关，聚合不唱名
        for t in order:
            need, reason = _needs(t)
            rel = os.path.relpath(t, root)
            if not need:
                if reason.startswith("未实例化"):
                    k = rel.split(os.sep + "方法" + os.sep)[0]
                    alien[k] = alien.get(k, 0) + 1
                else:
                    print(f"· 跳过 {rel}（{reason}）")
                skipped.append(rel)
                continue
            if a.plan:
                print(f"▶ 将点火 {rel}（{reason}）")
                fired.append(rel)
                continue
            print(f"▶ 点火 {rel}（{reason}）")
            task = load_task(t, root, case=case, executor=a.executor)
            cmd_run(task, a.sandbox and os.path.abspath(a.sandbox), a.engine,
                    a.timeout, a.io, a.run_id, a.stream,
                    getattr(a, "intent", "") or "")
            fired.append(rel)
        if alien:
            print("· 跳过 " + "、".join(f"{k}（{n} 站）" for k, n in
                                        sorted(alien.items()))
                  + "——本实例不在其实例库，与之无关")
        total_fired += fired
        if a.plan or not fired or round_no >= max_rounds:
            break
    verb = "计划点火" if a.plan else "已点火"
    print(f"\n推进完成：{verb} {len(total_fired)} 站（{round_no} 轮），"
          f"末轮跳过 {len(skipped)} 站"
          + ("（下游过期由级联在各自轮次现算）" if a.plan else ""))
    pending = [(os.path.relpath(t, root), r) for t in order
               for _, r in [_needs(t)]
               if r.startswith("待人办")]
    if pending:
        # 守恒律语义:这是边界辐射清单——待人办的消息是飞向体系外
        # (人=外部世界)的玻色子,静止在场合法;真正的悬账检查在案卷
        # 结账时(lint_case_closure),不在全局不动点(开放系统无全树静止)
        print(f"🧑 待人办 {len(pending)} 站（边界辐射:人工完成后 "
              f"backfill 办结入账——案卷结账前必须清零）：")
        for rel, r in pending:
            print(f"   ☐ {rel}（{r}）")
    if not a.plan and round_no >= max_rounds:
        leftover = [os.path.relpath(t, root) for t in order
                    if _needs(t)[0]]
        if leftover:
            print(f"⚠ 未收敛：轮数到限（{max_rounds}）仍有待点火："
                  f"{'，'.join(leftover)}")
            sys.exit(3)


def _account_candidates(root, tdir):
    """(实例号, 账目录) 候选——账是 M0:实例模式在 实例/<类>/<k>/__账/,
    单案在任务目录。引擎侧自持零依赖(与 core.task_accounts 同判定)。"""
    out = []
    md = os.path.join(tdir, "任务.md")
    if not os.path.isfile(md):
        mds = sorted(glob.glob(os.path.join(tdir, "__*任务定义*.md")))
        md = mds[0] if mds else ""
    body = open(md, encoding="utf-8", errors="replace").read() if md else ""
    marked = bool(md) and ("{实例}" in body or "{案卷号}" in body)
    if marked:
        kr = klass_rel(tdir, root)
        shelf = os.path.join(root, "实例", os.path.basename(kr)) if kr \
            else os.path.join(root, "实例")
        if not os.path.isdir(shelf):
            return out
        for k in sorted(os.listdir(shelf)):
            if k.startswith(".") or not os.path.isdir(os.path.join(shelf, k)):
                continue
            try:
                t = load_task(tdir, root, case=k)
            except SystemExit:
                continue
            out.append((k, rec_dir(t)))
    else:
        try:
            t = load_task(tdir, root)
        except SystemExit:
            return out
        out.append((None, rec_dir(t)))
    return out


def coverage_report(root):
    """MBT 覆盖报告(P3):验收挂树上。GraphWalker 停机条件映射——
    变迁覆盖=每个微任务点火≥1(有账);边覆盖=DAG 边(生产→消费经产物路径)
    两端的任务都有账≥1。缺口修法=推进(converge)。"""
    tdirs = sorted(find_tasks(root))
    fired, unfired = [], []
    for t in tdirs:
        hit = False
        for _c, rd in _account_candidates(root, t):
            try:
                if load_account(rd, root):
                    hit = True
                    break
            except (ValueError, OSError):
                continue
        (fired if hit else unfired).append(t)

    def _rows(t):
        cands = _account_candidates(root, t)
        if not cands:                      # 标记任务无实例:类级也读不出,跳过
            return []
        try:
            task = load_task(t, root, case=cands[0][0])
        except SystemExit:
            return []
        return task["rows"]

    prod_of, cons_of = {}, {}
    for t in tdirs:
        for r in _rows(t):
            p = r.get("path")
            if not p or r.get("pname") == "*":
                continue
            (prod_of if r["kind"] == "产物" else cons_of) \
                .setdefault(p, set()).add(t)
    edges, walked = [], []
    for p, ps in prod_of.items():
        for v in cons_of.get(p, ()):
            for u in ps:
                if u == v:
                    continue
                e = (u, v, p)
                edges.append(e)
                if u in fired and v in fired:
                    walked.append(e)
    return {"任务": tdirs, "已点火": fired, "未点火": unfired,
            "变迁覆盖": (len(fired), len(tdirs)),
            "边": edges, "已走过": walked,
            "边覆盖": (len(walked), len(edges))}


def cmd_coverage(pkg_root):
    """覆盖报告→终端:验收挂树上的读法。退出码 0=全覆盖,1=有缺口(常态)。"""
    rep = coverage_report(os.path.abspath(pkg_root))
    nf, nt = rep["变迁覆盖"]
    kw, ke = rep["边覆盖"]
    print(f"变迁覆盖 {nf}/{nt} · 边覆盖 {kw}/{ke}")
    for t in rep["未点火"]:
        print(f"  ◻ 未点火 {os.path.relpath(t, pkg_root)}")
    for u, v, p in rep["边"]:
        if (u, v, p) not in rep["已走过"]:
            print(f"  ◻ 未走过 {os.path.basename(u)}→{os.path.basename(v)}"
                  f"（{os.path.relpath(p, pkg_root)}）")
    sys.exit(0 if (nf == nt and kw == ke) else 1)


def main():
    ap = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("cmd", choices=["list", "preflight", "run", "backfill",
                                    "draft", "converge", "detect", "compile",
                                    "revert", "inform", "coverage", "login"])
    ap.add_argument("target", nargs="?", default=".")
    ap.add_argument("--root", help="强制指定树根（默认锚探针自校准）")
    ap.add_argument("--sandbox")
    ap.add_argument("--engine",
                    default=os.environ.get("I3DNA_LLM_CMD", DEFAULT_ENGINE),
                    help="引擎命令；`acp:<服务端命令>` 走 ACP 车道"
                         "（JSON-RPC over stdio，失败带 stopReason 死因），"
                         "如 acp:omp --thinking high acp")
    ap.add_argument("--io", choices=["write", "stdout"], default="write",
                    help="write=引擎agent直写产物路径(默认，OMP/claude 等带工具引擎)；"
                         "stdout=产物回传本引擎落盘(裸 LLM API 用)")
    ap.add_argument("--run-id", default=os.environ.get("I3DNA_RUN_ID", ""),
                    help="批次标识（意图分组）：同一条链的多次点火共用一个 ID，"
                         "落进各自 __结果.json 的「批次标识」；一条链跑前 "
                         "export I3DNA_RUN_ID=run-$(date +%%Y%%m%%d-%%H%%M%%S)。"
                         "实际血缘仍由 sha256 内容寻址承载，两者答不同问题")
    ap.add_argument("--output", action="append", default=[],
                    help="显式产物：名称@锚目录（任务定义/纯索引族用，可重复）")
    ap.add_argument("--timeout", type=int, default=1800)
    ap.add_argument("--note", default="")
    ap.add_argument("--status", default="",
                    help="login 专用：登录事件状态（登录成功/密码错误）")
    ap.add_argument("--stream", action="store_true",
                    help="流式输出引擎过程（OMP 自动加 --mode json，"
                         "💭思考/🔧工具/↩结果/🗣回复 逐行泵出；此模式下不做超时）")
    ap.add_argument("--plan", action="store_true",
                    help="converge 专用：只列推进计划不点火")
    ap.add_argument("--case", default=None,
                    help="实例号（M0 实例）：弧声明里的 {实例} 记号代入为 实例/<case>；"
                         "同一份 M1 定义可多实例并发，写集按构造不相交")
    ap.add_argument("--executor", default=None,
                    help="执行者覆盖（点火三元组第三维：执行者×方法×对象）。"
                         "格式 实例/部门/<部门号>（如 实例/部门/D01）；"
                         "缺省用任务.md 声明值，账记「执行者」字段")
    ap.add_argument("--intent", default="",
                    help="意图入账（101号 右键对话：话语原文）——落进"
                         " __结果.json「意图」字段；审计留痕字段，"
                         "不参与 sha 对账（lint 双读兼容：旧账无此键不比对）")
    ap.add_argument("--max-rounds", type=int, default=3,
                    help="converge 专用：固定点迭代轮数上限（守卫弧成环时"
                         "的停机保险，默认 3；到限仍有待点火则退出码 3 报告未收敛）")
    ap.add_argument("--accept", action="store_true",
                    help="compile 专用：接受暂存程序——落位转红＋记编译账")
    ap.add_argument("--discard", action="store_true",
                    help="compile 专用：弃编译暂存，零副作用")
    a = ap.parse_args()

    if a.cmd == "converge":
        cmd_converge(a.target, a)
        return

    if a.cmd == "coverage":
        cmd_coverage(a.target)
        return

    if a.cmd == "login":
        cmd_login(a.root or ".", a.target, a.status or "登录成功")
        return

    if a.cmd == "list":
        pkg = os.path.abspath(a.target)
        for d, kind in find_tasks(pkg).items():
            task = load_task(d, a.root, case=a.case)
            n_in = sum(1 for r in task["rows"] if r["kind"] == "输入")
            n_out = sum(1 for r in task["rows"] if r["kind"] == "产物")
            print(f"[{kind}] {os.path.relpath(d, pkg)}  （输入 {n_in} / 产物 {n_out}）")
        return

    task = load_task(a.target, a.root, a.output, case=a.case,
                     executor=a.executor,
                     tolerate_content_marks=(a.cmd == "draft"))
    if a.cmd == "preflight":
        print(f"微任务：{relroot(task['task_dir'], task['root'])}"
              f"（{task['kind']}族，树根 {task['root']}）")
        sys.exit(0 if preflight(task) else 1)
    if a.cmd == "run":
        cmd_run(task, a.sandbox and os.path.abspath(a.sandbox), a.engine, a.timeout,
                a.io, a.run_id, a.stream, a.intent)
    if a.cmd == "backfill":
        cmd_backfill(task, a.note, a.intent)
    if a.cmd == "draft":
        try:
            草稿 = json.load(sys.stdin)
        except ValueError as e:
            sys.exit(f"draft 载荷不是 JSON（stdin）：{e}")
        cmd_draft(task, 草稿)
    if a.cmd == "detect":
        cmd_detect(task, a.engine, a.timeout, a.stream)
    if a.cmd == "inform":
        cmd_inform(task, a.engine, a.timeout, a.stream)
    if a.cmd == "compile":
        cmd_compile(task, a.engine, a.timeout, a.stream, a.accept, a.discard)
    if a.cmd == "revert":
        cmd_revert(task)


if __name__ == "__main__":
    main()
