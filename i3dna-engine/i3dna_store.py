#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""i3dna_store — I3DNA 领域模型的持久化层（底物适配器）。

分层律：领域模型（弧/账/任务定义）是纯数据形状，**怎么落盘归本层**——
markdown 底物（obsidian 族）与勘察院 Excel 底物（参数表/索引 xlsx 族）
是同一份领域对象的两种编码。引擎/工具只见领域对象，不见 xlsx 列头与
yaml frontmatter。

两个正交轴（不要混）：
  1. 任务定义底物——按任务目录**自动识别**（任务.md → md；参数表/索引
     xlsx → xlsx）。一个包可混住两族（现状三代包格式通用）。
  2. 账底物——包级配置，默认 json（与 lint/explorer 现状全兼容）；
     勘察院全 Excel 包可切 xlsx（树根 __底物.yaml 的「账:」或环境变量
     I3DNA_ACCOUNT_STORE=xlsx）。读永远双回退（json→xlsx），写按配置，
     迁移期两底物并存不炸账。

本文件零 engine 依赖（engine import 本层，不反向）。
"""
import glob
import json
import os
import re
import sys

openpyxl = None          # 懒加载：md 树零 xlsx 依赖


def _need_openpyxl():
    global openpyxl
    if openpyxl is None:
        try:
            import openpyxl as _o
        except ImportError:
            sys.exit("需要 openpyxl：pip install openpyxl（仅 xlsx 底物需要）")
        openpyxl = _o
    return openpyxl


# ── 表格工具（xlsx 底物共用；engine 从这里 import）──────────

def norm(v):
    return "" if v is None else str(v).strip()


def is_coord_row(cells):
    vals = [c for c in cells if c != ""]
    if len(vals) < 3:
        return False
    if vals[0] == "*":
        vals = vals[1:]
    return all(re.fullmatch(r"\d+", v) for v in vals) and \
        [int(v) for v in vals] == list(range(len(vals)))


def data_rows(ws):
    for row in ws.iter_rows(values_only=True):
        cells = [norm(c) for c in row]
        first = cells[0] if cells else ""
        if re.fullmatch(r"\d+", first) and not is_coord_row(cells):
            yield cells


def headers_of(ws):
    for row in ws.iter_rows(values_only=True):
        return [norm(c) for c in row]
    return []


# ── 领域对象形状（dict 即领域对象——不引入类开销；键集是契约）──

ARC_KEYS = ("desc", "ptype", "pdir", "pver", "pname", "kind")
ACCOUNT_META_KEYS = ("状态", "批次标识", "平台钉", "IO模式", "引擎", "开始",
                     "结束", "回填时间", "备注", "出处存疑", "意图")


def new_arc(kind, pdir="", pname="", desc="", ptype="", pver=""):
    """弧的构造器——唯一合法入口，防键名漂移。"""
    return {"desc": desc, "ptype": ptype, "pdir": pdir, "pver": pver,
            "pname": pname, "kind": kind}


# ── 任务定义底物：识别与装载 ────────────────────────────────

def md_decl_path(task_dir):
    return os.path.join(task_dir, "任务.md")


def recipe_paths(task_dir):
    """勘察院参数表（8.5 模板族）：__*大模型智能体*版本.xlsx。"""
    return sorted(glob.glob(os.path.join(task_dir,
                                          "__*大模型智能体*版本.xlsx")))


def taskdef_md_paths(task_dir):
    """任务定义 md（7.26 族）：__*任务定义*.md。"""
    return sorted(glob.glob(os.path.join(task_dir, "__*任务定义*.md")))


def index_paths(task_dir):
    return sorted(glob.glob(os.path.join(task_dir, "__*索引文件*.xlsx")))


def classify_kind(ptype, pdir, target_axis):
    """客户类型串自带角色标记优先（7.17/7.27 族），无标记退回目标轴位置推断。

    角色标记只认类型串**末段**（模块专名会撞角色词，8-06 实测回归）。"""
    tail = ptype.rstrip("】").split("_")[-1]
    if "输入" in tail:
        return "输入"
    if "成果" in tail or "生成" in tail:
        return "产物"
    segs = pdir.replace("\\", "/").split("/")
    return "产物" if target_axis in segs else "输入"


def frontmatter_rows(md_path):
    """md 底物（obsidian 族）：`任务.md` 的 frontmatter 声明弧，正文=任务指令。"""
    import yaml
    text = open(md_path, encoding="utf-8").read()
    m = re.match(r"^---\s*\n(.*?)\n---\s*\n?(.*)$", text, re.S)
    if not m:
        sys.exit(f"{md_path} 缺 frontmatter（--- yaml ---）")
    fm = yaml.safe_load(m.group(1)) or {}
    rows = []
    for kind_key, kind in (("输入", "输入"), ("产物", "产物")):
        for item in fm.get(kind_key, []) or []:
            optional, cond, role, retract = False, None, None, False
            drain = False
            if isinstance(item, str):
                rel, desc = item, os.path.basename(item)
            else:
                rel = item.get("路径", "")
                desc = item.get("描述") or os.path.basename(rel)
                optional = bool(item.get("可缺"))
                cond = item.get("使能条件")
                role = item.get("角色")          # 弧角色契约：意图(缺省)|事实
                retract = bool(item.get("回收"))
                drain = str(item.get("清空") or "").strip() \
                    in ("真", "True", "true")    # 空夹门（工单4）：目录弧
                                                # 清空才使能——绿语义的等
            rows.append({"desc": desc, "ptype": "", "pver": "*",
                         "pdir": os.path.dirname(rel) or "*",
                         "pname": os.path.basename(rel), "kind": kind,
                         "optional": optional, "cond": cond,
                         "role": role, "retract": retract, "drain": drain})
    return rows, (m.group(2) or "").strip(), fm

def is_fact_arc(row):
    """弧角色契约：事实件弧（角色=事实）。事实件被改=违规（报警），
    不是过期（重算）——审计语义 vs 编译语义的分界。"""
    return (row or {}).get("role") == "事实"


def class_kind(klass_dir):
    """类.md 元声明 → 范畴（实体|过程）。无声明=过程（缺省：方法容器）。"""
    md = os.path.join(klass_dir, "类.md")
    if not os.path.isfile(md):
        return None
    try:
        import yaml
        m = re.match(r"^---\s*\n(.*?)\n---",
                     open(md, encoding="utf-8").read(), re.S)
        if not m:
            return None
        fm = yaml.safe_load(m.group(1)) or {}
        k = str(fm.get("范畴", "")).strip()
        return k if k in ("实体", "过程") else None
    except Exception:
        return None


def param_rows(recipe, task_dir):
    """勘察院参数表 → 弧列表（列头契约：【参数文件目录/名称/类型/版本/描述】）。"""
    target_axis = os.path.basename(os.path.dirname(os.path.abspath(task_dir)))
    wb = _need_openpyxl().load_workbook(recipe, data_only=True)
    rows = []
    for ws in wb.worksheets:
        heads = headers_of(ws)
        if "【参数文件类型】" not in heads:
            continue
        idx = {h: i for i, h in enumerate(heads)}
        for cells in data_rows(ws):
            def col(h):
                i = idx.get(h)
                return cells[i] if i is not None and i < len(cells) else ""
            rows.append(new_arc(
                classify_kind(col("【参数文件类型】"), col("【参数文件目录】"),
                              target_axis),
                pdir=col("【参数文件目录】"), pname=col("【参数文件名称】"),
                desc=col("【描述】"), ptype=col("【参数文件类型】"),
                pver=col("【参数文件版本】")))
    wb.close()
    return rows


def index_rows_from_file(path):
    """索引文件外联行 → 输入弧（近端绑定行的外联列=远端目录）。"""
    rows = []
    wb = _need_openpyxl().load_workbook(path, read_only=True, data_only=True)
    for ws in wb.worksheets:
        if "【目录-文件名称】" not in "".join(headers_of(ws)):
            continue
        for cells in data_rows(ws):
            cells += [""] * (7 - len(cells))
            _, desc, name, _, _, _, extref = cells[:7]
            if name and name != "*" and extref not in ("", "*"):
                rows.append(new_arc("输入", pdir=extref, pname=name, desc=desc))
    wb.close()
    return rows


def index_input_rows(task_dir):
    rows = []
    for path in index_paths(task_dir):
        rows += index_rows_from_file(path)
    return rows


class MdDefStore:
    """md 任务定义底物（任务.md frontmatter + 正文指令）。"""
    kind = "frontmatter"

    def __init__(self, task_dir):
        self.task_dir = os.path.abspath(task_dir)
        self.path = md_decl_path(self.task_dir)

    def load_arcs(self):
        rows, instruction, _fm = frontmatter_rows(self.path)
        return instruction, rows

    def save_arcs(self, rows, instruction, extra_fm=None):
        """写回任务.md：frontmatter 由弧列表重生成（yaml 列表，路径=dir/name），
        正文=指令原文。原子替换。extra_fm 透传保留 frontmatter 其余键。"""
        import yaml
        fm = dict(extra_fm or {})
        for kind in ("输入", "产物"):
            items = []
            for r in rows:
                if r["kind"] != kind:
                    continue
                rel = "/".join(x for x in (r["pdir"], r["pname"]) if x)
                d = {"路径": rel, "描述": r["desc"]}
                if r.get("optional"):
                    d["可缺"] = True
                if r.get("cond"):
                    d["使能条件"] = r["cond"]
                items.append(d)
            if items:
                fm[kind] = items
        head = yaml.safe_dump(fm, allow_unicode=True, sort_keys=False)
        tmp = self.path + ".tmp"
        with open(tmp, "w", encoding="utf-8") as f:
            f.write(f"---\n{head}---\n\n{instruction}\n")
        os.replace(tmp, self.path)
        return self.path


class XlsxDefStore:
    """勘察院 Excel 任务定义底物（参数表优先；索引外联行兜底）。"""
    kind = "参数表"

    def __init__(self, task_dir):
        self.task_dir = os.path.abspath(task_dir)

    def load_arcs(self):
        recipes = recipe_paths(self.task_dir)
        if recipes:
            return "", param_rows(recipes[0], self.task_dir)
        return "", index_input_rows(self.task_dir)   # 纯索引族

    def save_arcs(self, rows, instruction=""):
        """往参数表追加弧行（机器代笔）：勘察院列头契约，序号自增。
        只支持追加不支持改写——参数表是客户签字件，改写必须人手工。"""
        recipes = recipe_paths(self.task_dir)
        if not recipes:
            raise NotImplementedError("纯索引族无参数表可写；弧登记改用 md 底物")
        path = recipes[0]
        wb = _need_openpyxl().load_workbook(path)
        ws = wb.worksheets[0]
        heads = headers_of(ws)
        col = {h: i for i, h in enumerate(heads)}
        seqs = [int(c[0]) for c in data_rows(ws)]
        seq = (max(seqs) + 1) if seqs else 1
        for r in rows:
            ws.append([""] * len(heads))
            row_idx = ws.max_row
            ptype = r["ptype"] or ("输入参数文件】" if r["kind"] == "输入"
                                   else "成果模型文件】")
            for key, val in (("【参数文件目录】", r["pdir"] or "*"),
                             ("【参数文件名称】", r["pname"]),
                             ("【参数文件类型】", ptype),
                             ("【参数文件版本】", r.get("pver", "")),
                             ("【描述】", r["desc"])):
                if key in col:
                    ws.cell(row=row_idx, column=col[key] + 1, value=val)
            seq += 1
        wb.save(path)
        wb.close()
        return path


class TaskDefMdStore:
    """任务定义 md 底物（7.26 族：__*任务定义*.md 正文=指令，索引行=输入）。"""
    kind = "任务定义"

    def __init__(self, task_dir):
        self.task_dir = os.path.abspath(task_dir)
        self.paths = taskdef_md_paths(self.task_dir)

    def load_arcs(self):
        instruction = open(self.paths[0], encoding="utf-8",
                           errors="replace").read()
        return instruction, index_input_rows(self.task_dir)

    def save_arcs(self, rows, instruction):
        tmp = self.paths[0] + ".tmp"
        with open(tmp, "w", encoding="utf-8") as f:
            f.write(instruction)
        os.replace(tmp, self.paths[0])
        return self.paths[0]


def open_def_store(task_dir):
    """任务目录 → 定义底物适配器（族自动识别；与 load_task 原判定同序）。"""
    task_dir = os.path.abspath(task_dir)
    if os.path.isfile(md_decl_path(task_dir)):
        return MdDefStore(task_dir)
    if recipe_paths(task_dir):
        return XlsxDefStore(task_dir)
    if taskdef_md_paths(task_dir):
        return TaskDefMdStore(task_dir)
    return None          # 纯索引族由调用方按轴判定（沿用 engine 原逻辑）


# ── 账底物：json（默认）/ xlsx（勘察院全 Excel 包）─────────

ACCOUNT_JSON = "__结果.json"
ACCOUNT_XLSX = "__账.xlsx"

_META_ROWS = [("状态", "状态"), ("批次标识", "批次标识"),
              ("平台钉", "平台钉"), ("IO模式", "IO模式"), ("引擎", "引擎"),
              ("开始", "开始"), ("结束", "结束"), ("回填时间", "回填时间"),
              ("备注", "备注"), ("出处存疑", "出处存疑"), ("意图", "意图")]


def account_format(root):
    """包级账底物：树根 __底物.yaml 的「账:」> 环境变量 > 默认 json。"""
    env = os.environ.get("I3DNA_ACCOUNT_STORE", "").strip().lower()
    if env in ("json", "xlsx"):
        return env
    cfg = os.path.join(root, "__底物.yaml")
    if os.path.isfile(cfg):
        try:
            import yaml
            v = (yaml.safe_load(open(cfg, encoding="utf-8")) or {}) \
                .get("账", "")
            if str(v).strip().lower() in ("json", "xlsx"):
                return str(v).strip().lower()
        except Exception:
            pass
    return "json"


class JsonAccountStore:
    """账的 json 底物（现状）：tmp+rename 原子写。"""
    fmt = "json"

    def path(self, rec_dir):
        return os.path.join(rec_dir, ACCOUNT_JSON)

    def load(self, rec_dir):
        p = self.path(rec_dir)
        if not os.path.isfile(p):
            return None
        return json.load(open(p, encoding="utf-8"))

    def save(self, rec_dir, payload):
        os.makedirs(rec_dir, exist_ok=True)
        out = self.path(rec_dir)
        # tmp 带 pid：并发双写不共享暂存名（8-19 猎证：固定名下对方抢先
        # rename 会让本方 FileNotFoundError 崩栈——单文件原子性照旧，末写
        # 者胜，账房哲学不手搓锁）
        tmp = f"{out}.tmp.{os.getpid()}"
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
        os.replace(tmp, out)
        return out


class XlsxAccountStore:
    """账的勘察院 Excel 底物：__账.xlsx，三 sheet（概要/输入清单/产物清单）。
    读双回退（json 优先）——迁移期旧 json 账仍可读，写按 xlsx。"""
    fmt = "xlsx"

    def path(self, rec_dir):
        return os.path.join(rec_dir, ACCOUNT_XLSX)

    def _rows_to_payload(self, wb, sheet, keys):
        ws = wb[sheet] if sheet in wb.sheetnames else None
        if ws is None:
            return []
        heads = headers_of(ws)
        out = []
        for cells in data_rows(ws):
            d = {}
            for i, h in enumerate(heads):
                if h in keys and i < len(cells):
                    d[h] = _coerce(cells[i])
            if d:
                out.append(d)
        return out

    def load(self, rec_dir):
        jp = os.path.join(rec_dir, ACCOUNT_JSON)
        if os.path.isfile(jp):                  # 回退：迁移期 json 优先可读
            return json.load(open(jp, encoding="utf-8"))
        p = self.path(rec_dir)
        if not os.path.isfile(p):
            return None
        wb = _need_openpyxl().load_workbook(p, data_only=True)
        payload = {}
        if "概要" in wb.sheetnames:            # 概要纵表：行=[序号, 键, 值]
            for cells in data_rows(wb["概要"]):
                if len(cells) >= 3 and cells[1]:
                    payload[cells[1]] = _coerce(cells[2])
        for key in ("输入清单", "产物清单"):
            payload[key] = self._rows_to_payload(
                wb, key, {"名称", "字节", "sha256", "验证动作"})
        for extra in ("缺失输入或产物", "验证动作"):
            if extra in wb.sheetnames:
                heads = headers_of(wb[extra])
                payload[extra] = [
                    {h: _coerce(c[i]) for i, h in enumerate(heads)
                     if h and h != "序号" and i < len(c)}
                    for c in data_rows(wb[extra])]
                if not payload[extra]:
                    del payload[extra]
        wb.close()
        return payload

    def save(self, rec_dir, payload):
        os.makedirs(rec_dir, exist_ok=True)
        out = self.path(rec_dir)
        wb = _need_openpyxl().Workbook()
        wb.remove(wb.active)                    # 删默认空 Sheet
        ws = wb.create_sheet("概要")            # 纵表：行=[序号, 键, 值]
        for i, (key, _) in enumerate(_META_ROWS, start=1):
            v = payload.get(key)
            if v is not None and v != "":
                ws.append([i, key, v])
        for key in ("输入清单", "产物清单", "验证动作", "缺失输入或产物"):
            items = payload.get(key) or []
            if not isinstance(items, list):
                continue
            s2 = wb.create_sheet(key)
            allkeys = []
            for it in items:
                for k in it:
                    if k not in allkeys:
                        allkeys.append(k)
            s2.append(["序号"] + (allkeys or ["名称", "字节", "sha256"]))
            seq = 1
            for it in items:
                s2.append([seq] + [it.get(k, "") for k in allkeys])
                seq += 1
        tmp = out + ".tmp"
        wb.save(tmp)
        wb.close()
        os.replace(tmp, out)
        return out


def _coerce(v):
    """xlsx 单元格标量 → payload 标量（数字串保数字，其余原样）。"""
    if isinstance(v, (int, float)):
        return v
    s = str(v)
    if re.fullmatch(r"\d+", s):
        return int(s)
    return s


def open_account_store(root):
    """包级账底物适配器。"""
    return JsonAccountStore() if account_format(root) == "json" \
        else XlsxAccountStore()


def account_exists(rec_dir, root=None):
    """账在否（按写底物；读兼容双底物）。"""
    if os.path.isfile(os.path.join(rec_dir, ACCOUNT_JSON)):
        return True
    if root is not None and account_format(root) == "xlsx":
        return os.path.isfile(os.path.join(rec_dir, ACCOUNT_XLSX))
    return False


def load_account(rec_dir, root=None):
    """统一读账：按包底物，xlsx 底物自带 json 回退。root 缺省按 json。"""
    return open_account_store(root or rec_dir).load(rec_dir)


def save_account(rec_dir, payload, root=None):
    """统一写账：原子落位，返回写入路径。"""
    return open_account_store(root or rec_dir).save(rec_dir, payload)
